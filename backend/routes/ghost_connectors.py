"""
GoFarther AI — Universal App Connector System

Routes:
  GET  /ghost/connectors              — list all available apps + connection status
  POST /ghost/connectors/{app_id}/connect    — save API key / credentials
  DELETE /ghost/connectors/{app_id}/disconnect — remove credentials
  POST /ghost/connectors/{app_id}/action     — execute an action on a connected app
"""

from __future__ import annotations

import logging
import json
import os
import uuid
from datetime import datetime, timezone
from typing import Any, Optional

import httpx
from fastapi import APIRouter, HTTPException, Header, Depends, UploadFile, File, Form
from fastapi.responses import Response
from pydantic import BaseModel
from sqlalchemy import Column, String, Text, DateTime, select, and_
from sqlalchemy.dialects.postgresql import UUID
from sqlalchemy.ext.asyncio import AsyncSession

from db import get_db, Base

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/ghost/connectors", tags=["ghost-connectors"])


# ── Auth helper ──────────────────────────────────────────────────────────────

def _verify_auth(authorization: str) -> dict:
    from routes.ghost_auth import verify_ghost_token
    token = authorization.replace("Bearer ", "")
    return verify_ghost_token(token)


# ── Encrypted credential storage (DB-backed) ────────────────────────────────

try:
    from cryptography.fernet import Fernet
except ImportError:
    Fernet = None  # type: ignore

_CONNECTOR_KEY = os.getenv("CONNECTOR_ENCRYPTION_KEY") or ""
if not _CONNECTOR_KEY and os.getenv("SMTP_ENCRYPTION_KEY", ""):
    logging.getLogger(__name__).warning("CONNECTOR_ENCRYPTION_KEY not set — falling back to SMTP_ENCRYPTION_KEY (set a separate key in production)")
    _CONNECTOR_KEY = os.getenv("SMTP_ENCRYPTION_KEY", "")
_connector_fernet = Fernet(_CONNECTOR_KEY.encode()) if Fernet and _CONNECTOR_KEY else None


class GhostConnectorCred(Base):
    __tablename__ = "ghost_connector_creds"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    user_id = Column(UUID(as_uuid=True), nullable=False, index=True)
    app_id = Column(String(100), nullable=False, index=True)
    encrypted_creds = Column(Text, nullable=False)
    connected_at = Column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))


def _encrypt_creds(creds: dict) -> str:
    """Encrypt credentials dict to a Fernet token string."""
    raw = json.dumps(creds)
    if _connector_fernet:
        return _connector_fernet.encrypt(raw.encode()).decode()
    logger.warning("CONNECTOR_ENCRYPTION_KEY not set — storing credentials unencrypted")
    return raw


def _decrypt_creds(ciphertext: str) -> dict:
    """Decrypt credentials string back to a dict."""
    if _connector_fernet:
        try:
            return json.loads(_connector_fernet.decrypt(ciphertext.encode()).decode())
        except Exception:
            pass  # Fallback: try as plain JSON (for pre-encryption rows)
    try:
        return json.loads(ciphertext)
    except Exception:
        return {}


async def _get_creds(user_id, app_id: str, db: AsyncSession) -> dict | None:
    """Load and decrypt credentials from DB."""
    result = await db.execute(
        select(GhostConnectorCred).where(
            and_(GhostConnectorCred.user_id == user_id, GhostConnectorCred.app_id == app_id)
        )
    )
    row = result.scalar_one_or_none()
    if not row:
        return None
    return _decrypt_creds(row.encrypted_creds)


async def _set_creds(user_id, app_id: str, creds: dict, db: AsyncSession):
    """Encrypt and upsert credentials in DB."""
    encrypted = _encrypt_creds(creds)
    result = await db.execute(
        select(GhostConnectorCred).where(
            and_(GhostConnectorCred.user_id == user_id, GhostConnectorCred.app_id == app_id)
        )
    )
    row = result.scalar_one_or_none()
    if row:
        row.encrypted_creds = encrypted
        row.connected_at = datetime.now(timezone.utc)
    else:
        db.add(GhostConnectorCred(
            user_id=user_id, app_id=app_id,
            encrypted_creds=encrypted,
        ))
    await db.flush()


async def _del_creds(user_id, app_id: str, db: AsyncSession):
    """Delete credentials from DB."""
    result = await db.execute(
        select(GhostConnectorCred).where(
            and_(GhostConnectorCred.user_id == user_id, GhostConnectorCred.app_id == app_id)
        )
    )
    row = result.scalar_one_or_none()
    if row:
        await db.delete(row)
        await db.flush()


async def _get_connected_app_ids(user_id, db: AsyncSession) -> set[str]:
    """Get all connected app IDs for a user."""
    result = await db.execute(
        select(GhostConnectorCred.app_id).where(GhostConnectorCred.user_id == user_id)
    )
    return {row[0] for row in result.all()}


# ── App Registry ─────────────────────────────────────────────────────────────
# Each app: id, name, category, icon (Ionicons name), auth_fields, actions

APP_REGISTRY: dict[str, dict] = {
    # ── CRM ──────────────────────────────────────────────────────────────
    "hubspot": {
        "name": "HubSpot", "category": "CRM", "icon": "people",
        "auth_fields": [{"key": "api_key", "label": "Private App Token", "secure": True}],
        "setup": "Go to HubSpot → Settings → Integrations → Private Apps → Create → Copy the token.",
        "actions": ["get_contacts", "create_contact", "get_deals", "create_deal", "search"],
    },
    "salesforce": {
        "name": "Salesforce", "category": "CRM", "icon": "cloud",
        "auth_fields": [{"key": "api_key", "label": "Access Token", "secure": True}, {"key": "instance_url", "label": "Instance URL"}],
        "setup": "Go to Salesforce → Setup → Apps → App Manager → New Connected App → Copy the Consumer Key and generate an access token.",
        "actions": ["get_leads", "create_lead", "get_opportunities", "create_case", "search"],
    },
    "pipedrive": {
        "name": "Pipedrive", "category": "CRM", "icon": "funnel",
        "auth_fields": [{"key": "api_key", "label": "API Token", "secure": True}],
        "setup": "Go to Pipedrive → Settings → Personal preferences → API → Copy your API token.",
        "actions": ["get_deals", "create_deal", "get_persons", "create_person", "search"],
    },
    "gohighlevel": {
        "name": "GoHighLevel", "category": "CRM", "icon": "rocket",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}, {"key": "location_id", "label": "Location ID"}],
        "setup": "Go to GoHighLevel → Settings → Business Profile → API Keys → Create API Key.",
        "actions": ["get_contacts", "create_contact", "get_opportunities", "create_opportunity", "search"],
    },
    "zoho_crm": {
        "name": "Zoho CRM", "category": "CRM", "icon": "globe",
        "auth_fields": [{"key": "api_key", "label": "OAuth Token", "secure": True}],
        "setup": "Go to Zoho API Console → Self Client → Generate token with scope: ZohoCRM.modules.ALL.",
        "actions": ["get_leads", "create_lead", "get_deals", "search"],
    },
    "ringy": {
        "name": "Ringy", "category": "CRM", "icon": "call",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Ringy → Settings → Account Settings → Manage Account → API Keys → Create API Key (enable permissions: Lead data, Call data, Call recordings, Lead sold products, Create appointment). Paste the key here.",
        "actions": ["get_lead", "get_call", "get_sold_products", "get_call_recordings", "create_appointment"],
        "action_hints": {
            "get_lead": "lead_id=<UUID of the lead to look up>",
            "get_call": "call_id=<UUID of the call to look up>",
            "get_sold_products": "start_date=YYYY-MM-DD HH:mm:ss|end_date=YYYY-MM-DD HH:mm:ss (both optional, defaults to last 30 days)",
            "get_call_recordings": "start_date=YYYY-MM-DD HH:mm:ss|end_date=YYYY-MM-DD HH:mm:ss (both optional, defaults to last 30 days)",
            "create_appointment": "start=YYYY-MM-DD HH:mm:ss (required, UTC)|lead_id=<UUID> OR lead_phone=<phone number> (one required)|lead_first_name=...|lead_last_name=...|comments=...|duration_minutes=30",
        },
    },
    "close": {
        "name": "Close", "category": "CRM", "icon": "checkmark-circle",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Close → Settings → API Keys → Generate API Key.",
        "actions": ["get_leads", "create_lead", "get_opportunities", "search"],
    },
    "copper": {
        "name": "Copper", "category": "CRM", "icon": "diamond",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}, {"key": "email", "label": "Account Email"}],
        "setup": "Go to Copper → Settings → Integrations → API Keys → Generate.",
        "actions": ["get_leads", "create_lead", "get_opportunities", "search"],
    },
    "freshsales": {
        "name": "Freshsales", "category": "CRM", "icon": "leaf",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}, {"key": "domain", "label": "Freshsales Domain"}],
        "setup": "Go to Freshsales → Settings → API Settings → Copy your API Key. Domain is your-org.freshsales.io.",
        "actions": ["get_contacts", "create_contact", "get_deals", "search"],
    },
    "monday_crm": {
        "name": "Monday CRM", "category": "CRM", "icon": "grid",
        "auth_fields": [{"key": "api_key", "label": "API Token", "secure": True}],
        "setup": "Go to Monday.com → Profile picture → Developers → My Access Tokens → Copy token.",
        "actions": ["get_items", "create_item", "get_boards", "search"],
    },
    "keap": {
        "name": "Keap", "category": "CRM", "icon": "key",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Keap → Settings → API → Generate Personal Access Token.",
        "actions": ["get_contacts", "create_contact", "get_opportunities", "search"],
    },
    "insightly": {
        "name": "Insightly", "category": "CRM", "icon": "eye",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Insightly → User Settings → API → Copy API Key.",
        "actions": ["get_contacts", "create_contact", "get_opportunities", "search"],
    },
    "nutshell": {
        "name": "Nutshell", "category": "CRM", "icon": "nutrition",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}, {"key": "email", "label": "Account Email"}],
        "setup": "Go to Nutshell → Setup → API → Enable API access and copy your key.",
        "actions": ["get_leads", "create_lead", "get_contacts", "search"],
    },
    "less_annoying_crm": {
        "name": "Less Annoying CRM", "category": "CRM", "icon": "happy",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}, {"key": "user_code", "label": "User Code"}],
        "setup": "Go to Less Annoying CRM → Settings → Programmer API → Copy your API key and user code.",
        "actions": ["get_contacts", "create_contact", "search"],
    },
    "liondesk": {
        "name": "LionDesk", "category": "CRM", "icon": "paw",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to LionDesk → Settings → API → Generate API Key.",
        "actions": ["get_contacts", "create_contact", "search"],
    },
    "follow_up_boss": {
        "name": "Follow Up Boss", "category": "CRM", "icon": "person-add",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Follow Up Boss → Admin → API → Copy your API Key.",
        "actions": ["get_leads", "create_lead", "search"],
    },
    "kvcore": {
        "name": "kvCORE", "category": "CRM", "icon": "home",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to kvCORE → Settings → API → Copy your API Key.",
        "actions": ["get_leads", "create_lead", "search"],
    },
    "chime": {
        "name": "Chime", "category": "CRM", "icon": "notifications",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Chime → Settings → Integrations → API → Copy your API Key.",
        "actions": ["get_leads", "create_lead", "search"],
    },

    # ── Accounting & Finance ─────────────────────────────────────────────
    "quickbooks": {
        "name": "QuickBooks", "category": "Accounting", "icon": "calculator",
        "auth_fields": [{"key": "api_key", "label": "Access Token", "secure": True}, {"key": "realm_id", "label": "Realm/Company ID"}],
        "setup": "Go to developer.intuit.com → Create App → Get OAuth tokens → Copy Access Token and Realm ID.",
        "actions": ["get_invoices", "create_invoice", "get_customers", "get_expenses", "create_expense"],
    },
    "xero": {
        "name": "Xero", "category": "Accounting", "icon": "cash",
        "auth_fields": [{"key": "api_key", "label": "Access Token", "secure": True}, {"key": "tenant_id", "label": "Tenant ID"}],
        "setup": "Go to developer.xero.com → My Apps → Create → Copy Access Token and Tenant ID.",
        "actions": ["get_invoices", "create_invoice", "get_contacts", "get_accounts"],
    },
    "freshbooks": {
        "name": "FreshBooks", "category": "Accounting", "icon": "book",
        "auth_fields": [{"key": "api_key", "label": "Bearer Token", "secure": True}, {"key": "account_id", "label": "Account ID"}],
        "setup": "Go to FreshBooks → Settings → Developer → Create App → Copy Bearer Token and Account ID.",
        "actions": ["get_invoices", "create_invoice", "get_clients", "get_expenses"],
    },
    "wave": {
        "name": "Wave", "category": "Accounting", "icon": "water",
        "auth_fields": [{"key": "api_key", "label": "API Token", "secure": True}],
        "setup": "Go to Wave → Settings → Developer → Create Token.",
        "actions": ["get_invoices", "get_customers", "get_accounts"],
    },
    "sage": {
        "name": "Sage", "category": "Accounting", "icon": "leaf",
        "auth_fields": [{"key": "api_key", "label": "Access Token", "secure": True}],
        "setup": "Go to developer.sage.com → My Apps → Create → Copy Access Token.",
        "actions": ["get_invoices", "create_invoice", "get_contacts"],
    },
    "zoho_books": {
        "name": "Zoho Books", "category": "Accounting", "icon": "document-text",
        "auth_fields": [{"key": "api_key", "label": "OAuth Token", "secure": True}, {"key": "org_id", "label": "Organization ID"}],
        "setup": "Go to Zoho API Console → Self Client → Generate token with scope: ZohoBooks.fullaccess.all.",
        "actions": ["get_invoices", "create_invoice", "get_contacts", "get_expenses"],
    },
    "billcom": {
        "name": "Bill.com", "category": "Accounting", "icon": "receipt",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}, {"key": "org_id", "label": "Organization ID"}],
        "setup": "Go to Bill.com → Settings → Developer → API → Copy your credentials.",
        "actions": ["get_invoices", "get_bills", "get_vendors"],
    },
    "gusto": {
        "name": "Gusto", "category": "Accounting", "icon": "people",
        "auth_fields": [{"key": "api_key", "label": "API Token", "secure": True}],
        "setup": "Go to developer.gusto.com → Create App → Copy API Token.",
        "actions": ["get_employees", "get_payroll", "get_company"],
    },
    "adp": {
        "name": "ADP", "category": "Accounting", "icon": "briefcase",
        "auth_fields": [{"key": "client_id", "label": "Client ID"}, {"key": "client_secret", "label": "Client Secret", "secure": True}],
        "setup": "Go to ADP Marketplace → My Apps → Create → Copy Client ID and Secret.",
        "actions": ["get_employees", "get_payroll"],
    },
    "plaid": {
        "name": "Plaid", "category": "Accounting", "icon": "card",
        "auth_fields": [{"key": "client_id", "label": "Client ID"}, {"key": "secret", "label": "Secret", "secure": True}],
        "setup": "Go to dashboard.plaid.com → Team → Keys → Copy Client ID and Secret.",
        "actions": ["get_accounts", "get_transactions", "get_balance"],
    },

    # ── Project Management ───────────────────────────────────────────────
    "asana": {
        "name": "Asana", "category": "Project Management", "icon": "checkmark-done",
        "auth_fields": [{"key": "api_key", "label": "Personal Access Token", "secure": True}],
        "setup": "Go to Asana → My Profile Settings → Apps → Manage Developer Apps → Personal Access Tokens → Create.",
        "actions": ["get_tasks", "create_task", "get_projects", "update_task", "search"],
    },
    "trello": {
        "name": "Trello", "category": "Project Management", "icon": "albums",
        "auth_fields": [{"key": "api_key", "label": "API Key"}, {"key": "token", "label": "Token", "secure": True}],
        "setup": "Go to trello.com/power-ups/admin → New → Copy API Key. Then visit trello.com/1/authorize?key=YOUR_KEY&scope=read,write&response_type=token to get your Token.",
        "actions": ["get_cards", "create_card", "get_boards", "move_card", "search"],
    },
    "monday": {
        "name": "Monday.com", "category": "Project Management", "icon": "grid",
        "auth_fields": [{"key": "api_key", "label": "API Token", "secure": True}],
        "setup": "Go to Monday.com → Profile picture → Developers → My Access Tokens → Copy.",
        "actions": ["get_items", "create_item", "get_boards", "update_item", "search"],
    },
    "clickup": {
        "name": "ClickUp", "category": "Project Management", "icon": "flash",
        "auth_fields": [{"key": "api_key", "label": "API Token", "secure": True}],
        "setup": "Go to ClickUp → Settings → Apps → Generate API Token.",
        "actions": ["get_tasks", "create_task", "get_spaces", "update_task", "search"],
    },
    "notion": {
        "name": "Notion", "category": "Project Management", "icon": "document",
        "auth_fields": [{"key": "api_key", "label": "Integration Token", "secure": True}],
        "setup": "Go to notion.so/my-integrations → New Integration → Copy the Internal Integration Token. Then share your pages/databases with the integration.",
        "actions": ["get_pages", "create_page", "search", "update_page", "query_database"],
    },
    "jira": {
        "name": "Jira", "category": "Project Management", "icon": "bug",
        "auth_fields": [{"key": "api_key", "label": "API Token", "secure": True}, {"key": "email", "label": "Email"}, {"key": "domain", "label": "Jira Domain"}],
        "setup": "Go to id.atlassian.com/manage-profile/security → Create API Token. Domain is your-org.atlassian.net.",
        "actions": ["get_issues", "create_issue", "update_issue", "search", "get_projects"],
    },
    "linear": {
        "name": "Linear", "category": "Project Management", "icon": "git-branch",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Linear → Settings → API → Personal API Keys → Create.",
        "actions": ["get_issues", "create_issue", "update_issue", "search"],
    },
    "basecamp": {
        "name": "Basecamp", "category": "Project Management", "icon": "flag",
        "auth_fields": [{"key": "api_key", "label": "Access Token", "secure": True}, {"key": "account_id", "label": "Account ID"}],
        "setup": "Go to launchpad.37signals.com/integrations → Register App → Copy Access Token.",
        "actions": ["get_todos", "create_todo", "get_projects"],
    },
    "wrike": {
        "name": "Wrike", "category": "Project Management", "icon": "layers",
        "auth_fields": [{"key": "api_key", "label": "Access Token", "secure": True}],
        "setup": "Go to Wrike → Apps & Integrations → API → Create Token.",
        "actions": ["get_tasks", "create_task", "get_folders", "search"],
    },
    "teamwork": {
        "name": "Teamwork", "category": "Project Management", "icon": "people-circle",
        "auth_fields": [{"key": "api_key", "label": "API Token", "secure": True}, {"key": "domain", "label": "Teamwork Domain"}],
        "setup": "Go to Teamwork → Settings → API → Generate Token.",
        "actions": ["get_tasks", "create_task", "get_projects"],
    },
    "todoist": {
        "name": "Todoist", "category": "Project Management", "icon": "checkbox",
        "auth_fields": [{"key": "api_key", "label": "API Token", "secure": True}],
        "setup": "Go to todoist.com/prefs/integrations → Developer → Copy API Token.",
        "actions": ["get_tasks", "create_task", "update_task", "get_projects"],
    },

    # ── Communication ────────────────────────────────────────────────────
    "slack": {
        "name": "Slack", "category": "Communication", "icon": "chatbubbles",
        "auth_fields": [{"key": "api_key", "label": "Bot Token (xoxb-...)", "secure": True}],
        "setup": "Go to api.slack.com/apps → Create App → OAuth & Permissions → Install to Workspace → Copy Bot Token.",
        "actions": ["send_message", "get_channels", "get_messages", "search"],
    },
    "discord": {
        "name": "Discord", "category": "Communication", "icon": "game-controller",
        "auth_fields": [{"key": "api_key", "label": "Bot Token", "secure": True}],
        "setup": "Go to discord.com/developers → New Application → Bot → Copy Token.",
        "actions": ["send_message", "get_channels", "get_messages"],
    },
    "teams": {
        "name": "Microsoft Teams", "category": "Communication", "icon": "people",
        "auth_fields": [{"key": "api_key", "label": "Access Token", "secure": True}],
        "setup": "Go to Azure Portal → App Registrations → New → Copy Token (requires Graph API permissions).",
        "actions": ["send_message", "get_channels", "get_messages"],
    },
    "zoom": {
        "name": "Zoom", "category": "Communication", "icon": "videocam",
        "auth_fields": [{"key": "api_key", "label": "JWT Token", "secure": True}],
        "setup": "Go to marketplace.zoom.us → Create App → JWT → Copy Token.",
        "actions": ["create_meeting", "get_meetings", "get_recordings"],
    },
    "twilio": {
        "name": "Twilio", "category": "Communication", "icon": "call",
        "auth_fields": [{"key": "account_sid", "label": "Account SID"}, {"key": "auth_token", "label": "Auth Token", "secure": True}],
        "setup": "Go to twilio.com/console → Copy Account SID and Auth Token.",
        "actions": ["send_sms", "make_call", "get_messages"],
    },
    "whatsapp_business": {
        "name": "WhatsApp Business", "category": "Communication", "icon": "logo-whatsapp",
        "auth_fields": [{"key": "api_key", "label": "Access Token", "secure": True}, {"key": "phone_id", "label": "Phone Number ID"}],
        "setup": "Go to developers.facebook.com → My Apps → WhatsApp → API Setup → Copy token and Phone Number ID.",
        "actions": ["send_message", "get_messages"],
    },
    "telegram": {
        "name": "Telegram", "category": "Communication", "icon": "paper-plane",
        "auth_fields": [{"key": "api_key", "label": "Bot Token", "secure": True}],
        "setup": "Message @BotFather on Telegram → /newbot → Copy the token.",
        "actions": ["send_message", "get_updates"],
    },
    "intercom": {
        "name": "Intercom", "category": "Communication", "icon": "chatbox",
        "auth_fields": [{"key": "api_key", "label": "Access Token", "secure": True}],
        "setup": "Go to Intercom → Settings → Integrations → Developer Hub → New App → Copy Access Token.",
        "actions": ["get_contacts", "create_contact", "send_message", "get_conversations"],
    },

    # ── Calendar & Scheduling ────────────────────────────────────────────
    "google_calendar": {
        "name": "Google Calendar", "category": "Calendar", "icon": "calendar",
        "auth_fields": [{"key": "api_key", "label": "API Key / OAuth Token", "secure": True}],
        "setup": "Go to console.cloud.google.com → APIs & Services → Credentials → Create API Key or OAuth Token.",
        "actions": ["get_events", "create_event", "update_event", "delete_event"],
    },
    "outlook_calendar": {
        "name": "Outlook Calendar", "category": "Calendar", "icon": "mail",
        "auth_fields": [{"key": "api_key", "label": "Access Token", "secure": True}],
        "setup": "Go to Azure Portal → App Registrations → New → Certificates & Secrets → Copy Token.",
        "actions": ["get_events", "create_event", "update_event"],
    },
    "calendly": {
        "name": "Calendly", "category": "Calendar", "icon": "time",
        "auth_fields": [{"key": "api_key", "label": "Personal Access Token", "secure": True}],
        "setup": "Go to calendly.com/integrations/api → Create Token → Copy.",
        "actions": ["get_events", "get_event_types", "get_availability"],
    },
    "acuity": {
        "name": "Acuity Scheduling", "category": "Calendar", "icon": "alarm",
        "auth_fields": [{"key": "user_id", "label": "User ID"}, {"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Acuity → Integrations → API → Copy User ID and API Key.",
        "actions": ["get_appointments", "create_appointment", "get_availability"],
    },
    "calcom": {
        "name": "Cal.com", "category": "Calendar", "icon": "today",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Cal.com → Settings → Developer → API Keys → Create.",
        "actions": ["get_bookings", "get_event_types", "get_availability"],
    },

    # ── E-commerce ───────────────────────────────────────────────────────
    "shopify": {
        "name": "Shopify", "category": "E-commerce", "icon": "cart",
        "auth_fields": [{"key": "api_key", "label": "Admin API Access Token", "secure": True}, {"key": "store_url", "label": "Store URL (your-store.myshopify.com)"}],
        "setup": "Go to Shopify Admin → Settings → Apps → Develop apps → Create → Admin API Access Token.",
        "actions": ["get_orders", "get_products", "get_customers", "create_product", "get_inventory"],
    },
    "stripe": {
        "name": "Stripe", "category": "E-commerce", "icon": "card",
        "auth_fields": [{"key": "api_key", "label": "Secret Key (sk_...)", "secure": True}],
        "setup": "Go to dashboard.stripe.com → Developers → API Keys → Copy Secret Key.",
        "actions": ["get_payments", "get_customers", "create_payment_link", "get_invoices", "get_balance"],
    },
    "square": {
        "name": "Square", "category": "E-commerce", "icon": "cube",
        "auth_fields": [{"key": "api_key", "label": "Access Token", "secure": True}],
        "setup": "Go to developer.squareup.com → Applications → Credentials → Copy Access Token.",
        "actions": ["get_payments", "get_catalog", "get_customers", "get_orders"],
    },
    "woocommerce": {
        "name": "WooCommerce", "category": "E-commerce", "icon": "storefront",
        "auth_fields": [{"key": "consumer_key", "label": "Consumer Key"}, {"key": "consumer_secret", "label": "Consumer Secret", "secure": True}, {"key": "store_url", "label": "Store URL"}],
        "setup": "Go to WooCommerce → Settings → Advanced → REST API → Add Key → Copy Consumer Key and Secret.",
        "actions": ["get_orders", "get_products", "get_customers", "create_product"],
    },
    "paypal": {
        "name": "PayPal", "category": "E-commerce", "icon": "logo-paypal",
        "auth_fields": [{"key": "client_id", "label": "Client ID"}, {"key": "secret", "label": "Secret", "secure": True}],
        "setup": "Go to developer.paypal.com → My Apps → Create → Copy Client ID and Secret.",
        "actions": ["get_transactions", "get_balance", "create_invoice"],
    },
    "amazon_seller": {
        "name": "Amazon Seller", "category": "E-commerce", "icon": "logo-amazon",
        "auth_fields": [{"key": "api_key", "label": "Access Key", "secure": True}, {"key": "secret_key", "label": "Secret Key", "secure": True}],
        "setup": "Go to sellercentral.amazon.com → Settings → User Permissions → Developer Access.",
        "actions": ["get_orders", "get_inventory", "get_products"],
    },
    "etsy": {
        "name": "Etsy", "category": "E-commerce", "icon": "pricetag",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to etsy.com/developers → Create App → Copy API Key.",
        "actions": ["get_orders", "get_listings", "get_shop"],
    },

    # ── Storage & Docs ───────────────────────────────────────────────────
    "google_drive": {
        "name": "Google Drive", "category": "Storage", "icon": "folder",
        "auth_fields": [{"key": "api_key", "label": "OAuth Token", "secure": True}],
        "setup": "Go to console.cloud.google.com → APIs → Enable Drive API → Create OAuth credentials.",
        "actions": ["list_files", "search_files", "get_file", "create_file"],
    },
    "dropbox": {
        "name": "Dropbox", "category": "Storage", "icon": "cloud-download",
        "auth_fields": [{"key": "api_key", "label": "Access Token", "secure": True}],
        "setup": "Go to dropbox.com/developers → My Apps → Create → Generate Access Token.",
        "actions": ["list_files", "search_files", "get_file"],
    },
    "onedrive": {
        "name": "OneDrive", "category": "Storage", "icon": "cloud",
        "auth_fields": [{"key": "api_key", "label": "Access Token", "secure": True}],
        "setup": "Go to Azure Portal → App Registrations → Create → Get Token via Microsoft Graph API.",
        "actions": ["list_files", "search_files", "get_file"],
    },
    "box": {
        "name": "Box", "category": "Storage", "icon": "cube",
        "auth_fields": [{"key": "api_key", "label": "Developer Token", "secure": True}],
        "setup": "Go to developer.box.com → My Apps → Create → Configuration → Generate Developer Token.",
        "actions": ["list_files", "search_files", "get_file"],
    },
    "google_sheets": {
        "name": "Google Sheets", "category": "Storage", "icon": "grid",
        "auth_fields": [{"key": "api_key", "label": "OAuth Token", "secure": True}],
        "setup": "Go to console.cloud.google.com → APIs → Enable Sheets API → Create OAuth credentials.",
        "actions": ["get_spreadsheet", "update_cells", "create_spreadsheet", "search"],
    },
    "excel_online": {
        "name": "Microsoft Excel Online", "category": "Storage", "icon": "grid",
        "auth_fields": [{"key": "access_token", "label": "Microsoft Graph Access Token", "secure": True}],
        "oauth_flow": "microsoft",
        "setup": "Tap 'Connect with Microsoft' to sign in with your Microsoft account. You'll be asked to allow GoFarther to read and write your Excel files in OneDrive.",
        "actions": [
            # Core
            "list_workbooks", "get_worksheets", "read_range", "write_range",
            "add_row", "create_workbook", "get_cell_value",
            # Sheet management
            "add_worksheet", "rename_worksheet", "delete_worksheet", "copy_worksheet",
            # Row/range manipulation
            "delete_row", "clear_range", "set_formula",
            # Tables
            "create_table", "add_table_row",
            # Formatting
            "format_range", "set_number_format", "autofit_columns",
            # Charts
            "create_chart", "delete_chart",
            # Queries & analysis
            "find_cell", "sum_column", "get_last_row", "filter_rows", "sort_range",
            # File-level utilities
            "download_as_pdf", "download_workbook", "protect_sheet", "share_workbook",
            # Tier 6 — advanced
            "calculate_workbook", "list_tables", "list_pivot_tables", "refresh_pivot",
            "list_comments", "set_cell_comment",
            "insert_rows", "insert_columns",
            "set_column_width", "set_row_height",
            "freeze_panes", "unfreeze_panes",
            "merge_cells", "unmerge_cells",
            "create_named_range", "list_named_ranges",
            "add_hyperlink", "unprotect_sheet",
            "range_details", "set_conditional_format",
            "execute_function",
        ],
        "action_hints": {
            "list_workbooks": "no params — lists every .xlsx in the user's OneDrive",
            "get_worksheets": "workbook_id=<partial name or empty if only 1 file>",
            "read_range": "workbook_id=<partial name or empty>|range=<A1 notation, e.g. A1:C10>",
            "write_range": "workbook_id=<partial name>|range=<A1 notation>|values=<comma-separated or 2D JSON array>",
            "add_row": "workbook_id=<partial name or empty>|values=<comma-separated, e.g. coffee,50>",
            "create_workbook": "name=<any name, .xlsx added automatically>",
            "get_cell_value": "workbook_id=<partial name>|cell=<A1 notation, e.g. B2>",
            "add_worksheet": "workbook_id=<partial name>|name=<new sheet name>",
            "rename_worksheet": "workbook_id=<partial name>|worksheet=<current sheet name>|name=<new name>",
            "delete_worksheet": "workbook_id=<partial name>|worksheet=<sheet name to delete>",
            "copy_worksheet": "workbook_id=<partial name>|worksheet=<source sheet name>|name=<new sheet name>",
            "delete_row": "workbook_id=<partial name>|row=<1-based row number>",
            "clear_range": "workbook_id=<partial name>|range=<A1 notation>",
            "set_formula": "workbook_id=<partial name>|cell=<A1 notation, e.g. B11>|formula=<Excel formula like =SUM(B2:B10)>",
            "create_table": "workbook_id=<partial name>|range=<A1 notation with headers, e.g. A1:D10>|name=<table name>",
            "add_table_row": "workbook_id=<partial name>|table=<table name>|values=<comma-separated or JSON array>",
            "format_range": "workbook_id=<partial name>|range=<A1>|bold=<true/false>|italic=<true/false>|color=<hex like #FF0000>|fill=<hex background>",
            "set_number_format": "workbook_id=<partial name>|range=<A1>|format=<Excel format code, e.g. $#,##0.00 or 0.00% or m/d/yyyy>",
            "autofit_columns": "workbook_id=<partial name>|range=<A1 to autofit, e.g. A:D or A1:D1>",
            "create_chart": "workbook_id=<partial name>|range=<data range>|type=<ColumnClustered, Bar, Line, Pie, Scatter>|title=<chart title>",
            "delete_chart": "workbook_id=<partial name>|chart=<chart name>",
            "find_cell": "workbook_id=<partial name>|search=<text or number to find>",
            "sum_column": "workbook_id=<partial name>|column=<column letter like B>",
            "get_last_row": "workbook_id=<partial name> — returns the last used row number",
            "filter_rows": "workbook_id=<partial name>|column=<column letter>|value=<value to match>",
            "sort_range": "workbook_id=<partial name>|range=<A1>|column=<0-based column index>|ascending=<true/false>",
            "download_as_pdf": "workbook_id=<partial name> — returns a download URL for a PDF version",
            "download_workbook": "workbook_id=<partial name>|format=<optional target format: xlsx (default), pdf, csv, txt, docx, ods, html> — returns the file for download in the requested format. xlsx/pdf come back as a URL; any other format is converted server-side and returned inline as base64.",
            "protect_sheet": "workbook_id=<partial name>|worksheet=<sheet name, defaults to Sheet1>",
            "share_workbook": "workbook_id=<partial name>|scope=<view or edit, defaults to view>",
            "calculate_workbook": "workbook_id=<partial name>|type=<Recalculate, Full, or FullRebuild> — forces formula recalculation",
            "list_tables": "workbook_id=<partial name> — lists every Excel Table in the workbook",
            "list_pivot_tables": "workbook_id=<partial name>|worksheet=<sheet name, optional>",
            "refresh_pivot": "workbook_id=<partial name>|worksheet=<sheet name>|pivot=<pivot table name, optional — refreshes all if omitted>",
            "list_comments": "workbook_id=<partial name> — lists every cell comment in the workbook",
            "set_cell_comment": "workbook_id=<partial name>|worksheet=<sheet name>|cell=<A1>|text=<comment text>",
            "insert_rows": "workbook_id=<partial name>|worksheet=<sheet name>|row=<1-based row>|count=<how many to insert, defaults 1>",
            "insert_columns": "workbook_id=<partial name>|worksheet=<sheet name>|column=<column letter>|count=<how many to insert, defaults 1>",
            "set_column_width": "workbook_id=<partial name>|worksheet=<sheet name>|column=<letter or range like A:C>|width=<points, e.g. 120>",
            "set_row_height": "workbook_id=<partial name>|worksheet=<sheet name>|row=<row number or range like 1:3>|height=<points, e.g. 24>",
            "freeze_panes": "workbook_id=<partial name>|worksheet=<sheet name>|rows=<number of top rows to freeze>|columns=<number of left columns to freeze>",
            "unfreeze_panes": "workbook_id=<partial name>|worksheet=<sheet name>",
            "merge_cells": "workbook_id=<partial name>|worksheet=<sheet name>|range=<A1 notation>|across=<true/false, true merges per-row>",
            "unmerge_cells": "workbook_id=<partial name>|worksheet=<sheet name>|range=<A1 notation>",
            "create_named_range": "workbook_id=<partial name>|name=<name of the named range>|reference=<A1 notation like Sheet1!A1:B10>|comment=<optional>",
            "list_named_ranges": "workbook_id=<partial name> — lists every workbook-level named range",
            "add_hyperlink": "workbook_id=<partial name>|worksheet=<sheet name>|cell=<A1>|url=<https://...>|display=<link text, optional>",
            "unprotect_sheet": "workbook_id=<partial name>|worksheet=<sheet name>",
            "range_details": "workbook_id=<partial name>|worksheet=<sheet name>|range=<A1> — returns values, formulas, formats, and used range info",
            "set_conditional_format": "workbook_id=<partial name>|worksheet=<sheet name>|range=<A1>|rule=<colorScale, dataBar, iconSet, top, bottom, aboveAverage, presetCriteria, custom>|color=<hex for simple rules>",
            "execute_function": "function=<Excel function name like VLOOKUP, SUMIF, XLOOKUP>|args=<JSON array of arguments, can include range refs like Sheet1!A1:B10>|workbook_id=<partial name>",
        },
    },
    "airtable": {
        "name": "Airtable", "category": "Storage", "icon": "apps",
        "auth_fields": [{"key": "api_key", "label": "Personal Access Token", "secure": True}],
        "setup": "Go to airtable.com/create/tokens → Create Token → Add scopes (data.records:read, data.records:write).",
        "actions": ["get_records", "create_record", "update_record", "search"],
    },

    # ── Email (mailbox access: read / search / reply / send / manage) ────
    "outlook_mail": {
        "name": "Microsoft Outlook", "category": "Email", "icon": "mail",
        "auth_fields": [{"key": "access_token", "label": "Microsoft Graph Access Token", "secure": True}],
        "oauth_flow": "microsoft",
        "setup": "Tap 'Connect with Microsoft' to sign in and allow GoFarther to read, send, and manage messages in your Outlook inbox.",
        "actions": [
            "list_inbox", "search_emails", "read_email", "reply_to_email",
            "send_email", "mark_read", "mark_unread", "archive", "delete",
            "move_to_folder", "list_folders", "download_attachment",
        ],
        "action_hints": {
            "list_inbox": "folder=<folder name, defaults to Inbox>|limit=<1-50, default 20> — returns id, from, subject, snippet, received date",
            "search_emails": "query=<text>|from=<email>|subject=<text>|unread=<true/false>|limit=<1-50>",
            "read_email": "message_id=<id from list_inbox/search> — returns full body + attachment list",
            "reply_to_email": "message_id=<id>|body=<html or plain text>|reply_all=<true/false, default false>",
            "send_email": "to=<email(s) comma-separated>|subject=<subject>|body=<html>|cc=<optional>|bcc=<optional>",
            "mark_read": "message_id=<id>",
            "mark_unread": "message_id=<id>",
            "archive": "message_id=<id> — moves to Archive folder",
            "delete": "message_id=<id> — moves to Deleted Items",
            "move_to_folder": "message_id=<id>|folder=<destination folder name or id>",
            "list_folders": "no params — returns all mail folders",
            "download_attachment": "message_id=<id>|attachment_id=<id from read_email>",
        },
    },
    "gmail": {
        "name": "Gmail", "category": "Email", "icon": "mail",
        "auth_fields": [{"key": "access_token", "label": "Google OAuth Access Token", "secure": True}],
        "oauth_flow": "google",
        "setup": "Tap 'Connect with Google' to sign in and allow GoFarther to read, send, and manage messages in your Gmail account.",
        "actions": [
            "list_inbox", "search_emails", "read_email", "reply_to_email",
            "send_email", "mark_read", "mark_unread", "archive", "delete",
            "move_to_folder", "list_folders", "download_attachment",
        ],
        "action_hints": {
            "list_inbox": "label=<label name, defaults to INBOX>|limit=<1-50, default 20>",
            "search_emails": "query=<Gmail search query, e.g. 'from:boss@acme.com subject:invoice'>|limit=<1-50>",
            "read_email": "message_id=<id>",
            "reply_to_email": "message_id=<id>|body=<html or plain>|reply_all=<true/false>",
            "send_email": "to=<email(s)>|subject=<subject>|body=<html>|cc=<optional>|bcc=<optional>",
            "mark_read": "message_id=<id>",
            "mark_unread": "message_id=<id>",
            "archive": "message_id=<id> — removes INBOX label",
            "delete": "message_id=<id> — moves to Trash",
            "move_to_folder": "message_id=<id>|folder=<label name>",
            "list_folders": "no params — returns all labels (Gmail calls folders 'labels')",
            "download_attachment": "message_id=<id>|attachment_id=<id>",
        },
    },
    "neo_mail": {
        "name": "Neo Business Email", "category": "Email", "icon": "mail",
        "auth_fields": [
            {"key": "username", "label": "Email Address (e.g. you@yourdomain.com)"},
            {"key": "app_password", "label": "Password", "secure": True},
        ],
        "setup": "Enter your Neo email address and password. Server settings (imap.neo.space / smtp.neo.space) are configured automatically. If Neo rejects the password, generate an app-specific password in your Neo account settings.",
        "setup_url": "https://app.neo.space/mail/",
        "actions": [
            "list_inbox", "search_emails", "read_email", "reply_to_email",
            "send_email", "mark_read", "mark_unread", "archive", "delete",
            "move_to_folder", "list_folders", "download_attachment",
        ],
        "action_hints": {
            "list_inbox": "folder=<folder name, defaults to INBOX>|limit=<1-50, default 20>",
            "search_emails": "query=<text>|from=<email>|subject=<text>|unread=<true/false>|limit=<1-50>",
            "read_email": "message_id=<uid>|folder=<folder name, defaults to INBOX>",
            "reply_to_email": "message_id=<uid>|body=<html>|folder=<folder name>",
            "send_email": "to=<email(s)>|subject=<subject>|body=<html>|cc=<optional>|bcc=<optional>",
            "mark_read": "message_id=<uid>|folder=<folder name>",
            "mark_unread": "message_id=<uid>|folder=<folder name>",
            "archive": "message_id=<uid>|folder=<source folder>",
            "delete": "message_id=<uid>|folder=<folder name>",
            "move_to_folder": "message_id=<uid>|folder=<source>|to=<destination folder>",
            "list_folders": "no params",
            "download_attachment": "message_id=<uid>|folder=<folder>|attachment_index=<0-based index from read_email>",
        },
    },
    "titan_mail": {
        "name": "Titan Email", "category": "Email", "icon": "mail",
        "auth_fields": [
            {"key": "username", "label": "Email Address (e.g. you@yourdomain.com)"},
            {"key": "app_password", "label": "Password", "secure": True},
        ],
        "setup": "Enter your Titan email address and password. Server settings (imap.titan.email / smtp.titan.email) are configured automatically. If Titan rejects the password, generate an app-specific password in your Titan account settings.",
        "setup_url": "https://app.titan.email/",
        "actions": [
            "list_inbox", "search_emails", "read_email", "reply_to_email",
            "send_email", "mark_read", "mark_unread", "archive", "delete",
            "move_to_folder", "list_folders", "download_attachment",
        ],
        "action_hints": {
            "list_inbox": "folder=<folder name, defaults to INBOX>|limit=<1-50, default 20>",
            "search_emails": "query=<text>|from=<email>|subject=<text>|unread=<true/false>|limit=<1-50>",
            "read_email": "message_id=<uid>|folder=<folder name, defaults to INBOX>",
            "reply_to_email": "message_id=<uid>|body=<html>|folder=<folder name>",
            "send_email": "to=<email(s)>|subject=<subject>|body=<html>|cc=<optional>|bcc=<optional>",
            "mark_read": "message_id=<uid>|folder=<folder name>",
            "mark_unread": "message_id=<uid>|folder=<folder name>",
            "archive": "message_id=<uid>|folder=<source folder>",
            "delete": "message_id=<uid>|folder=<folder name>",
            "move_to_folder": "message_id=<uid>|folder=<source>|to=<destination folder>",
            "list_folders": "no params",
            "download_attachment": "message_id=<uid>|folder=<folder>|attachment_index=<0-based index from read_email>",
        },
    },
    # The following IMAP-based email presets are generated from a compact
    # preset table below the registry. Each one only asks for email + app
    # password; the hosts/ports are hardcoded in the adapter wrapper so
    # the user doesn't need to know their provider's server settings.
    #
    # Entries are inserted by the _inject_mail_presets() call right after
    # this registry literal closes, so they land in this dict at
    # module-init time but the source file stays readable.

    "imap_mail": {
        "name": "Email (IMAP)", "category": "Email", "icon": "mail",
        "auth_fields": [
            {"key": "username", "label": "Email Address"},
            {"key": "app_password", "label": "App Password", "secure": True},
            # Optional overrides — only needed when autodetection fails. The
            # frontend can render these as an "Advanced" expandable section.
            {"key": "imap_host", "label": "IMAP Server (optional — autodetected from your email)"},
            {"key": "imap_port", "label": "IMAP Port (optional, default 993)"},
            {"key": "smtp_host", "label": "SMTP Server (optional — autodetected)"},
            {"key": "smtp_port", "label": "SMTP Port (optional, default 587)"},
        ],
        "setup": "Enter your email address and password. GoFarther auto-detects the IMAP and SMTP servers for most providers (Gmail, Yahoo, iCloud, Titan, Neo, FastMail, Zoho, and thousands more) using the same autoconfig database Thunderbird uses. Only fill the server fields if auto-detection fails. Most providers require an 'app password' instead of your regular password.",
        "actions": [
            "list_inbox", "search_emails", "read_email", "reply_to_email",
            "send_email", "mark_read", "mark_unread", "archive", "delete",
            "move_to_folder", "list_folders", "download_attachment",
        ],
        "action_hints": {
            "list_inbox": "folder=<folder name, defaults to INBOX>|limit=<1-50, default 20>",
            "search_emails": "query=<text>|from=<email>|subject=<text>|unread=<true/false>|limit=<1-50>",
            "read_email": "message_id=<uid from list_inbox>|folder=<folder name, defaults to INBOX>",
            "reply_to_email": "message_id=<uid>|body=<html>|folder=<folder name>",
            "send_email": "to=<email(s)>|subject=<subject>|body=<html>|cc=<optional>|bcc=<optional>",
            "mark_read": "message_id=<uid>|folder=<folder name>",
            "mark_unread": "message_id=<uid>|folder=<folder name>",
            "archive": "message_id=<uid>|folder=<source folder> — moves to Archive",
            "delete": "message_id=<uid>|folder=<folder name>",
            "move_to_folder": "message_id=<uid>|folder=<source>|to=<destination folder>",
            "list_folders": "no params",
            "download_attachment": "message_id=<uid>|folder=<folder>|attachment_index=<0-based index from read_email>",
        },
    },

    # ── Email Marketing ──────────────────────────────────────────────────
    "mailchimp": {
        "name": "Mailchimp", "category": "Email Marketing", "icon": "mail",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Mailchimp → Account → Extras → API Keys → Create.",
        "actions": ["get_lists", "get_campaigns", "create_campaign", "get_subscribers", "add_subscriber"],
    },
    "convertkit": {
        "name": "ConvertKit", "category": "Email Marketing", "icon": "send",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to ConvertKit → Settings → Advanced → API → Copy API Key.",
        "actions": ["get_subscribers", "add_subscriber", "get_sequences", "get_forms"],
    },
    "klaviyo": {
        "name": "Klaviyo", "category": "Email Marketing", "icon": "megaphone",
        "auth_fields": [{"key": "api_key", "label": "Private API Key", "secure": True}],
        "setup": "Go to Klaviyo → Account → Settings → API Keys → Create Private Key.",
        "actions": ["get_lists", "get_campaigns", "get_profiles", "add_profile"],
    },
    "activecampaign": {
        "name": "ActiveCampaign", "category": "Email Marketing", "icon": "pulse",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}, {"key": "api_url", "label": "API URL"}],
        "setup": "Go to ActiveCampaign → Settings → Developer → Copy API URL and Key.",
        "actions": ["get_contacts", "create_contact", "get_campaigns", "get_lists"],
    },
    "constant_contact": {
        "name": "Constant Contact", "category": "Email Marketing", "icon": "at",
        "auth_fields": [{"key": "api_key", "label": "Access Token", "secure": True}],
        "setup": "Go to app.constantcontact.com/pages/dma/portal → My Applications → Create → Copy Token.",
        "actions": ["get_contacts", "add_contact", "get_campaigns"],
    },
    "brevo": {
        "name": "Brevo (Sendinblue)", "category": "Email Marketing", "icon": "mail-open",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Brevo → Settings → SMTP & API → API Keys → Generate.",
        "actions": ["get_contacts", "create_contact", "get_campaigns", "send_email"],
    },

    # ── HR & Recruiting ──────────────────────────────────────────────────
    "bamboohr": {
        "name": "BambooHR", "category": "HR", "icon": "people",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}, {"key": "subdomain", "label": "Subdomain"}],
        "setup": "Go to BambooHR → Account → API Keys → Add New Key.",
        "actions": ["get_employees", "get_directory", "get_time_off"],
    },
    "greenhouse": {
        "name": "Greenhouse", "category": "HR", "icon": "leaf",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Greenhouse → Configure → Dev Center → API Credentials → Create.",
        "actions": ["get_candidates", "get_jobs", "create_candidate"],
    },
    "lever": {
        "name": "Lever", "category": "HR", "icon": "swap-horizontal",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Lever → Settings → Integrations → API → Generate Key.",
        "actions": ["get_opportunities", "get_postings", "create_opportunity"],
    },

    # ── Customer Support ─────────────────────────────────────────────────
    "zendesk": {
        "name": "Zendesk", "category": "Support", "icon": "headset",
        "auth_fields": [{"key": "api_key", "label": "API Token", "secure": True}, {"key": "email", "label": "Email"}, {"key": "subdomain", "label": "Subdomain"}],
        "setup": "Go to Zendesk → Admin → Channels → API → Add Token.",
        "actions": ["get_tickets", "create_ticket", "update_ticket", "search", "get_users"],
    },
    "freshdesk": {
        "name": "Freshdesk", "category": "Support", "icon": "help-circle",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}, {"key": "domain", "label": "Domain"}],
        "setup": "Go to Freshdesk → Profile → Your API Key.",
        "actions": ["get_tickets", "create_ticket", "update_ticket", "search"],
    },
    "helpscout": {
        "name": "HelpScout", "category": "Support", "icon": "help-buoy",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to HelpScout → Your Profile → API Keys → Generate.",
        "actions": ["get_conversations", "create_conversation", "get_customers"],
    },
    "livechat": {
        "name": "LiveChat", "category": "Support", "icon": "chatbubble-ellipses",
        "auth_fields": [{"key": "api_key", "label": "Personal Access Token", "secure": True}],
        "setup": "Go to LiveChat → Settings → Integrations → Personal Access Tokens → Create.",
        "actions": ["get_chats", "get_agents", "send_message"],
    },

    # ── Legal & Signatures ───────────────────────────────────────────────
    "docusign": {
        "name": "DocuSign", "category": "Legal", "icon": "create",
        "auth_fields": [{"key": "api_key", "label": "Access Token", "secure": True}, {"key": "account_id", "label": "Account ID"}],
        "setup": "Go to developers.docusign.com → My Account → API → Create Key and Access Token.",
        "actions": ["send_envelope", "get_envelopes", "get_envelope_status"],
    },
    "hellosign": {
        "name": "HelloSign", "category": "Legal", "icon": "pencil",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to HelloSign → Settings → API → Copy API Key.",
        "actions": ["send_signature_request", "get_signature_requests"],
    },
    "pandadoc": {
        "name": "PandaDoc", "category": "Legal", "icon": "document-attach",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to PandaDoc → Settings → Integrations → API → Create Key.",
        "actions": ["get_documents", "create_document", "send_document"],
    },
    "contractsafe": {
        "name": "ContractSafe", "category": "Legal", "icon": "shield-checkmark",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to ContractSafe → Settings → API → Generate Key.",
        "actions": ["get_contracts", "search"],
    },

    # ── Social Media ─────────────────────────────────────────────────────
    "instagram": {
        "name": "Instagram", "category": "Social Media", "icon": "logo-instagram",
        "auth_fields": [{"key": "api_key", "label": "Access Token", "secure": True}],
        "setup": "Go to developers.facebook.com → Create App → Instagram Graph API → Generate Token.",
        "actions": ["get_posts", "create_post", "get_insights"],
    },
    "facebook_pages": {
        "name": "Facebook Pages", "category": "Social Media", "icon": "logo-facebook",
        "auth_fields": [{"key": "api_key", "label": "Page Access Token", "secure": True}, {"key": "page_id", "label": "Page ID"}],
        "setup": "Go to developers.facebook.com → Graph API Explorer → Get Page Token.",
        "actions": ["get_posts", "create_post", "get_insights"],
    },
    "twitter": {
        "name": "Twitter / X", "category": "Social Media", "icon": "logo-twitter",
        "auth_fields": [{"key": "api_key", "label": "Bearer Token", "secure": True}],
        "setup": "Go to developer.twitter.com → Projects & Apps → Keys and Tokens → Generate Bearer Token.",
        "actions": ["get_tweets", "create_tweet", "search"],
    },
    "linkedin": {
        "name": "LinkedIn", "category": "Social Media", "icon": "logo-linkedin",
        "auth_fields": [{"key": "api_key", "label": "Access Token", "secure": True}],
        "setup": "Go to linkedin.com/developers → Create App → Auth → Generate Token.",
        "actions": ["get_profile", "create_post", "get_connections"],
    },
    "tiktok": {
        "name": "TikTok", "category": "Social Media", "icon": "musical-notes",
        "auth_fields": [{"key": "api_key", "label": "Access Token", "secure": True}],
        "setup": "Go to developers.tiktok.com → My Apps → Create → Get Access Token.",
        "actions": ["get_videos", "get_user_info"],
    },
    "buffer": {
        "name": "Buffer", "category": "Social Media", "icon": "share-social",
        "auth_fields": [{"key": "api_key", "label": "Access Token", "secure": True}],
        "setup": "Go to buffer.com/developers → My Apps → Create → Copy Access Token.",
        "actions": ["get_profiles", "create_update", "get_updates"],
    },
    "hootsuite": {
        "name": "Hootsuite", "category": "Social Media", "icon": "planet",
        "auth_fields": [{"key": "api_key", "label": "Access Token", "secure": True}],
        "setup": "Go to developer.hootsuite.com → My Apps → Create → Copy Access Token.",
        "actions": ["get_social_profiles", "schedule_message", "get_messages"],
    },

    # ── Healthcare ───────────────────────────────────────────────────────
    "athenahealth": {
        "name": "Athenahealth", "category": "Healthcare", "icon": "medkit",
        "auth_fields": [{"key": "client_id", "label": "Client ID"}, {"key": "client_secret", "label": "Client Secret", "secure": True}],
        "setup": "Go to developer.athenahealth.com → My Apps → Create → Copy Client ID and Secret.",
        "actions": ["get_patients", "get_appointments", "create_appointment"],
    },
    "drchrono": {
        "name": "DrChrono", "category": "Healthcare", "icon": "fitness",
        "auth_fields": [{"key": "api_key", "label": "Access Token", "secure": True}],
        "setup": "Go to drchrono.com → API → Generate Access Token.",
        "actions": ["get_patients", "get_appointments", "create_appointment"],
    },
    "simplepractice": {
        "name": "SimplePractice", "category": "Healthcare", "icon": "heart",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to SimplePractice → Settings → Integrations → API → Generate Key.",
        "actions": ["get_clients", "get_appointments", "create_appointment"],
    },

    # ── Automation ───────────────────────────────────────────────────────
    "zapier": {
        "name": "Zapier Webhooks", "category": "Automation", "icon": "flash",
        "auth_fields": [{"key": "webhook_url", "label": "Webhook URL"}],
        "setup": "Go to zapier.com → Create Zap → Trigger: Webhooks by Zapier → Catch Hook → Copy the webhook URL. This lets GoFarther AI trigger any of your 5,000+ Zapier integrations.",
        "actions": ["trigger_webhook"],
    },
    "make": {
        "name": "Make (Integromat)", "category": "Automation", "icon": "construct",
        "auth_fields": [{"key": "webhook_url", "label": "Webhook URL"}],
        "setup": "Go to make.com → Create Scenario → Add Webhook module → Copy URL.",
        "actions": ["trigger_webhook"],
    },
    "n8n": {
        "name": "n8n", "category": "Automation", "icon": "git-network",
        "auth_fields": [{"key": "webhook_url", "label": "Webhook URL"}],
        "setup": "Go to n8n → Create Workflow → Add Webhook node → Copy production URL.",
        "actions": ["trigger_webhook"],
    },
    "ifttt": {
        "name": "IFTTT", "category": "Automation", "icon": "link",
        "auth_fields": [{"key": "webhook_url", "label": "Webhook URL"}, {"key": "event_name", "label": "Event Name"}],
        "setup": "Go to ifttt.com/maker_webhooks → Settings → Copy URL. Create an Applet with Webhooks trigger.",
        "actions": ["trigger_webhook"],
    },

    # ── Finance & Banking ───────────────────────────────────────────────
    "brex": {
        "name": "Brex", "category": "Finance", "icon": "card",
        "auth_fields": [{"key": "api_key", "label": "API Token", "secure": True}],
        "setup": "Go to Brex → Settings → Developer → API Keys → Create Token.",
        "actions": ["get_transactions", "get_accounts", "create_expense"],
    },
    "mercury": {
        "name": "Mercury", "category": "Finance", "icon": "trending-up",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Mercury → Settings → API → Generate API Key.",
        "actions": ["get_transactions", "get_accounts", "get_balance"],
    },
    "ramp": {
        "name": "Ramp", "category": "Finance", "icon": "card",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Ramp → Settings → Developer → API Keys → Create.",
        "actions": ["get_transactions", "get_cards", "get_expenses"],
    },
    "wise": {
        "name": "Wise Business", "category": "Finance", "icon": "swap-horizontal",
        "auth_fields": [{"key": "api_key", "label": "API Token", "secure": True}],
        "setup": "Go to Wise → Settings → API Tokens → Create Token with full access.",
        "actions": ["get_transactions", "get_balance", "create_transfer"],
    },

    # ── Real Estate ─────────────────────────────────────────────────────
    "propertybase": {
        "name": "Propertybase", "category": "Real Estate", "icon": "home",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Propertybase → Settings → API → Generate Key.",
        "actions": ["get_listings", "get_leads", "create_lead", "search"],
    },
    "boomtown": {
        "name": "BoomTown", "category": "Real Estate", "icon": "business",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to BoomTown → Settings → API → Copy Key.",
        "actions": ["get_leads", "create_lead", "search"],
    },

    # ── Legal ───────────────────────────────────────────────────────────
    "clio": {
        "name": "Clio", "category": "Legal", "icon": "briefcase",
        "auth_fields": [{"key": "api_key", "label": "Access Token", "secure": True}],
        "setup": "Go to app.clio.com → Settings → API → Create App → Copy Access Token.",
        "actions": ["get_matters", "create_matter", "get_contacts", "get_tasks", "create_task"],
    },
    "lawpay": {
        "name": "LawPay", "category": "Legal", "icon": "cash",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to LawPay → Settings → API → Generate Key.",
        "actions": ["get_transactions", "create_payment_link", "get_invoices"],
    },
    "mycase": {
        "name": "MyCase", "category": "Legal", "icon": "folder",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to MyCase → Settings → Integrations → API → Copy Key.",
        "actions": ["get_cases", "create_case", "get_contacts", "get_tasks"],
    },
    "practicepanther": {
        "name": "PracticePanther", "category": "Legal", "icon": "shield",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to PracticePanther → Settings → API → Generate Key.",
        "actions": ["get_matters", "get_contacts", "get_invoices", "create_task"],
    },

    # ── Education ───────────────────────────────────────────────────────
    "canvas_lms": {
        "name": "Canvas LMS", "category": "Education", "icon": "school",
        "auth_fields": [{"key": "api_key", "label": "Access Token", "secure": True}, {"key": "domain", "label": "Canvas Domain (e.g. school.instructure.com)"}],
        "setup": "Go to Canvas → Account → Settings → New Access Token → Generate.",
        "actions": ["get_courses", "get_assignments", "get_grades", "create_assignment"],
    },
    "google_classroom": {
        "name": "Google Classroom", "category": "Education", "icon": "school",
        "auth_fields": [{"key": "api_key", "label": "OAuth Token", "secure": True}],
        "setup": "Go to console.cloud.google.com → Enable Classroom API → Create OAuth credentials.",
        "actions": ["get_courses", "get_assignments", "create_assignment", "get_students"],
    },

    # ── Restaurants & POS ───────────────────────────────────────────────
    "toast": {
        "name": "Toast POS", "category": "POS", "icon": "restaurant",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}, {"key": "restaurant_id", "label": "Restaurant GUID"}],
        "setup": "Go to Toast → Developer Portal → Create App → Copy API Key and Restaurant GUID.",
        "actions": ["get_orders", "get_menu", "get_employees", "get_revenue"],
    },
    "clover": {
        "name": "Clover", "category": "POS", "icon": "card",
        "auth_fields": [{"key": "api_key", "label": "API Token", "secure": True}, {"key": "merchant_id", "label": "Merchant ID"}],
        "setup": "Go to Clover → Developer Dashboard → Create App → Copy Token and Merchant ID.",
        "actions": ["get_orders", "get_inventory", "get_employees", "get_revenue"],
    },
    "lightspeed": {
        "name": "Lightspeed", "category": "POS", "icon": "flash",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Lightspeed → Settings → API → Generate Key.",
        "actions": ["get_sales", "get_inventory", "get_customers", "get_products"],
    },

    # ── Field Service ───────────────────────────────────────────────────
    "servicetitan": {
        "name": "ServiceTitan", "category": "Field Service", "icon": "construct",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}, {"key": "tenant_id", "label": "Tenant ID"}],
        "setup": "Go to ServiceTitan → Settings → Integrations → API → Copy Key and Tenant ID.",
        "actions": ["get_jobs", "create_job", "get_customers", "get_invoices", "schedule_job"],
    },
    "jobber": {
        "name": "Jobber", "category": "Field Service", "icon": "hammer",
        "auth_fields": [{"key": "api_key", "label": "API Token", "secure": True}],
        "setup": "Go to Jobber → Settings → API → Generate Token.",
        "actions": ["get_jobs", "create_job", "get_clients", "get_invoices", "schedule_visit"],
    },
    "housecall_pro": {
        "name": "Housecall Pro", "category": "Field Service", "icon": "home",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Housecall Pro → Settings → API → Generate Key.",
        "actions": ["get_jobs", "create_job", "get_customers", "get_invoices"],
    },

    # ── Logistics & Shipping ────────────────────────────────────────────
    "shipstation": {
        "name": "ShipStation", "category": "Logistics", "icon": "boat",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}, {"key": "api_secret", "label": "API Secret", "secure": True}],
        "setup": "Go to ShipStation → Settings → Account → API Settings → Generate Keys.",
        "actions": ["get_orders", "create_label", "get_shipments", "track_package"],
    },
    "shippo": {
        "name": "Shippo", "category": "Logistics", "icon": "airplane",
        "auth_fields": [{"key": "api_key", "label": "API Token", "secure": True}],
        "setup": "Go to Shippo → Settings → API → Copy Live Token.",
        "actions": ["get_shipments", "create_shipment", "get_rates", "create_label"],
    },
    "easypost": {
        "name": "EasyPost", "category": "Logistics", "icon": "cube",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to easypost.com → Account → API Keys → Copy Production Key.",
        "actions": ["create_shipment", "get_rates", "create_label", "track_package"],
    },

    # ── Design ──────────────────────────────────────────────────────────
    "figma": {
        "name": "Figma", "category": "Design", "icon": "color-palette",
        "auth_fields": [{"key": "api_key", "label": "Personal Access Token", "secure": True}],
        "setup": "Go to Figma → Settings → Account → Personal Access Tokens → Create.",
        "actions": ["get_files", "get_projects", "get_comments", "export_image"],
    },
    "canva": {
        "name": "Canva", "category": "Design", "icon": "brush",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to canva.com/developers → Create App → Copy API Key.",
        "actions": ["get_designs", "create_design", "export_design"],
    },

    # ── Analytics ───────────────────────────────────────────────────────
    "google_analytics": {
        "name": "Google Analytics", "category": "Analytics", "icon": "analytics",
        "auth_fields": [{"key": "api_key", "label": "Service Account JSON Key", "secure": True}, {"key": "property_id", "label": "Property ID"}],
        "setup": "Go to console.cloud.google.com → Enable Analytics API → Create Service Account → Download JSON key.",
        "actions": ["get_report", "get_realtime", "get_users", "get_pageviews"],
    },
    "mixpanel": {
        "name": "Mixpanel", "category": "Analytics", "icon": "bar-chart",
        "auth_fields": [{"key": "api_key", "label": "Service Account", "secure": True}, {"key": "project_id", "label": "Project ID"}],
        "setup": "Go to Mixpanel → Settings → Project Settings → Service Accounts → Create.",
        "actions": ["get_events", "get_funnels", "get_retention", "get_users"],
    },
    "segment": {
        "name": "Segment", "category": "Analytics", "icon": "pie-chart",
        "auth_fields": [{"key": "api_key", "label": "Write Key", "secure": True}],
        "setup": "Go to Segment → Sources → Your Source → Settings → API Keys → Copy Write Key.",
        "actions": ["track_event", "identify_user", "get_events"],
    },

    # ── Dev Tools ───────────────────────────────────────────────────────
    "github": {
        "name": "GitHub", "category": "Dev Tools", "icon": "logo-github",
        "auth_fields": [{"key": "api_key", "label": "Personal Access Token", "secure": True}],
        "setup": "Go to GitHub → Settings → Developer Settings → Personal Access Tokens → Generate.",
        "actions": ["get_repos", "get_issues", "create_issue", "get_prs", "search"],
    },
    "gitlab": {
        "name": "GitLab", "category": "Dev Tools", "icon": "git-branch",
        "auth_fields": [{"key": "api_key", "label": "Personal Access Token", "secure": True}],
        "setup": "Go to GitLab → Settings → Access Tokens → Create.",
        "actions": ["get_projects", "get_issues", "create_issue", "get_pipelines"],
    },
    "vercel": {
        "name": "Vercel", "category": "Dev Tools", "icon": "triangle",
        "auth_fields": [{"key": "api_key", "label": "API Token", "secure": True}],
        "setup": "Go to Vercel → Settings → Tokens → Create.",
        "actions": ["get_deployments", "get_projects", "trigger_deploy"],
    },

    # ── Video ───────────────────────────────────────────────────────────
    "youtube": {
        "name": "YouTube", "category": "Video", "icon": "logo-youtube",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to console.cloud.google.com → Enable YouTube Data API → Create API Key.",
        "actions": ["get_videos", "get_analytics", "get_comments", "search"],
    },
    "loom": {
        "name": "Loom", "category": "Video", "icon": "videocam",
        "auth_fields": [{"key": "api_key", "label": "Developer Token", "secure": True}],
        "setup": "Go to loom.com → Settings → Developer → Generate Token.",
        "actions": ["get_videos", "get_shared_videos"],
    },

    # ── Surveys & Forms ─────────────────────────────────────────────────
    "typeform": {
        "name": "Typeform", "category": "Surveys", "icon": "list",
        "auth_fields": [{"key": "api_key", "label": "Personal Access Token", "secure": True}],
        "setup": "Go to Typeform → Settings → Personal Tokens → Generate.",
        "actions": ["get_forms", "get_responses", "create_form"],
    },
    "surveymonkey": {
        "name": "SurveyMonkey", "category": "Surveys", "icon": "clipboard",
        "auth_fields": [{"key": "api_key", "label": "Access Token", "secure": True}],
        "setup": "Go to developer.surveymonkey.com → My Apps → Create → Copy Token.",
        "actions": ["get_surveys", "get_responses", "create_survey"],
    },
    "tally": {
        "name": "Tally", "category": "Surveys", "icon": "checkmark-done",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to tally.so → Settings → Integrations → API → Copy Key.",
        "actions": ["get_forms", "get_submissions"],
    },

    # ── Appointments ────────────────────────────────────────────────────
    "vagaro": {
        "name": "Vagaro", "category": "Appointments", "icon": "calendar",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Vagaro → Settings → Developer → API → Generate Key.",
        "actions": ["get_appointments", "create_appointment", "get_clients", "get_services"],
    },
    "mindbody": {
        "name": "Mindbody", "category": "Appointments", "icon": "fitness",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}, {"key": "site_id", "label": "Site ID"}],
        "setup": "Go to Mindbody → Developer Portal → Create App → Copy API Key and Site ID.",
        "actions": ["get_classes", "get_appointments", "get_clients", "book_class"],
    },
    "fresha": {
        "name": "Fresha", "category": "Appointments", "icon": "cut",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Fresha → Settings → Integrations → API → Generate Key.",
        "actions": ["get_appointments", "get_clients", "get_services"],
    },
    "booksy": {
        "name": "Booksy", "category": "Appointments", "icon": "time",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Booksy → Settings → Integrations → API → Copy Key.",
        "actions": ["get_appointments", "get_clients", "get_services"],
    },

    # ── Insurance ───────────────────────────────────────────────────────
    "applied_epic": {
        "name": "Applied Epic", "category": "Insurance", "icon": "shield-checkmark",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Applied Epic → Admin → API → Generate Key.",
        "actions": ["get_policies", "get_clients", "create_client", "search"],
    },
    "hawksoft": {
        "name": "HawkSoft", "category": "Insurance", "icon": "shield",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to HawkSoft → Settings → API → Generate Key.",
        "actions": ["get_policies", "get_clients", "search"],
    },
    "ezlynx": {
        "name": "EZLynx", "category": "Insurance", "icon": "document-text",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to EZLynx → Settings → API → Copy Key.",
        "actions": ["get_policies", "get_clients", "create_quote", "search"],
    },
    "agency_zoom": {
        "name": "Agency Zoom", "category": "Insurance", "icon": "trending-up",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Agency Zoom → Settings → Integrations → API → Generate Key.",
        "actions": ["get_leads", "create_lead", "get_policies", "search"],
    },
    "better_agency": {
        "name": "Better Agency", "category": "Insurance", "icon": "people",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Better Agency → Settings → API → Copy Key.",
        "actions": ["get_leads", "create_lead", "get_clients", "search"],
    },

    # ── Construction ────────────────────────────────────────────────────
    "procore": {
        "name": "Procore", "category": "Construction", "icon": "build",
        "auth_fields": [{"key": "api_key", "label": "Access Token", "secure": True}],
        "setup": "Go to developers.procore.com → Create App → Generate OAuth Token.",
        "actions": ["get_projects", "get_rfis", "create_rfi", "get_submittals", "search"],
    },
    "buildertrend": {
        "name": "Buildertrend", "category": "Construction", "icon": "hammer",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Buildertrend → Settings → API → Generate Key.",
        "actions": ["get_projects", "get_schedules", "get_financials", "search"],
    },
    "coconstruct": {
        "name": "CoConstruct", "category": "Construction", "icon": "home",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to CoConstruct → Settings → API → Copy Key.",
        "actions": ["get_projects", "get_clients", "get_financials"],
    },
    "plangrid": {
        "name": "PlanGrid", "category": "Construction", "icon": "map",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to PlanGrid → Settings → API → Generate Key.",
        "actions": ["get_projects", "get_sheets", "get_issues", "search"],
    },

    # ── Automotive ──────────────────────────────────────────────────────
    "dealersocket": {
        "name": "DealerSocket", "category": "Automotive", "icon": "car",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to DealerSocket → Admin → API → Generate Key.",
        "actions": ["get_leads", "create_lead", "get_inventory", "search"],
    },
    "vinsolutions": {
        "name": "VinSolutions", "category": "Automotive", "icon": "car-sport",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to VinSolutions → Settings → API → Copy Key.",
        "actions": ["get_leads", "create_lead", "get_inventory", "search"],
    },

    # ── Nonprofit ───────────────────────────────────────────────────────
    "bloomerang": {
        "name": "Bloomerang", "category": "Nonprofit", "icon": "heart",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Bloomerang → Settings → API → Copy Key.",
        "actions": ["get_donors", "create_donor", "get_donations", "search"],
    },
    "donorperfect": {
        "name": "DonorPerfect", "category": "Nonprofit", "icon": "gift",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to DonorPerfect → Admin → API → Generate Key.",
        "actions": ["get_donors", "create_donor", "get_donations", "search"],
    },
    "givebutter": {
        "name": "Givebutter", "category": "Nonprofit", "icon": "heart-circle",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Givebutter → Settings → API → Copy Key.",
        "actions": ["get_campaigns", "get_donations", "get_supporters"],
    },

    # ── Hospitality ─────────────────────────────────────────────────────
    "guesty": {
        "name": "Guesty", "category": "Hospitality", "icon": "bed",
        "auth_fields": [{"key": "api_key", "label": "API Token", "secure": True}],
        "setup": "Go to Guesty → Marketplace → Open API → Generate Token.",
        "actions": ["get_listings", "get_reservations", "get_guests", "search"],
    },
    "hostaway": {
        "name": "Hostaway", "category": "Hospitality", "icon": "key",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}, {"key": "account_id", "label": "Account ID"}],
        "setup": "Go to Hostaway → Settings → API → Copy Key and Account ID.",
        "actions": ["get_listings", "get_reservations", "get_guests"],
    },
    "cloudbeds": {
        "name": "Cloudbeds", "category": "Hospitality", "icon": "business",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Cloudbeds → Settings → API → Generate Key.",
        "actions": ["get_reservations", "get_rooms", "get_guests", "search"],
    },

    # ── Fitness ─────────────────────────────────────────────────────────
    "gymmaster": {
        "name": "GymMaster", "category": "Fitness", "icon": "barbell",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to GymMaster → Settings → API → Generate Key.",
        "actions": ["get_members", "get_classes", "get_bookings"],
    },
    "glofox": {
        "name": "Glofox", "category": "Fitness", "icon": "fitness",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Glofox → Settings → API → Copy Key.",
        "actions": ["get_members", "get_classes", "get_bookings"],
    },
    "wellnessliving": {
        "name": "WellnessLiving", "category": "Fitness", "icon": "body",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to WellnessLiving → Settings → Developer → API → Generate Key.",
        "actions": ["get_clients", "get_classes", "get_appointments", "get_bookings"],
    },

    # ── Dental ──────────────────────────────────────────────────────────
    "dentrix": {
        "name": "Dentrix", "category": "Dental", "icon": "medkit",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Dentrix → Settings → Developer API → Generate Key.",
        "actions": ["get_patients", "get_appointments", "get_procedures", "search"],
    },
    "open_dental": {
        "name": "Open Dental", "category": "Dental", "icon": "medical",
        "auth_fields": [{"key": "api_key", "label": "Developer Key", "secure": True}],
        "setup": "Go to Open Dental → Setup → API → Enable and copy Developer Key.",
        "actions": ["get_patients", "get_appointments", "create_appointment", "search"],
    },
    "curve_dental": {
        "name": "Curve Dental", "category": "Dental", "icon": "pulse",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to Curve Dental → Admin → API → Generate Key.",
        "actions": ["get_patients", "get_appointments", "get_procedures"],
    },

    # ── Government ──────────────────────────────────────────────────────
    "govpilot": {
        "name": "GovPilot", "category": "Government", "icon": "flag",
        "auth_fields": [{"key": "api_key", "label": "API Key", "secure": True}],
        "setup": "Go to GovPilot → Admin → API → Generate Key.",
        "actions": ["get_permits", "get_requests", "create_request", "search"],
    },
    "accela": {
        "name": "Accela", "category": "Government", "icon": "document",
        "auth_fields": [{"key": "api_key", "label": "App ID", "secure": True}, {"key": "api_secret", "label": "App Secret", "secure": True}],
        "setup": "Go to developer.accela.com → Create App → Copy App ID and Secret.",
        "actions": ["get_records", "create_record", "get_inspections", "search"],
    },

    # ── Oracle Products ──────────────────────────────────────────────────
    "oracle_netsuite": {
        "name": "Oracle NetSuite", "category": "ERP", "icon": "server",
        "auth_fields": [
            {"key": "account_id", "label": "Account ID", "secure": False},
            {"key": "consumer_key", "label": "Consumer Key", "secure": True},
            {"key": "consumer_secret", "label": "Consumer Secret", "secure": True},
            {"key": "token_id", "label": "Token ID", "secure": True},
            {"key": "token_secret", "label": "Token Secret", "secure": True},
        ],
        "setup": "Go to NetSuite → Setup → Integration → Manage Integrations → New. Enable Token-Based Auth. Create an access token under Setup → Users/Roles → Access Tokens.",
        "actions": ["get_customers", "create_customer", "get_invoices", "create_invoice", "get_sales_orders", "create_sales_order", "get_items", "get_vendors", "get_purchase_orders", "get_journal_entries", "search"],
    },
    "oracle_cloud_erp": {
        "name": "Oracle Cloud ERP", "category": "ERP", "icon": "cloud",
        "auth_fields": [
            {"key": "base_url", "label": "Cloud ERP URL", "secure": False},
            {"key": "username", "label": "Username", "secure": False},
            {"key": "password", "label": "Password", "secure": True},
        ],
        "setup": "Go to Oracle Cloud → Navigator → Tools → Security Console. Create a user with API access. Your base URL is like https://servername.fa.us2.oraclecloud.com.",
        "actions": ["get_invoices", "create_invoice", "get_purchase_orders", "get_journal_entries", "get_general_ledger", "get_payables", "get_receivables", "get_assets", "search"],
    },
    "oracle_cx_sales": {
        "name": "Oracle CX Sales", "category": "CRM", "icon": "people",
        "auth_fields": [
            {"key": "base_url", "label": "CX Sales URL", "secure": False},
            {"key": "username", "label": "Username", "secure": False},
            {"key": "password", "label": "Password", "secure": True},
        ],
        "setup": "Go to Oracle CX Sales → Navigator → Tools → Manage Users. Create a user with REST API access. Base URL is like https://servername.cx.us2.oraclecloud.com.",
        "actions": ["get_accounts", "create_account", "get_contacts", "create_contact", "get_opportunities", "create_opportunity", "get_leads", "get_activities", "search"],
    },
    "oracle_cx_service": {
        "name": "Oracle CX Service", "category": "Support", "icon": "headset",
        "auth_fields": [
            {"key": "base_url", "label": "CX Service URL", "secure": False},
            {"key": "username", "label": "Username", "secure": False},
            {"key": "password", "label": "Password", "secure": True},
        ],
        "setup": "Go to Oracle CX Service (B2C Service) → Configuration → Staff Management → Profiles. Enable REST API access for your profile.",
        "actions": ["get_incidents", "create_incident", "get_contacts", "create_contact", "get_service_requests", "update_incident", "search"],
    },
    "oracle_hcm": {
        "name": "Oracle Cloud HCM", "category": "HR", "icon": "person",
        "auth_fields": [
            {"key": "base_url", "label": "HCM Cloud URL", "secure": False},
            {"key": "username", "label": "Username", "secure": False},
            {"key": "password", "label": "Password", "secure": True},
        ],
        "setup": "Go to Oracle HCM Cloud → Navigator → Tools → Security Console. Create a user with HCM REST API access.",
        "actions": ["get_employees", "get_departments", "get_jobs", "get_absences", "get_payroll", "get_benefits", "get_compensation", "search"],
    },
    "oracle_epm": {
        "name": "Oracle EPM Cloud", "category": "Finance", "icon": "calculator",
        "auth_fields": [
            {"key": "base_url", "label": "EPM Cloud URL", "secure": False},
            {"key": "username", "label": "Username", "secure": False},
            {"key": "password", "label": "Password", "secure": True},
        ],
        "setup": "Go to Oracle EPM Cloud → Navigator → Tools → Access Control. Assign the Service Administrator or appropriate role for REST API access.",
        "actions": ["get_planning_data", "submit_data", "run_business_rule", "get_reports", "get_consolidation", "get_budgets", "search"],
    },
    "oracle_scm": {
        "name": "Oracle SCM Cloud", "category": "Logistics", "icon": "cube",
        "auth_fields": [
            {"key": "base_url", "label": "SCM Cloud URL", "secure": False},
            {"key": "username", "label": "Username", "secure": False},
            {"key": "password", "label": "Password", "secure": True},
        ],
        "setup": "Go to Oracle SCM Cloud → Navigator → Tools → Security Console. Create a user with SCM REST API privileges.",
        "actions": ["get_inventory", "get_shipments", "get_purchase_orders", "get_work_orders", "get_suppliers", "get_warehouses", "search"],
    },
    "oracle_apex": {
        "name": "Oracle APEX", "category": "Dev Tools", "icon": "code",
        "auth_fields": [
            {"key": "base_url", "label": "APEX Workspace URL", "secure": False},
            {"key": "api_key", "label": "API Key", "secure": True},
        ],
        "setup": "Go to Oracle APEX → SQL Workshop → RESTful Services → Register Schema. Create a module and template to expose REST endpoints.",
        "actions": ["run_query", "get_tables", "execute_procedure", "get_apps", "search"],
    },
    "oracle_analytics": {
        "name": "Oracle Analytics Cloud", "category": "Analytics", "icon": "bar-chart",
        "auth_fields": [
            {"key": "base_url", "label": "Analytics Cloud URL", "secure": False},
            {"key": "username", "label": "Username", "secure": False},
            {"key": "password", "label": "Password", "secure": True},
        ],
        "setup": "Go to Oracle Analytics Cloud → Console → Users. Ensure your account has BI Service Administrator or BI Consumer role for API access.",
        "actions": ["get_reports", "get_dashboards", "run_analysis", "get_datasets", "search"],
    },
    "oracle_commerce": {
        "name": "Oracle Commerce Cloud", "category": "E-commerce", "icon": "cart",
        "auth_fields": [
            {"key": "base_url", "label": "Commerce Cloud URL", "secure": False},
            {"key": "app_key", "label": "Application Key", "secure": True},
        ],
        "setup": "Go to Oracle Commerce Cloud → Settings → Web APIs → Register Application. Copy the Application Key.",
        "actions": ["get_products", "get_orders", "get_customers", "get_inventory", "get_collections", "create_product", "search"],
    },
}


# ── IMAP email provider presets (injected into APP_REGISTRY at init) ────
#
# Each tuple is (app_id, name, setup_help, description). All presets share
# the same auth_fields (email + app_password) and expose the full mail
# action surface because they all delegate to _imap_mail_adapter with the
# hosts pre-injected. Adding a new provider is a one-liner here — no need
# to copy-paste a 50-line registry entry.

_MAIL_PRESETS = [
    # (app_id, display name, setup help, setup_url)
    # The actual adapter functions are wired in the ADAPTERS map near
    # the bottom of this file — they don't need to be referenced here,
    # and referencing them at module import time would crash because
    # those functions are defined much later in the file.
    ("yahoo_mail",       "Yahoo Mail",              "Tap 'Get App Password' to open Yahoo's security page, turn on 2-step verification, then generate an app password for 'GoFarther' and paste it below.", "https://login.yahoo.com/account/security"),
    ("icloud_mail",      "iCloud Mail",             "Tap 'Get App Password' to open Apple ID, then go to Sign-In and Security → App-Specific Passwords and create one named 'GoFarther'.", "https://appleid.apple.com/account/manage"),
    ("zoho_mail",        "Zoho Mail",               "Tap 'Get App Password' to open Zoho's security page and create an app password named 'GoFarther'.", "https://accounts.zoho.com/home#security/app_password"),
    ("fastmail_mail",    "Fastmail",                "Tap 'Get App Password' to open Fastmail's password settings and create an app password with Mail access.", "https://app.fastmail.com/settings/security/integrations/apppassword/new"),
    ("aol_mail",         "AOL Mail",                "Tap 'Get App Password' to open AOL's account security page and generate an app password for 'GoFarther'.", "https://login.aol.com/account/security"),
    ("gmx_mail",         "GMX Mail",                "Tap 'Get App Password' to open GMX settings. Enable POP3/IMAP access, then use your regular GMX password below.", "https://www.gmx.com/mail/customer-center/"),
    ("mailru_mail",      "Mail.ru",                 "Tap 'Get App Password' to open Mail.ru security settings and create a password for external apps.", "https://account.mail.ru/user/2-step-auth/passwords"),
    ("yandex_mail",      "Yandex Mail",             "Tap 'Get App Password' to open Yandex ID security and create an app password for mail.", "https://id.yandex.com/security/app-passwords"),
    ("protonmail_mail",  "ProtonMail (Bridge)",     "ProtonMail only supports third-party apps via Proton Bridge. Tap 'Open Proton Bridge' to download it, then use the Bridge-generated username and password below.", "https://proton.me/mail/bridge"),
    ("hostinger_mail",   "Hostinger Email",         "Tap 'Open hPanel' to log in. Use your regular email password — Hostinger business email works with normal credentials.", "https://hpanel.hostinger.com/email"),
    ("godaddy_mail",     "GoDaddy Workspace",       "Tap 'Open GoDaddy' to log in. Use your regular Workspace Email password.", "https://sso.godaddy.com/v1/login?app=email"),
    ("namecheap_mail",   "Namecheap Private Email", "Tap 'Open Namecheap' to log in to Private Email. Use your regular mailbox password.", "https://privateemail.com/appsuite/"),
    ("ionos_mail",       "IONOS Email",             "Tap 'Open IONOS' to log in. Use your regular IONOS mailbox password.", "https://login.ionos.com/"),
    ("mailboxorg_mail",  "Mailbox.org",             "Tap 'Open Mailbox.org' to log in. Use your regular password — IMAP/SMTP works with normal credentials.", "https://login.mailbox.org/"),
    ("posteo_mail",      "Posteo",                  "Tap 'Open Posteo' to log in. Use your regular Posteo password.", "https://posteo.de/en/my-account"),
    ("mailfence_mail",   "Mailfence",               "Tap 'Open Mailfence' to log in and generate an app password in Account → Security.", "https://mailfence.com/flow/#/login"),
]

_MAIL_PRESET_ACTIONS = [
    "list_inbox", "search_emails", "read_email", "reply_to_email",
    "send_email", "mark_read", "mark_unread", "archive", "delete",
    "move_to_folder", "list_folders", "download_attachment",
]

_MAIL_PRESET_HINTS = {
    "list_inbox": "folder=<folder name, defaults to INBOX>|limit=<1-50, default 20>",
    "search_emails": "query=<text>|from=<email>|subject=<text>|unread=<true/false>|limit=<1-50>",
    "read_email": "message_id=<uid>|folder=<folder name, defaults to INBOX>",
    "reply_to_email": "message_id=<uid>|body=<html>|folder=<folder name>",
    "send_email": "to=<email(s)>|subject=<subject>|body=<html>|cc=<optional>|bcc=<optional>",
    "mark_read": "message_id=<uid>|folder=<folder name>",
    "mark_unread": "message_id=<uid>|folder=<folder name>",
    "archive": "message_id=<uid>|folder=<source folder>",
    "delete": "message_id=<uid>|folder=<folder name>",
    "move_to_folder": "message_id=<uid>|folder=<source>|to=<destination folder>",
    "list_folders": "no params",
    "download_attachment": "message_id=<uid>|folder=<folder>|attachment_index=<0-based index from read_email>",
}

for _app_id, _name, _setup, _setup_url in _MAIL_PRESETS:
    APP_REGISTRY[_app_id] = {
        "name": _name,
        "category": "Email",
        "icon": "mail",
        "auth_fields": [
            {"key": "username", "label": "Email Address"},
            {"key": "app_password", "label": "Password", "secure": True},
        ],
        "setup": f"Enter your {_name} address and password. {_setup}",
        # Deep link to the provider's own credential page — the Settings
        # screen renders this as a "Get App Password →" button so the
        # user never has to google "how do I get my <provider> app
        # password".
        "setup_url": _setup_url,
        "actions": list(_MAIL_PRESET_ACTIONS),
        "action_hints": dict(_MAIL_PRESET_HINTS),
    }

# NOTE: the "flip OAuth-capable mail presets to mail_oauth flow when
# env vars are configured" loop moved to right after
# MAIL_OAUTH_PROVIDERS is defined further down the file, because
# Python import ordering forbids referencing MAIL_OAUTH_PROVIDERS here.

# ── Category order for display ───────────────────────────────────────────

CATEGORY_ORDER = [
    "CRM", "ERP", "Accounting", "Finance", "Project Management", "Communication",
    "Email", "Calendar", "E-commerce", "Storage", "Email Marketing",
    "HR", "Support", "Legal", "Social Media",
    "Healthcare", "Dental", "Real Estate", "Insurance", "Construction",
    "Automotive", "Field Service", "POS", "Hospitality", "Fitness",
    "Logistics", "Design", "Analytics", "Dev Tools",
    "Video", "Surveys", "Appointments", "Education",
    "Nonprofit", "Government", "Automation",
]


# ── Pydantic models ──────────────────────────────────────────────────────

class ConnectRequest(BaseModel):
    credentials: dict[str, str]


class ActionRequest(BaseModel):
    action: str
    params: dict[str, Any] = {}


# ── Endpoints ────────────────────────────────────────────────────────────

@router.get("")
async def list_connectors(authorization: str = Header(...), db: AsyncSession = Depends(get_db)):
    """List all available apps with connection status per user."""
    payload = _verify_auth(authorization)
    user_id = payload.get("sub")
    connected_ids = await _get_connected_app_ids(user_id, db)

    result = []
    for app_id, info in APP_REGISTRY.items():
        result.append({
            "id": app_id,
            "name": info["name"],
            "category": info["category"],
            "icon": info["icon"],
            "auth_fields": info["auth_fields"],
            "setup": info["setup"],
            "setup_url": info.get("setup_url"),
            "actions": info["actions"],
            "action_hints": info.get("action_hints", {}),
            "oauth_flow": info.get("oauth_flow"),
            "connected": app_id in connected_ids,
        })
    # Sort connected apps to the top of the list so they're easier to find
    # in the Settings → Connect Apps screen. Within each group (connected /
    # not connected) we preserve the original APP_REGISTRY order, which is
    # already grouped by category.
    result.sort(key=lambda c: (0 if c["connected"] else 1))
    return {"connectors": result, "categories": CATEGORY_ORDER}


@router.post("/{app_id}/connect")
async def connect_app(app_id: str, body: ConnectRequest, authorization: str = Header(...), db: AsyncSession = Depends(get_db)):
    """Save credentials for an app (encrypted at rest)."""
    payload = _verify_auth(authorization)
    user_id = payload.get("sub")

    if app_id not in APP_REGISTRY:
        raise HTTPException(404, f"Unknown app: {app_id}")

    # Validate required fields
    required = [f["key"] for f in APP_REGISTRY[app_id]["auth_fields"]]
    missing = [k for k in required if not body.credentials.get(k)]
    if missing:
        raise HTTPException(400, f"Missing required fields: {', '.join(missing)}")

    await _set_creds(user_id, app_id, body.credentials, db)
    # Audit log
    from routes.ghost_auth import _audit_log
    await _audit_log(db, payload.get("email", ""), "connector_connected", f"Connected app: {app_id}")
    await db.commit()
    return {"status": "connected", "app": app_id}


@router.delete("/{app_id}/disconnect")
async def disconnect_app(app_id: str, authorization: str = Header(...), db: AsyncSession = Depends(get_db)):
    """Remove credentials for an app."""
    payload = _verify_auth(authorization)
    user_id = payload.get("sub")
    await _del_creds(user_id, app_id, db)
    # Audit log
    from routes.ghost_auth import _audit_log
    await _audit_log(db, payload.get("email", ""), "connector_disconnected", f"Disconnected app: {app_id}")
    await db.commit()
    return {"status": "disconnected", "app": app_id}


# ── Microsoft OAuth 2.0 flow (for Excel Online, OneDrive, etc.) ──────────────
# Uses authorization code flow with client secret. The mobile app opens the
# authorize URL in the system browser, Microsoft redirects to our backend
# callback, we exchange the code for tokens and store them encrypted.

def _ms_client_id() -> str:
    # Prefer the dedicated env var, fall back to the Teams bot's App ID since
    # both point at the same Azure AD App Registration in practice.
    return os.getenv("MICROSOFT_CLIENT_ID") or os.getenv("TEAMS_APP_ID") or ""


def _ms_client_secret() -> str:
    return os.getenv("MICROSOFT_CLIENT_SECRET") or os.getenv("TEAMS_APP_PASSWORD") or ""


def _ms_redirect_uri() -> str:
    return os.getenv("MICROSOFT_REDIRECT_URI") or "https://isibi-backend.onrender.com/api/ghost/connectors/oauth/microsoft/callback"


def _ms_tenant() -> str:
    # If the Azure app is single-tenant the tenant GUID is required in the
    # authorize/token URLs. For multi-tenant apps "common" works.
    return os.getenv("MICROSOFT_TENANT_ID") or os.getenv("TEAMS_TENANT_ID") or "common"


# ── Google OAuth (Gmail, Drive, Calendar, etc.) ────────────────────────
def _google_client_id() -> str:
    return os.getenv("GOOGLE_CLIENT_ID") or ""


def _google_client_secret() -> str:
    return os.getenv("GOOGLE_CLIENT_SECRET") or ""


def _google_redirect_uri() -> str:
    return os.getenv("GOOGLE_REDIRECT_URI") or "https://isibi-backend.onrender.com/api/ghost/connectors/oauth/google/callback"


_GOOGLE_AUTH_URL = "https://accounts.google.com/o/oauth2/v2/auth"
_GOOGLE_TOKEN_URL = "https://oauth2.googleapis.com/token"

_GOOGLE_SCOPES = {
    # Full Gmail mailbox access: read, modify, compose, send, delete drafts.
    # The `https://mail.google.com/` scope is the catch-all for IMAP-equivalent
    # access — needed for move_to_folder, permanent delete, etc.
    "gmail": "https://mail.google.com/ https://www.googleapis.com/auth/gmail.modify https://www.googleapis.com/auth/gmail.send openid email",
}


# ── Generic mail OAuth (Yahoo, AOL, Zoho, Fastmail, Mail.ru, Yandex) ────
#
# All of these are IMAP-based providers that ALSO expose OAuth 2.0 for
# third-party developers. We use OAuth to obtain an access_token, then
# authenticate to IMAP + SMTP via XOAUTH2 instead of a password. The user
# never enters credentials into GoFarther — they sign in on the provider's
# own page, exactly like Gmail / Outlook.
#
# Each provider is declared in MAIL_OAUTH_PROVIDERS with its auth/token
# URLs, scopes, env var names, which connector app_ids it powers, and the
# pinned IMAP/SMTP hosts to use. The oauth_start / callback / refresh
# code is fully generic — adding another OAuth-capable provider is just
# one more dict entry.
#
# Yahoo and AOL share the same OAuth backend (both owned by Yahoo), so a
# single dev-app registration covers both.

MAIL_OAUTH_PROVIDERS: dict[str, dict] = {
    "yahoo": {
        "auth_url":  "https://api.login.yahoo.com/oauth2/request_auth",
        "token_url": "https://api.login.yahoo.com/oauth2/get_token",
        "scopes":    "mail-w",  # Yahoo: mail-w grants full mailbox access
        "client_id_env":     "YAHOO_CLIENT_ID",
        "client_secret_env": "YAHOO_CLIENT_SECRET",
        "app_ids":   ["yahoo_mail", "aol_mail"],
        "imap_host": "imap.mail.yahoo.com", "imap_port": 993,
        "smtp_host": "smtp.mail.yahoo.com", "smtp_port": 587,
        "display":   "Yahoo",
    },
    "zoho": {
        "auth_url":  "https://accounts.zoho.com/oauth/v2/auth",
        "token_url": "https://accounts.zoho.com/oauth/v2/token",
        "scopes":    "ZohoMail.messages.ALL ZohoMail.folders.ALL ZohoMail.accounts.READ",
        "client_id_env":     "ZOHO_CLIENT_ID",
        "client_secret_env": "ZOHO_CLIENT_SECRET",
        "app_ids":   ["zoho_mail"],
        "imap_host": "imap.zoho.com", "imap_port": 993,
        "smtp_host": "smtp.zoho.com", "smtp_port": 587,
        "display":   "Zoho",
    },
    "fastmail": {
        "auth_url":  "https://api.fastmail.com/oauth/authorize",
        "token_url": "https://api.fastmail.com/oauth/refresh",
        "scopes":    "urn:ietf:params:jmap:mail urn:ietf:params:jmap:submission",
        "client_id_env":     "FASTMAIL_CLIENT_ID",
        "client_secret_env": "FASTMAIL_CLIENT_SECRET",
        "app_ids":   ["fastmail_mail"],
        "imap_host": "imap.fastmail.com", "imap_port": 993,
        "smtp_host": "smtp.fastmail.com", "smtp_port": 587,
        "display":   "Fastmail",
    },
    "mailru": {
        "auth_url":  "https://oauth.mail.ru/login",
        "token_url": "https://oauth.mail.ru/token",
        "scopes":    "userinfo",
        "client_id_env":     "MAILRU_CLIENT_ID",
        "client_secret_env": "MAILRU_CLIENT_SECRET",
        "app_ids":   ["mailru_mail"],
        "imap_host": "imap.mail.ru", "imap_port": 993,
        "smtp_host": "smtp.mail.ru", "smtp_port": 587,
        "display":   "Mail.ru",
    },
    "yandex": {
        "auth_url":  "https://oauth.yandex.com/authorize",
        "token_url": "https://oauth.yandex.com/token",
        "scopes":    "mail:imap_full mail:smtp",
        "client_id_env":     "YANDEX_CLIENT_ID",
        "client_secret_env": "YANDEX_CLIENT_SECRET",
        "app_ids":   ["yandex_mail"],
        "imap_host": "imap.yandex.com", "imap_port": 993,
        "smtp_host": "smtp.yandex.com", "smtp_port": 587,
        "display":   "Yandex",
    },
}


def _mail_oauth_client(provider_key: str) -> tuple[str, str]:
    """Return (client_id, client_secret) for a mail OAuth provider from
    env vars. Empty strings if not configured — caller should check."""
    cfg = MAIL_OAUTH_PROVIDERS.get(provider_key) or {}
    return (
        os.getenv(cfg.get("client_id_env") or "", ""),
        os.getenv(cfg.get("client_secret_env") or "", ""),
    )


def _mail_oauth_redirect_uri(provider_key: str) -> str:
    """Per-provider redirect URI. Each provider needs its own callback
    path so the client can register it in their dev console separately.
    Override via <PROVIDER>_REDIRECT_URI env var if needed."""
    env_name = f"{provider_key.upper()}_REDIRECT_URI"
    return os.getenv(env_name) or f"https://isibi-backend.onrender.com/api/ghost/connectors/oauth/{provider_key}/callback"


def _mail_oauth_provider_for_app(app_id: str) -> Optional[str]:
    """Reverse-lookup: which OAuth provider key powers this app_id?"""
    for provider_key, cfg in MAIL_OAUTH_PROVIDERS.items():
        if app_id in cfg.get("app_ids", []):
            return provider_key
    return None


def _mail_oauth_configured(provider_key: str) -> bool:
    cid, secret = _mail_oauth_client(provider_key)
    return bool(cid and secret)


# Flip OAuth-capable mail presets to the generic "mail_oauth" flow *only*
# when the provider's client ID + secret are configured on the server.
# Runs AFTER both APP_REGISTRY and MAIL_OAUTH_PROVIDERS are defined —
# positioning matters because Python evaluates module-level code in
# order. Until the env vars are set, the mail preset entries stay as
# password-based IMAP tiles.
for _provider_key, _cfg in MAIL_OAUTH_PROVIDERS.items():
    if not _mail_oauth_configured(_provider_key):
        continue
    for _app_id in _cfg.get("app_ids", []):
        if _app_id not in APP_REGISTRY:
            continue
        APP_REGISTRY[_app_id]["oauth_flow"] = "mail_oauth"
        APP_REGISTRY[_app_id]["auth_fields"] = [
            {"key": "access_token", "label": f"{_cfg['display']} OAuth Access Token", "secure": True},
        ]
        APP_REGISTRY[_app_id]["setup"] = (
            f"Tap 'Connect with {_cfg['display']}' to sign in. "
            f"GoFarther never sees your password — the token comes straight "
            f"from {_cfg['display']} and is stored encrypted."
        )


def _ms_auth_url() -> str:
    return f"https://login.microsoftonline.com/{_ms_tenant()}/oauth2/v2.0/authorize"


def _ms_token_url() -> str:
    return f"https://login.microsoftonline.com/{_ms_tenant()}/oauth2/v2.0/token"


_MS_SCOPES = {
    "excel_online": "Files.ReadWrite offline_access",
    # Outlook Mail needs mailbox read/write + send. Mail.ReadWrite covers
    # marking read, moving, deleting, and flagging; Mail.Send is separate.
    "outlook_mail": "Mail.ReadWrite Mail.Send offline_access",
    # Future: onedrive, teams can share the same flow with different scopes
}

# In-memory map of oauth states → {user_id, app_id, created_at}
# (short-lived — states expire after 10 minutes)
_oauth_states: dict[str, dict] = {}


def _prune_old_oauth_states() -> None:
    cutoff = datetime.now(timezone.utc).timestamp() - 600  # 10 min
    for state, data in list(_oauth_states.items()):
        if data.get("ts", 0) < cutoff:
            _oauth_states.pop(state, None)


@router.post("/{app_id}/oauth/start")
async def oauth_start(app_id: str, authorization: str = Header(...)):
    """Begin OAuth flow for Microsoft apps. Returns an authorize URL the
    client should open in the system browser."""
    payload = _verify_auth(authorization)
    user_id = payload.get("sub")

    if app_id not in APP_REGISTRY:
        raise HTTPException(404, f"Unknown app: {app_id}")

    info = APP_REGISTRY[app_id]
    flow = info.get("oauth_flow")
    # Mail OAuth providers (yahoo/zoho/fastmail/mailru/yandex) all share
    # one generic handler driven by the MAIL_OAUTH_PROVIDERS config table.
    mail_provider = _mail_oauth_provider_for_app(app_id) if flow == "mail_oauth" else None
    if flow not in ("microsoft", "google", "mail_oauth"):
        raise HTTPException(400, f"{info['name']} does not use an OAuth flow we handle")

    _prune_old_oauth_states()
    state = uuid.uuid4().hex
    _oauth_states[state] = {
        "user_id": str(user_id),
        "app_id": app_id,
        "ts": datetime.now(timezone.utc).timestamp(),
    }

    from urllib.parse import urlencode

    if flow == "microsoft":
        client_id = _ms_client_id()
        redirect_uri = _ms_redirect_uri()
        if not client_id:
            raise HTTPException(
                500,
                "Microsoft OAuth is not configured on the server. "
                "Set MICROSOFT_CLIENT_ID (or TEAMS_APP_ID) and MICROSOFT_CLIENT_SECRET.",
            )
        scope = _MS_SCOPES.get(app_id, "Files.ReadWrite offline_access")
        params = {
            "client_id": client_id,
            "response_type": "code",
            "redirect_uri": redirect_uri,
            "response_mode": "query",
            "scope": scope,
            "state": state,
            "prompt": "select_account",
        }
        authorize_url = f"{_ms_auth_url()}?{urlencode(params)}"
        return {"authorize_url": authorize_url, "state": state}

    if flow == "google":
        client_id = _google_client_id()
        redirect_uri = _google_redirect_uri()
        if not client_id:
            raise HTTPException(
                500,
                "Google OAuth is not configured on the server. "
                "Set GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET.",
            )
        scope = _GOOGLE_SCOPES.get(app_id, "openid email")
        params = {
            "client_id": client_id,
            "response_type": "code",
            "redirect_uri": redirect_uri,
            "scope": scope,
            "state": state,
            "access_type": "offline",       # gets us a refresh_token
            "prompt": "consent",            # forces the refresh_token on every connect
            "include_granted_scopes": "true",
        }
        authorize_url = f"{_GOOGLE_AUTH_URL}?{urlencode(params)}"
        return {"authorize_url": authorize_url, "state": state}

    # Generic mail OAuth (yahoo / zoho / fastmail / mailru / yandex)
    if not mail_provider:
        raise HTTPException(400, f"{info['name']} is not wired to a known mail OAuth provider")
    cfg = MAIL_OAUTH_PROVIDERS[mail_provider]
    client_id, _secret = _mail_oauth_client(mail_provider)
    if not client_id:
        raise HTTPException(
            500,
            f"{cfg['display']} OAuth is not configured on the server. "
            f"Set {cfg['client_id_env']} and {cfg['client_secret_env']} in Render env vars.",
        )
    params = {
        "client_id": client_id,
        "response_type": "code",
        "redirect_uri": _mail_oauth_redirect_uri(mail_provider),
        "scope": cfg["scopes"],
        "state": state,
    }
    # Zoho wants offline access explicitly to return a refresh token
    if mail_provider == "zoho":
        params["access_type"] = "offline"
        params["prompt"] = "consent"
    # Yandex uses force_confirm instead of prompt=consent
    if mail_provider == "yandex":
        params["force_confirm"] = "yes"
    authorize_url = f"{cfg['auth_url']}?{urlencode(params)}"
    return {"authorize_url": authorize_url, "state": state}


@router.get("/oauth/microsoft/callback")
async def oauth_microsoft_callback(
    code: Optional[str] = None,
    state: Optional[str] = None,
    error: Optional[str] = None,
    error_description: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    """Handle Microsoft OAuth redirect. Exchange code for tokens, save them,
    and show a success page the user can close."""
    from fastapi.responses import HTMLResponse

    def _html(title: str, body: str, ok: bool = True) -> HTMLResponse:
        color = "#16a34a" if ok else "#dc2626"
        return HTMLResponse(
            f"""<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>{title}</title></head><body style="font-family:-apple-system,system-ui,sans-serif;text-align:center;padding:40px 20px;background:#fafafa"><div style="max-width:420px;margin:0 auto;background:#fff;border-radius:16px;padding:32px 24px;box-shadow:0 2px 12px rgba(0,0,0,.06)"><div style="font-size:48px;margin-bottom:12px">{'✅' if ok else '⚠️'}</div><h1 style="color:{color};font-size:22px;margin:0 0 12px">{title}</h1><p style="color:#555;font-size:15px;line-height:1.5;margin:0 0 20px">{body}</p><p style="color:#999;font-size:13px;margin:0">You can close this window and return to the app.</p></div></body></html>"""
        )

    if error:
        return _html("Connection Failed", error_description or error, ok=False)
    if not code or not state:
        return _html("Connection Failed", "Missing authorization code or state.", ok=False)

    _prune_old_oauth_states()
    session = _oauth_states.pop(state, None)
    if not session:
        return _html(
            "Connection Failed",
            "This link has expired or was already used. Please try connecting again from the app.",
            ok=False,
        )

    user_id = session["user_id"]
    app_id = session["app_id"]

    client_id = _ms_client_id()
    client_secret = _ms_client_secret()
    redirect_uri = _ms_redirect_uri()
    if not client_id or not client_secret:
        return _html("Server Misconfigured", "Microsoft OAuth env vars are not set.", ok=False)

    # Exchange code for tokens
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            r = await client.post(
                _ms_token_url(),
                data={
                    "client_id": client_id,
                    "client_secret": client_secret,
                    "code": code,
                    "redirect_uri": redirect_uri,
                    "grant_type": "authorization_code",
                    "scope": _MS_SCOPES.get(app_id, "Files.ReadWrite offline_access"),
                },
                headers={"Content-Type": "application/x-www-form-urlencoded"},
            )
            if r.status_code != 200:
                logger.error(f"Microsoft token exchange failed: {r.status_code} {r.text}")
                return _html("Token Exchange Failed", r.text[:200], ok=False)
            tokens = r.json()
    except Exception as e:
        logger.exception("Microsoft token exchange errored")
        return _html("Token Exchange Error", str(e), ok=False)

    access_token = tokens.get("access_token")
    refresh_token = tokens.get("refresh_token")
    expires_in = int(tokens.get("expires_in") or 3600)
    if not access_token:
        return _html("Invalid Response", "Microsoft did not return an access token.", ok=False)

    expires_at = datetime.now(timezone.utc).timestamp() + expires_in - 60  # 60s buffer
    creds = {
        "access_token": access_token,
        "refresh_token": refresh_token,
        "expires_at": expires_at,
        "token_type": tokens.get("token_type", "Bearer"),
    }

    await _set_creds(user_id, app_id, creds, db)
    await db.commit()
    logger.info(f"Microsoft OAuth success: user={user_id} app={app_id}")
    app_name = APP_REGISTRY.get(app_id, {}).get("name", app_id)
    return _html(
        "Connected!",
        f"{app_name} is now linked to your GoFarther account.",
        ok=True,
    )


@router.get("/oauth/google/callback")
async def oauth_google_callback(
    code: Optional[str] = None,
    state: Optional[str] = None,
    error: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    """Handle Google OAuth redirect. Same shape as the Microsoft callback —
    exchange the code for tokens, persist them, show a success/error page."""
    from fastapi.responses import HTMLResponse

    def _html(title: str, body: str, ok: bool = True) -> HTMLResponse:
        color = "#16a34a" if ok else "#dc2626"
        return HTMLResponse(
            f"""<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>{title}</title></head><body style="font-family:-apple-system,system-ui,sans-serif;text-align:center;padding:40px 20px;background:#fafafa"><div style="max-width:420px;margin:0 auto;background:#fff;border-radius:16px;padding:32px 24px;box-shadow:0 2px 12px rgba(0,0,0,.06)"><div style="font-size:48px;margin-bottom:12px">{'✅' if ok else '⚠️'}</div><h1 style="color:{color};font-size:22px;margin:0 0 12px">{title}</h1><p style="color:#555;font-size:15px;line-height:1.5;margin:0 0 20px">{body}</p><p style="color:#999;font-size:13px;margin:0">You can close this window and return to the app.</p></div></body></html>"""
        )

    if error:
        return _html("Connection Failed", error, ok=False)
    if not code or not state:
        return _html("Connection Failed", "Missing authorization code or state.", ok=False)

    _prune_old_oauth_states()
    session = _oauth_states.pop(state, None)
    if not session:
        return _html(
            "Connection Failed",
            "This link has expired or was already used. Please try connecting again from the app.",
            ok=False,
        )

    user_id = session["user_id"]
    app_id = session["app_id"]

    client_id = _google_client_id()
    client_secret = _google_client_secret()
    redirect_uri = _google_redirect_uri()
    if not client_id or not client_secret:
        return _html("Server Misconfigured", "Google OAuth env vars are not set.", ok=False)

    try:
        async with httpx.AsyncClient(timeout=30) as client:
            r = await client.post(
                _GOOGLE_TOKEN_URL,
                data={
                    "client_id": client_id,
                    "client_secret": client_secret,
                    "code": code,
                    "redirect_uri": redirect_uri,
                    "grant_type": "authorization_code",
                },
                headers={"Content-Type": "application/x-www-form-urlencoded"},
            )
            if r.status_code != 200:
                logger.error(f"Google token exchange failed: {r.status_code} {r.text}")
                return _html("Token Exchange Failed", r.text[:200], ok=False)
            tokens = r.json()
    except Exception as e:
        logger.exception("Google token exchange errored")
        return _html("Token Exchange Error", str(e), ok=False)

    access_token = tokens.get("access_token")
    refresh_token = tokens.get("refresh_token")
    expires_in = int(tokens.get("expires_in") or 3600)
    if not access_token:
        return _html("Invalid Response", "Google did not return an access token.", ok=False)

    expires_at = datetime.now(timezone.utc).timestamp() + expires_in - 60
    creds = {
        "access_token": access_token,
        "refresh_token": refresh_token,
        "expires_at": expires_at,
        "token_type": tokens.get("token_type", "Bearer"),
    }

    await _set_creds(user_id, app_id, creds, db)
    await db.commit()
    logger.info(f"Google OAuth success: user={user_id} app={app_id}")
    app_name = APP_REGISTRY.get(app_id, {}).get("name", app_id)
    return _html(
        "Connected!",
        f"{app_name} is now linked to your GoFarther account.",
        ok=True,
    )


async def _refresh_google_token(user_id, app_id: str, creds: dict, db: AsyncSession) -> dict:
    """Refresh a Google access token if it's expired or about to be."""
    expires_at = float(creds.get("expires_at") or 0)
    now = datetime.now(timezone.utc).timestamp()
    if expires_at > now + 30:
        return creds

    refresh_token = creds.get("refresh_token")
    if not refresh_token:
        return creds

    client_id = _google_client_id()
    client_secret = _google_client_secret()
    if not client_id or not client_secret:
        return creds

    try:
        async with httpx.AsyncClient(timeout=30) as client:
            r = await client.post(
                _GOOGLE_TOKEN_URL,
                data={
                    "client_id": client_id,
                    "client_secret": client_secret,
                    "grant_type": "refresh_token",
                    "refresh_token": refresh_token,
                },
                headers={"Content-Type": "application/x-www-form-urlencoded"},
            )
            if r.status_code != 200:
                logger.warning(f"Google token refresh failed: {r.status_code} {r.text}")
                return creds
            tokens = r.json()
    except Exception:
        logger.exception("Google token refresh errored")
        return creds

    new_creds = dict(creds)
    new_creds["access_token"] = tokens.get("access_token", creds.get("access_token"))
    # Google only returns a new refresh_token on the first consent, so keep the old one
    if tokens.get("refresh_token"):
        new_creds["refresh_token"] = tokens["refresh_token"]
    new_creds["expires_at"] = now + int(tokens.get("expires_in") or 3600) - 60

    await _set_creds(user_id, app_id, new_creds, db)
    await db.commit()
    return new_creds


@router.get("/oauth/{provider}/callback")
async def oauth_mail_generic_callback(
    provider: str,
    code: Optional[str] = None,
    state: Optional[str] = None,
    error: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    """Generic OAuth callback for all mail providers in MAIL_OAUTH_PROVIDERS
    (yahoo, zoho, fastmail, mailru, yandex). Exchanges the code for an
    access_token + refresh_token and stores it the same way the Microsoft
    and Google callbacks do."""
    from fastapi.responses import HTMLResponse

    def _html(title: str, body: str, ok: bool = True) -> HTMLResponse:
        color = "#16a34a" if ok else "#dc2626"
        return HTMLResponse(
            f"""<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>{title}</title></head><body style="font-family:-apple-system,system-ui,sans-serif;text-align:center;padding:40px 20px;background:#fafafa"><div style="max-width:420px;margin:0 auto;background:#fff;border-radius:16px;padding:32px 24px;box-shadow:0 2px 12px rgba(0,0,0,.06)"><div style="font-size:48px;margin-bottom:12px">{'✅' if ok else '⚠️'}</div><h1 style="color:{color};font-size:22px;margin:0 0 12px">{title}</h1><p style="color:#555;font-size:15px;line-height:1.5;margin:0 0 20px">{body}</p><p style="color:#999;font-size:13px;margin:0">You can close this window and return to the app.</p></div></body></html>"""
        )

    if provider not in MAIL_OAUTH_PROVIDERS:
        return _html("Unknown Provider", f"No mail OAuth provider named '{provider}'.", ok=False)
    cfg = MAIL_OAUTH_PROVIDERS[provider]

    if error:
        return _html("Connection Failed", error, ok=False)
    if not code or not state:
        return _html("Connection Failed", "Missing authorization code or state.", ok=False)

    _prune_old_oauth_states()
    session = _oauth_states.pop(state, None)
    if not session:
        return _html(
            "Connection Failed",
            "This link has expired or was already used. Please try connecting again from the app.",
            ok=False,
        )

    user_id = session["user_id"]
    app_id = session["app_id"]

    client_id, client_secret = _mail_oauth_client(provider)
    redirect_uri = _mail_oauth_redirect_uri(provider)
    if not client_id or not client_secret:
        return _html("Server Misconfigured", f"{cfg['display']} OAuth env vars are not set.", ok=False)

    try:
        async with httpx.AsyncClient(timeout=30) as client:
            r = await client.post(
                cfg["token_url"],
                data={
                    "client_id": client_id,
                    "client_secret": client_secret,
                    "code": code,
                    "redirect_uri": redirect_uri,
                    "grant_type": "authorization_code",
                },
                headers={"Content-Type": "application/x-www-form-urlencoded"},
            )
            if r.status_code != 200:
                logger.error(f"{provider} token exchange failed: {r.status_code} {r.text}")
                return _html("Token Exchange Failed", r.text[:200], ok=False)
            tokens = r.json()
    except Exception as e:
        logger.exception(f"{provider} token exchange errored")
        return _html("Token Exchange Error", str(e), ok=False)

    access_token = tokens.get("access_token")
    refresh_token = tokens.get("refresh_token")
    expires_in = int(tokens.get("expires_in") or 3600)
    if not access_token:
        return _html("Invalid Response", f"{cfg['display']} did not return an access token.", ok=False)

    expires_at = datetime.now(timezone.utc).timestamp() + expires_in - 60
    creds = {
        "access_token": access_token,
        "refresh_token": refresh_token,
        "expires_at": expires_at,
        "token_type": tokens.get("token_type", "Bearer"),
        # Pin the IMAP/SMTP hosts so the adapter doesn't need to autodetect
        "imap_host": cfg["imap_host"],
        "imap_port": cfg["imap_port"],
        "smtp_host": cfg["smtp_host"],
        "smtp_port": cfg["smtp_port"],
        # XOAUTH2 uses the email address as the "user" — we store it here
        # for convenience but it gets filled in on first use if missing.
        # Yahoo / Mail.ru / Yandex expose the address on a userinfo endpoint;
        # Zoho / Fastmail put it in the token response.
        "username": tokens.get("email") or tokens.get("mail") or "",
        "oauth_provider": provider,
    }

    await _set_creds(user_id, app_id, creds, db)
    await db.commit()
    logger.info(f"{provider} mail OAuth success: user={user_id} app={app_id}")
    app_name = APP_REGISTRY.get(app_id, {}).get("name", app_id)
    return _html(
        "Connected!",
        f"{app_name} is now linked to your GoFarther account.",
        ok=True,
    )


async def _refresh_mail_oauth_token(user_id, app_id: str, creds: dict, db: AsyncSession) -> dict:
    """Refresh a mail OAuth access token if it's expired or close to it.
    Works for any provider in MAIL_OAUTH_PROVIDERS."""
    expires_at = float(creds.get("expires_at") or 0)
    now = datetime.now(timezone.utc).timestamp()
    if expires_at > now + 30:
        return creds

    refresh_token = creds.get("refresh_token")
    if not refresh_token:
        return creds

    provider = creds.get("oauth_provider") or _mail_oauth_provider_for_app(app_id)
    if not provider or provider not in MAIL_OAUTH_PROVIDERS:
        return creds
    cfg = MAIL_OAUTH_PROVIDERS[provider]

    client_id, client_secret = _mail_oauth_client(provider)
    if not client_id or not client_secret:
        return creds

    try:
        async with httpx.AsyncClient(timeout=30) as client:
            r = await client.post(
                cfg["token_url"],
                data={
                    "client_id": client_id,
                    "client_secret": client_secret,
                    "grant_type": "refresh_token",
                    "refresh_token": refresh_token,
                },
                headers={"Content-Type": "application/x-www-form-urlencoded"},
            )
            if r.status_code != 200:
                logger.warning(f"{provider} token refresh failed: {r.status_code} {r.text}")
                return creds
            tokens = r.json()
    except Exception:
        logger.exception(f"{provider} token refresh errored")
        return creds

    new_creds = dict(creds)
    new_creds["access_token"] = tokens.get("access_token", creds.get("access_token"))
    if tokens.get("refresh_token"):
        new_creds["refresh_token"] = tokens["refresh_token"]
    new_creds["expires_at"] = now + int(tokens.get("expires_in") or 3600) - 60

    await _set_creds(user_id, app_id, new_creds, db)
    await db.commit()
    return new_creds


async def _refresh_microsoft_token(user_id, app_id: str, creds: dict, db: AsyncSession) -> dict:
    """If the Microsoft access token is expired or close to it, use the
    refresh token to get a new one and persist it. Returns updated creds."""
    expires_at = float(creds.get("expires_at") or 0)
    now = datetime.now(timezone.utc).timestamp()
    if expires_at > now + 30:  # still fresh
        return creds

    refresh_token = creds.get("refresh_token")
    if not refresh_token:
        return creds  # caller will get a 401 from Graph

    client_id = _ms_client_id()
    client_secret = _ms_client_secret()
    if not client_id or not client_secret:
        return creds

    try:
        async with httpx.AsyncClient(timeout=30) as client:
            r = await client.post(
                _ms_token_url(),
                data={
                    "client_id": client_id,
                    "client_secret": client_secret,
                    "grant_type": "refresh_token",
                    "refresh_token": refresh_token,
                    "scope": _MS_SCOPES.get(app_id, "Files.ReadWrite offline_access"),
                },
                headers={"Content-Type": "application/x-www-form-urlencoded"},
            )
            if r.status_code != 200:
                logger.warning(f"Microsoft token refresh failed: {r.status_code} {r.text}")
                return creds
            tokens = r.json()
    except Exception:
        logger.exception("Microsoft token refresh errored")
        return creds

    new_creds = dict(creds)
    new_creds["access_token"] = tokens.get("access_token", creds.get("access_token"))
    if tokens.get("refresh_token"):
        new_creds["refresh_token"] = tokens["refresh_token"]
    new_creds["expires_at"] = now + int(tokens.get("expires_in") or 3600) - 60

    await _set_creds(user_id, app_id, new_creds, db)
    await db.commit()
    return new_creds


@router.post("/{app_id}/action")
async def execute_action(app_id: str, body: ActionRequest, authorization: str = Header(...), db: AsyncSession = Depends(get_db)):
    """Execute an action on a connected app."""
    payload = _verify_auth(authorization)
    user_id = payload.get("sub")

    if app_id not in APP_REGISTRY:
        raise HTTPException(404, f"Unknown app: {app_id}")

    creds = await _get_creds(user_id, app_id, db)
    if not creds:
        raise HTTPException(400, f"{APP_REGISTRY[app_id]['name']} is not connected. Go to Settings → Connect Apps to set it up.")

    # Refresh OAuth tokens if they're close to expiring
    _flow = APP_REGISTRY[app_id].get("oauth_flow")
    if _flow == "microsoft":
        creds = await _refresh_microsoft_token(user_id, app_id, creds, db)
    elif _flow == "google":
        creds = await _refresh_google_token(user_id, app_id, creds, db)
    elif _flow == "mail_oauth":
        creds = await _refresh_mail_oauth_token(user_id, app_id, creds, db)

    if body.action not in APP_REGISTRY[app_id]["actions"]:
        raise HTTPException(400, f"Action '{body.action}' not available for {APP_REGISTRY[app_id]['name']}")

    adapter = ADAPTERS.get(app_id)
    if not adapter:
        raise HTTPException(501, f"{APP_REGISTRY[app_id]['name']} adapter not yet implemented")

    try:
        result = await adapter(body.action, body.params, creds)
        return {"status": "ok", "app": app_id, "action": body.action, "result": result}
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Connector action failed: {app_id}/{body.action}")
        raise HTTPException(500, f"{APP_REGISTRY[app_id]['name']} error: {str(e)}")


# ── Unified outbound email router ───────────────────────────────────────
#
# Single entry point for "send an email on behalf of this user". Everything
# that needs to send mail on behalf of a user — the plan executor,
# scheduled reports, the background worker, /ghost-tools-v2/send-email —
# routes through here.
#
# POLICY: outbound email ONLY goes through a connected mail app. There is
# no SMTP or Resend fallback for user-initiated emails. This is intentional:
#
#   - The email must show up in the user's real Sent folder.
#   - The from-address must be the user's real email so replies come back.
#   - Having silent fallbacks caused the LLM to pick inconsistently.
#
# Preference order when the user has multiple mail apps connected:
#
#   1. Gmail (Gmail API)
#   2. Outlook Mail (Microsoft Graph)
#   3. Neo / Titan / generic IMAP (SMTP submission)
#
# If the user has none connected, send_email_for_user returns a clean
# error that the caller surfaces to the user with instructions to connect
# an app. System-level emails (verification codes, password resets, login
# alerts) continue to use Resend directly via services/email.py and do
# NOT go through this router.
#
# Legacy GhostUser.smtp_* columns are also auto-migrated to an imap_mail
# connector entry the first time we see them, so users upgrading from the
# old "Send from My Email" tile don't need to reconnect.


_MAIL_CONNECTOR_PREFERENCE = (
    # OAuth / native-API connectors first (most reliable, best UX)
    "gmail", "outlook_mail",
    # IMAP presets (ordered by popularity)
    "neo_mail", "titan_mail",
    "yahoo_mail", "icloud_mail", "zoho_mail", "fastmail_mail",
    "aol_mail", "gmx_mail", "mailru_mail", "yandex_mail",
    "protonmail_mail",
    "hostinger_mail", "godaddy_mail", "namecheap_mail", "ionos_mail",
    "mailboxorg_mail", "posteo_mail", "mailfence_mail",
    # Generic IMAP catch-all last
    "imap_mail",
)


async def _migrate_legacy_smtp_to_imap_mail(user_id, smtp_settings: dict, db: AsyncSession) -> bool:
    """One-shot migration from GhostUser.smtp_* columns to an imap_mail
    connector entry. Idempotent — if the user already has imap_mail
    connected, this is a no-op. Returns True if a new connector row was
    created."""
    if not smtp_settings:
        return False
    existing = await _get_creds(user_id, "imap_mail", db)
    if existing:
        return False
    username = smtp_settings.get("smtp_user") or smtp_settings.get("smtp_from")
    password = smtp_settings.get("smtp_pass")
    if not (username and password):
        return False
    # We don't know the IMAP host from the legacy columns (they only stored
    # SMTP). Leave imap_host blank and let _autodetect_mail_servers fill it
    # in the first time the IMAP adapter runs against the username domain.
    new_creds = {
        "username": username,
        "app_password": password,
        "smtp_host": smtp_settings.get("smtp_host") or "",
        "smtp_port": int(smtp_settings.get("smtp_port") or 587),
        # Leave imap_host unset — the adapter will autodetect it from the
        # username domain when needed. For outbound-only legacy users that
        # never touch IMAP, the blank is harmless.
    }
    await _set_creds(user_id, "imap_mail", new_creds, db)
    await db.commit()
    logger.info(f"Migrated legacy SMTP → imap_mail connector: user={user_id}")
    return True


async def _refresh_mail_creds(user_id, app_id: str, creds: dict, db: AsyncSession) -> dict:
    """Run the right token-refresh helper for a given mail connector."""
    flow = APP_REGISTRY.get(app_id, {}).get("oauth_flow")
    if flow == "microsoft":
        return await _refresh_microsoft_token(user_id, app_id, creds, db)
    if flow == "google":
        return await _refresh_google_token(user_id, app_id, creds, db)
    if flow == "mail_oauth":
        return await _refresh_mail_oauth_token(user_id, app_id, creds, db)
    return creds


async def send_email_for_user(
    user_id,
    user_email: str,
    db: AsyncSession,
    *,
    to,
    subject: str,
    html: str,
    cc=None,
    bcc=None,
    attachments: list | None = None,
) -> dict:
    """Unified "send an email as this user" helper. Returns
    {sent: bool, via: str, error?: str, ...}.

    `to`/`cc`/`bcc` accept either a string (comma-separated) or a list."""
    # Normalize addresses for the connector adapters (they accept either).
    to_str = ",".join(to) if isinstance(to, list) else (to or "")
    cc_str = ",".join(cc) if isinstance(cc, list) else (cc or "")
    bcc_str = ",".join(bcc) if isinstance(bcc, list) else (bcc or "")

    # Build a params dict our connector adapters understand
    base_params = {
        "to": to_str,
        "subject": subject,
        "body": html,
    }
    if cc_str:
        base_params["cc"] = cc_str
    if bcc_str:
        base_params["bcc"] = bcc_str

    # Auto-migrate: if the user has legacy GhostUser.smtp_* columns and
    # does NOT yet have an imap_mail connector, copy the settings over so
    # their first outbound email works without them having to reconnect.
    try:
        from routes.ghost_auth import get_user_smtp
        legacy = await get_user_smtp(user_email, db)
        if legacy.get("smtp_host"):
            await _migrate_legacy_smtp_to_imap_mail(user_id, legacy, db)
    except Exception:
        logger.exception("Legacy SMTP migration check failed (ignored)")

    # Try each connected mail connector in preference order. First success
    # wins; any errors from individual connectors are logged and we move on
    # to the next connected one.
    tried: list[str] = []
    last_error: str | None = None
    for app_id in _MAIL_CONNECTOR_PREFERENCE:
        creds = await _get_creds(user_id, app_id, db)
        if not creds:
            continue
        tried.append(app_id)
        creds = await _refresh_mail_creds(user_id, app_id, creds, db)
        adapter = ADAPTERS.get(app_id)
        if not adapter:
            continue
        try:
            result = await adapter("send_email", base_params, creds)
        except Exception as e:
            logger.warning(f"send_email via {app_id} raised: {e}")
            last_error = f"{app_id}: {e}"
            continue
        if isinstance(result, dict) and "error" not in result:
            return {
                "sent": True,
                "via": app_id,
                "to": to_str,
                "subject": subject,
                # NOTE: attachments aren't yet piped through the connector
                # adapters — the plan executor's email step is still the
                # only path that supports attachments natively.
                "attachments_sent": False if attachments else None,
                **(result if isinstance(result, dict) else {}),
            }
        last_error = f"{app_id}: {result.get('error') if isinstance(result, dict) else result}"
        logger.warning(f"Mail connector {app_id} failed: {last_error}")

    # No connected mail app worked — return a clean, actionable error.
    if not tried:
        return {
            "sent": False,
            "error": (
                "No email app connected. Go to Settings → Connect Apps and "
                "connect Gmail, Outlook, Neo, Titan, or any email provider "
                "(IMAP) so GoFarther can send emails from your real account."
            ),
        }
    return {
        "sent": False,
        "error": (
            f"Could not send through any connected email app "
            f"({', '.join(tried)}). Last error: {last_error}"
        ),
    }


# ── Multi-step plan executor ─────────────────────────────────────────────

class PlanStep(BaseModel):
    id: str | None = None
    type: str  # "connector" | "excel_pdf" | "email"
    app: str | None = None          # for type=connector
    action: str | None = None       # for type=connector
    params: dict[str, Any] = {}


class PlanRequest(BaseModel):
    steps: list[PlanStep]


def _resolve_plan_refs(value: Any, outputs: dict[str, dict]) -> Any:
    """Replace $stepId.field references inside strings/dicts/lists with the
    matching output from a previous step. Supports exact-match ("$step.field")
    and inline substitution ("Report: $step.sum")."""
    import re as _re
    if isinstance(value, str):
        # Exact match → return the raw value (could be non-string, e.g. bytes)
        m = _re.fullmatch(r"\$([A-Za-z0-9_]+)\.([A-Za-z0-9_]+)", value)
        if m:
            step_id, field = m.group(1), m.group(2)
            out = outputs.get(step_id) or {}
            return out.get(field)
        # Inline substitution — only stringifies simple scalar values
        def _sub(match):
            sid, fld = match.group(1), match.group(2)
            out = outputs.get(sid) or {}
            v = out.get(fld)
            if v is None:
                return match.group(0)
            if isinstance(v, (bytes, bytearray)):
                return f"<{len(v)} bytes>"
            return str(v)
        return _re.sub(r"\$([A-Za-z0-9_]+)\.([A-Za-z0-9_]+)", _sub, value)
    if isinstance(value, dict):
        return {k: _resolve_plan_refs(v, outputs) for k, v in value.items()}
    if isinstance(value, list):
        return [_resolve_plan_refs(v, outputs) for v in value]
    return value


def _xlsx_base_name(n: str) -> str:
    """Same fuzzy normalization the Excel adapter uses for workbook matching:
    lowercase, strip repeated .xlsx extensions, drop spaces/underscores/dashes."""
    s = (n or "").lower().strip()
    while s.endswith(".xlsx"):
        s = s[:-5]
    for ch in (" ", "_", "-"):
        s = s.replace(ch, "")
    return s


async def _list_all_xlsx(client: "httpx.AsyncClient", token: str) -> list[dict]:
    """Module-level copy of the Excel adapter's _collect_all_xlsx. Returns
    every .xlsx in the user's OneDrive (root, first-level subfolders, and
    the search index). Kept identical so the plan executor's PDF helper
    matches the same files the Excel adapter would."""
    base = "https://graph.microsoft.com/v1.0"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    seen: dict[str, dict] = {}

    r = await client.get(
        f"{base}/me/drive/root/children",
        headers=headers,
        params={"$top": 200, "$select": "id,name,folder,webUrl,lastModifiedDateTime"},
    )
    subfolder_ids: list[str] = []
    if r.status_code == 200:
        for it in (r.json() or {}).get("value", []):
            if it.get("folder"):
                if it.get("id"):
                    subfolder_ids.append(it["id"])
                continue
            name = it.get("name") or ""
            if name.lower().endswith(".xlsx") and it.get("id"):
                seen[it["id"]] = it

    for fid in subfolder_ids[:20]:
        rr = await client.get(
            f"{base}/me/drive/items/{fid}/children",
            headers=headers,
            params={"$top": 200, "$select": "id,name,webUrl,lastModifiedDateTime"},
        )
        if rr.status_code == 200:
            for it in (rr.json() or {}).get("value", []):
                name = it.get("name") or ""
                if name.lower().endswith(".xlsx") and it.get("id"):
                    seen.setdefault(it["id"], it)

    rs = await client.get(
        f"{base}/me/drive/root/search(q='.xlsx')",
        headers=headers,
        params={"$top": 50, "$select": "id,name,webUrl,lastModifiedDateTime"},
    )
    if rs.status_code == 200:
        for it in (rs.json() or {}).get("value", []):
            name = it.get("name") or ""
            if name.lower().endswith(".xlsx") and it.get("id"):
                seen.setdefault(it["id"], it)
    return list(seen.values())


async def _resolve_xlsx(client: "httpx.AsyncClient", token: str, workbook_ref: str) -> tuple[str | None, str | None]:
    """Shared workbook resolver for the plan executor. Mirrors the logic in
    _excel_online_adapter._resolve_workbook_id: exact → substring → fallback,
    plus auto-pick when there's only one file.

    Returns (item_id, error_message).
    """
    # Graph item id heuristic (40+ chars, starts with "01", no separators)
    if (
        workbook_ref
        and len(workbook_ref) > 20
        and workbook_ref.startswith("01")
        and "." not in workbook_ref
        and "/" not in workbook_ref
        and " " not in workbook_ref
    ):
        return workbook_ref, None

    all_wbs = await _list_all_xlsx(client, token)
    if not all_wbs:
        return None, "No Excel files found in OneDrive."

    if not workbook_ref:
        if len(all_wbs) == 1:
            return all_wbs[0]["id"], None
        names = ", ".join([w.get("name", "") for w in all_wbs[:10]])
        return None, f"Multiple Excel files — specify which one: {names}"

    target = _xlsx_base_name(workbook_ref)
    exact = [w for w in all_wbs if _xlsx_base_name(w.get("name", "")) == target]
    if len(exact) == 1:
        return exact[0]["id"], None
    if len(exact) > 1:
        names = ", ".join([w.get("name", "") for w in exact[:10]])
        return None, f"Multiple matches for '{workbook_ref}': {names}"

    partial = [
        w for w in all_wbs
        if target in _xlsx_base_name(w.get("name", ""))
        or _xlsx_base_name(w.get("name", "")) in target
    ]
    if len(partial) == 1:
        return partial[0]["id"], None
    if len(partial) > 1:
        names = ", ".join([w.get("name", "") for w in partial[:10]])
        return None, f"Multiple matches for '{workbook_ref}': {names}"

    all_names = ", ".join([w.get("name", "") for w in all_wbs[:10]])
    return None, f"No Excel file matches '{workbook_ref}'. Your files: {all_names}"


async def _fetch_excel_pdf_bytes(workbook_ref: str, creds: dict) -> tuple[bytes | None, str | None]:
    """Download an xlsx file from OneDrive as a PDF via Graph.

    Uses the shared fuzzy resolver so "budget" matches "Test Budget.xlsx",
    "budget-2024.xlsx", etc. the same way the Excel adapter does.
    """
    import httpx
    token = creds.get("access_token") or creds.get("api_key")
    if not token:
        return None, "Excel Online is not connected."
    base = "https://graph.microsoft.com/v1.0"

    async with httpx.AsyncClient(timeout=60) as client:
        wid, err = await _resolve_xlsx(client, token, workbook_ref or "")
        if err:
            return None, err
        if not wid:
            return None, f"Workbook '{workbook_ref}' not found"
        # follow_redirects=True so httpx pulls the signed URL and returns bytes
        r = await client.get(
            f"{base}/me/drive/items/{wid}/content?format=pdf",
            headers={"Authorization": f"Bearer {token}"},
            follow_redirects=True,
        )
        if r.status_code == 200:
            return r.content, None
        return None, f"Graph PDF error {r.status_code}: {r.text[:200]}"


@router.post("/run_plan")
async def run_plan(body: PlanRequest, authorization: str = Header(...), db: AsyncSession = Depends(get_db)):
    """Execute a multi-step plan. Each step can be a connector action, a
    server-internal helper (excel_pdf), or an email send. Steps reference
    earlier outputs via "$stepId.field" strings.

    Returns {status, steps: [{id, type, ok, result/error}]}.
    """
    payload = _verify_auth(authorization)
    user_id = payload.get("sub")
    user_email = payload.get("email") or ""

    if not body.steps:
        raise HTTPException(400, "Plan has no steps")
    if len(body.steps) > 10:
        raise HTTPException(400, "Plan cannot exceed 10 steps")

    outputs: dict[str, dict] = {}
    step_results: list[dict] = []

    for idx, step in enumerate(body.steps):
        sid = step.id or f"s{idx}"
        resolved_params = _resolve_plan_refs(step.params, outputs)

        try:
            if step.type == "connector":
                if not step.app or not step.action:
                    raise ValueError("connector step requires app and action")
                if step.app not in APP_REGISTRY:
                    raise ValueError(f"Unknown app: {step.app}")
                creds = await _get_creds(user_id, step.app, db)
                if not creds:
                    raise ValueError(f"{APP_REGISTRY[step.app]['name']} is not connected")
                step_flow = APP_REGISTRY[step.app].get("oauth_flow")
                if step_flow == "microsoft":
                    creds = await _refresh_microsoft_token(user_id, step.app, creds, db)
                elif step_flow == "google":
                    creds = await _refresh_google_token(user_id, step.app, creds, db)
                elif step_flow == "mail_oauth":
                    creds = await _refresh_mail_oauth_token(user_id, step.app, creds, db)
                if step.action not in APP_REGISTRY[step.app]["actions"]:
                    raise ValueError(f"Action '{step.action}' not available for {step.app}")
                adapter = ADAPTERS.get(step.app)
                if not adapter:
                    raise ValueError(f"{step.app} adapter not implemented")
                result = await adapter(step.action, resolved_params, creds)
                outputs[sid] = result if isinstance(result, dict) else {"value": result}
                step_results.append({"id": sid, "type": "connector", "app": step.app, "action": step.action, "ok": "error" not in outputs[sid], "result": outputs[sid]})

            elif step.type == "excel_pdf":
                workbook_ref = resolved_params.get("workbook") or resolved_params.get("workbook_id") or resolved_params.get("filename") or ""
                creds = await _get_creds(user_id, "excel_online", db)
                if not creds:
                    raise ValueError("Excel Online is not connected")
                creds = await _refresh_microsoft_token(user_id, "excel_online", creds, db)
                pdf_bytes, err = await _fetch_excel_pdf_bytes(workbook_ref, creds)
                if err:
                    raise ValueError(err)
                filename = (workbook_ref.rsplit("/", 1)[-1] or "report").replace(".xlsx", "") + ".pdf"
                outputs[sid] = {"bytes": pdf_bytes, "filename": filename, "content_type": "application/pdf", "size": len(pdf_bytes or b"")}
                # Don't echo bytes back to the client — just the size.
                step_results.append({"id": sid, "type": "excel_pdf", "ok": True, "result": {"filename": filename, "size": len(pdf_bytes or b"")}})

            elif step.type == "convert_file":
                # Inputs (choose one): attach_from=<stepId>, url=<http url>,
                #                     content_base64=<base64 string>
                # Required: from_ext (or from_mime), to_ext (or to_mime)
                src_bytes: bytes | None = None
                src_name = resolved_params.get("filename") or "file"
                from_hint = resolved_params.get("from_ext") or resolved_params.get("from") or resolved_params.get("from_mime") or ""
                to_hint = resolved_params.get("to_ext") or resolved_params.get("to") or resolved_params.get("to_mime") or ""
                if not to_hint:
                    raise ValueError("convert_file step requires to_ext (e.g. 'pdf')")
                # 1. prior step bytes
                if resolved_params.get("attach_from"):
                    src = outputs.get(resolved_params["attach_from"]) or {}
                    src_bytes = src.get("bytes")
                    src_name = src.get("filename") or src_name
                    if not from_hint and src.get("content_type"):
                        from_hint = src["content_type"]
                # 2. arbitrary url
                if src_bytes is None and resolved_params.get("url"):
                    import httpx as _httpx
                    async with _httpx.AsyncClient(timeout=60, follow_redirects=True) as _c:
                        _r = await _c.get(resolved_params["url"])
                        if _r.status_code != 200:
                            raise ValueError(f"Could not download {resolved_params['url']}: HTTP {_r.status_code}")
                        src_bytes = _r.content
                        if not from_hint:
                            ct = _r.headers.get("content-type", "").split(";")[0].strip()
                            from_hint = ct or resolved_params["url"].rsplit(".", 1)[-1]
                        if "filename" not in resolved_params:
                            src_name = resolved_params["url"].rsplit("/", 1)[-1] or src_name
                # 3. inline base64 from client (used by standalone /convert path too)
                if src_bytes is None and resolved_params.get("content_base64"):
                    import base64 as _b64
                    src_bytes = _b64.b64decode(resolved_params["content_base64"])
                if src_bytes is None:
                    raise ValueError("convert_file needs attach_from, url, or content_base64")
                # Infer from_ext from filename if still missing
                if not from_hint and "." in src_name:
                    from_hint = src_name.rsplit(".", 1)[-1]
                from services.file_convert import convert_bytes_async as _convert_bytes_async
                out_name_base = (src_name.rsplit(".", 1)[0] or "file")
                out_bytes, out_mime, out_filename = await _convert_bytes_async(src_bytes, from_hint, to_hint, out_name=out_name_base)
                outputs[sid] = {"bytes": out_bytes, "filename": out_filename, "content_type": out_mime, "size": len(out_bytes)}
                step_results.append({"id": sid, "type": "convert_file", "ok": True, "result": {"filename": out_filename, "size": len(out_bytes), "from": from_hint, "to": to_hint}})

            elif step.type == "email":
                to = resolved_params.get("to") or ""
                subject = resolved_params.get("subject") or "Report"
                html = resolved_params.get("html") or resolved_params.get("body") or f"<p>{subject}</p>"
                cc = resolved_params.get("cc")
                bcc = resolved_params.get("bcc")
                # Attachments can reference prior step outputs by step id:
                #   {"attach_from": "pdf"}  →  use outputs["pdf"]
                # or be provided inline as {filename, content, content_type}.
                raw_attach = resolved_params.get("attachments") or []
                if isinstance(raw_attach, dict):
                    raw_attach = [raw_attach]
                attachments: list[dict] = []
                for a in raw_attach:
                    if not isinstance(a, dict):
                        continue
                    if a.get("attach_from"):
                        src = outputs.get(a["attach_from"]) or {}
                        if src.get("bytes"):
                            attachments.append({
                                "filename": a.get("filename") or src.get("filename") or "attachment.bin",
                                "content": src["bytes"],
                                "content_type": a.get("content_type") or src.get("content_type") or "application/octet-stream",
                            })
                    elif a.get("content"):
                        attachments.append(a)
                if not to:
                    raise ValueError("email step requires 'to'")
                # Route through the unified sender: connected mail app
                # first (so the email lands in the user's real Sent folder),
                # then legacy SMTP, then Resend.
                result = await send_email_for_user(
                    user_id,
                    user_email,
                    db,
                    to=to,
                    subject=subject,
                    html=html,
                    cc=cc,
                    bcc=bcc,
                    attachments=attachments or None,
                )
                if not result.get("sent"):
                    raise ValueError(result.get("error") or "Email could not be sent")
                outputs[sid] = {
                    "sent": True,
                    "to": to,
                    "subject": subject,
                    "attachment_count": len(attachments),
                    "via": result.get("via"),
                }
                step_results.append({"id": sid, "type": "email", "ok": True, "result": outputs[sid]})

            else:
                raise ValueError(f"Unknown step type: {step.type}")

        except Exception as e:
            logger.exception(f"Plan step {sid} ({step.type}) failed")
            step_results.append({"id": sid, "type": step.type, "ok": False, "error": str(e)})
            return {"status": "error", "failed_at": sid, "steps": step_results}

    return {"status": "ok", "steps": step_results}


@router.post("/convert")
async def convert_file_endpoint(
    file: UploadFile = File(...),
    to: str = Form(...),
    from_ext: str | None = Form(default=None),
    authorization: str = Header(...),
):
    """Ad-hoc file conversion. POST multipart with `file` + `to` (e.g. 'pdf')
    and optionally `from_ext`. Returns the converted file bytes directly with
    the right content-type header — the client can save it, attach it to a
    plan step, or share it.

    This is the simplest path for a user who uploads a file in chat and says
    "convert this to PDF". For plan-driven multi-step flows use the
    convert_file step of /run_plan instead.
    """
    _verify_auth(authorization)  # auth is required, but we don't need the sub
    from services.file_convert import convert_bytes_async as _convert_bytes_async

    data = await file.read()
    if not data:
        raise HTTPException(400, "Empty upload")

    # Prefer the explicit from_ext, then the uploaded filename suffix, then
    # the mime type the client sent.
    source_hint = (
        from_ext
        or (file.filename.rsplit(".", 1)[-1] if file.filename and "." in file.filename else "")
        or (file.content_type or "")
    )
    if not source_hint:
        raise HTTPException(400, "Could not detect source format — include from_ext")

    try:
        out_bytes, out_mime, out_filename = await _convert_bytes_async(
            data, source_hint, to,
            out_name=(file.filename.rsplit(".", 1)[0] if file.filename else "file"),
        )
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        logger.exception("File conversion failed")
        raise HTTPException(500, f"Conversion error: {e}")

    return Response(
        content=out_bytes,
        media_type=out_mime,
        headers={"Content-Disposition": f'attachment; filename="{out_filename}"'},
    )


@router.get("/email/autodetect")
async def email_autodetect(email: str, authorization: str = Header(...)):
    """Given an email address, return the detected IMAP + SMTP servers.
    Used by the Settings screen to pre-fill the "Advanced" section when a
    user is setting up a custom IMAP account, so they can see and override
    the guessed hosts before saving."""
    _verify_auth(authorization)
    if not email or "@" not in email:
        raise HTTPException(400, "A full email address is required")
    result = await _autodetect_mail_servers(email)
    if not result:
        return {"detected": False, "email": email}
    return {
        "detected": True,
        "email": email,
        "imap_host": result.get("imap_host"),
        "imap_port": result.get("imap_port"),
        "smtp_host": result.get("smtp_host"),
        "smtp_port": result.get("smtp_port"),
        "guessed": result.get("_guessed", False),
    }


@router.get("/convert/supported")
async def convert_supported(authorization: str = Header(...)):
    """List every conversion pair supported by /convert and the convert_file
    plan step. Used by the chat UI to gate the convert button.

    Returns pure-Python pairs under `pairs` and LibreOffice-backed pairs
    (only active when soffice is on PATH in the container) under `lo_pairs`.
    """
    _verify_auth(authorization)
    from services.file_convert import list_supported, list_supported_lo, _soffice_available
    return {
        "pairs": list_supported(),
        "lo_pairs": list_supported_lo(),
        "lo_available": _soffice_available(),
    }


# ── Adapter implementations ─────────────────────────────────────────────
# Each adapter is an async function: (action, params, creds) -> dict

async def _hubspot_adapter(action: str, params: dict, creds: dict) -> dict:
    token = creds["api_key"]
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    base = "https://api.hubapi.com"

    async with httpx.AsyncClient(timeout=30) as client:
        if action == "get_contacts":
            limit = params.get("limit", 20)
            r = await client.get(f"{base}/crm/v3/objects/contacts?limit={limit}&properties=firstname,lastname,email,phone", headers=headers)
            r.raise_for_status()
            contacts = [{"id": c["id"], **c.get("properties", {})} for c in r.json().get("results", [])]
            return {"contacts": contacts, "count": len(contacts)}

        if action == "create_contact":
            props = {}
            if params.get("email"): props["email"] = params["email"]
            if params.get("firstname"): props["firstname"] = params["firstname"]
            if params.get("lastname"): props["lastname"] = params["lastname"]
            if params.get("phone"): props["phone"] = params["phone"]
            if params.get("name"):
                parts = params["name"].split(" ", 1)
                props["firstname"] = parts[0]
                if len(parts) > 1: props["lastname"] = parts[1]
            r = await client.post(f"{base}/crm/v3/objects/contacts", headers=headers, json={"properties": props})
            r.raise_for_status()
            return {"contact": r.json(), "message": "Contact created"}

        if action == "get_deals":
            limit = params.get("limit", 20)
            r = await client.get(f"{base}/crm/v3/objects/deals?limit={limit}&properties=dealname,amount,dealstage,closedate", headers=headers)
            r.raise_for_status()
            deals = [{"id": d["id"], **d.get("properties", {})} for d in r.json().get("results", [])]
            return {"deals": deals, "count": len(deals)}

        if action == "create_deal":
            props = {}
            if params.get("name"): props["dealname"] = params["name"]
            if params.get("amount"): props["amount"] = params["amount"]
            if params.get("stage"): props["dealstage"] = params["stage"]
            r = await client.post(f"{base}/crm/v3/objects/deals", headers=headers, json={"properties": props})
            r.raise_for_status()
            return {"deal": r.json(), "message": "Deal created"}

        if action == "search":
            query = params.get("query", "")
            r = await client.post(f"{base}/crm/v3/objects/contacts/search", headers=headers, json={"query": query, "limit": 10})
            r.raise_for_status()
            return {"results": r.json().get("results", []), "count": r.json().get("total", 0)}

    return {"error": f"Unknown action: {action}"}


async def _salesforce_adapter(action: str, params: dict, creds: dict) -> dict:
    token = creds["api_key"]
    instance = creds.get("instance_url", "").rstrip("/")
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    base = f"{instance}/services/data/v59.0"

    async with httpx.AsyncClient(timeout=30) as client:
        if action == "get_leads":
            limit = params.get("limit", 20)
            r = await client.get(f"{base}/query?q=SELECT+Id,Name,Email,Phone,Status+FROM+Lead+LIMIT+{limit}", headers=headers)
            r.raise_for_status()
            return {"leads": r.json().get("records", []), "count": r.json().get("totalSize", 0)}

        if action == "create_lead":
            data = {"LastName": params.get("name", "Unknown")}
            if params.get("email"): data["Email"] = params["email"]
            if params.get("phone"): data["Phone"] = params["phone"]
            if params.get("company"): data["Company"] = params["company"]
            r = await client.post(f"{base}/sobjects/Lead", headers=headers, json=data)
            r.raise_for_status()
            return {"lead": r.json(), "message": "Lead created"}

        if action == "get_opportunities":
            limit = params.get("limit", 20)
            r = await client.get(f"{base}/query?q=SELECT+Id,Name,Amount,StageName+FROM+Opportunity+LIMIT+{limit}", headers=headers)
            r.raise_for_status()
            return {"opportunities": r.json().get("records", []), "count": r.json().get("totalSize", 0)}

        if action == "create_case":
            data = {"Subject": params.get("subject", "New Case")}
            if params.get("description"): data["Description"] = params["description"]
            r = await client.post(f"{base}/sobjects/Case", headers=headers, json=data)
            r.raise_for_status()
            return {"case": r.json(), "message": "Case created"}

        if action == "search":
            query = params.get("query", "")
            r = await client.get(f"{base}/search?q=FIND+{{{query}}}+IN+ALL+FIELDS+RETURNING+Lead,Contact,Opportunity", headers=headers)
            r.raise_for_status()
            return {"results": r.json().get("searchRecords", [])}

    return {"error": f"Unknown action: {action}"}


async def _ringy_adapter(action: str, params: dict, creds: dict) -> dict:
    """Ringy CRM adapter.

    Ringy's public API uses POST requests with the API key in the JSON body
    (not a header), and every lookup is per-ID rather than collection-based.
    We expose the 5 documented endpoints:

      - get_lead              → POST /get-lead              (one lead by id)
      - get_call              → POST /get-calls             (one call by id)
      - get_sold_products     → POST /get-lead-sold-products (date-range list)
      - get_call_recordings   → POST /get-call-recordings    (date-range list)
      - create_appointment    → POST /create-appointment     (write)

    Note: Ringy does NOT offer a "list all leads" endpoint. The closest
    thing is to call get_sold_products for a date range and walk the
    returned leadIds — callers that want "recent leads" can do that.
    """
    api_key = creds.get("api_key")
    if not api_key:
        # User has an old SID/Auth Token-based connection from the Lead Vendor
        # era. Tell them to reconnect with a real API key.
        return {
            "error": (
                "Ringy is connected with old credentials. Please reconnect "
                "in Settings → My Apps → Ringy using an API Key from "
                "Ringy → Settings → Account Settings → Manage Account → API Keys."
            )
        }

    base = "https://app.ringy.com/api/public/external"

    async with httpx.AsyncClient(timeout=30) as client:
        if action == "get_lead":
            lead_id = params.get("lead_id") or params.get("leadId")
            if not lead_id:
                return {"error": "lead_id is required"}
            r = await client.post(
                f"{base}/get-lead",
                json={"apiKey": api_key, "leadId": lead_id},
            )
            r.raise_for_status()
            return {"lead": r.json()}

        if action == "get_call":
            call_id = params.get("call_id") or params.get("callId")
            if not call_id:
                return {"error": "call_id is required"}
            r = await client.post(
                f"{base}/get-calls",
                json={"apiKey": api_key, "callId": call_id},
            )
            r.raise_for_status()
            return {"call": r.json()}

        if action == "get_sold_products":
            start = params.get("start_date") or params.get("startDate")
            end = params.get("end_date") or params.get("endDate")
            # Default to the last 30 days if the caller didn't specify a range
            if not start or not end:
                from datetime import datetime as _dt, timedelta as _td
                _now = _dt.utcnow()
                end = _now.strftime("%Y-%m-%d %H:%M:%S")
                start = (_now - _td(days=30)).strftime("%Y-%m-%d 00:00:00")
            body: dict = {"apiKey": api_key, "startDate": start, "endDate": end}
            if params.get("limit"):
                body["limit"] = int(params["limit"])
            r = await client.post(f"{base}/get-lead-sold-products", json=body)
            r.raise_for_status()
            data = r.json() or []
            return {
                "sold_products": data,
                "count": len(data),
                "start_date": start,
                "end_date": end,
            }

        if action == "get_call_recordings":
            start = params.get("start_date") or params.get("startDate")
            end = params.get("end_date") or params.get("endDate")
            if not start or not end:
                from datetime import datetime as _dt, timedelta as _td
                _now = _dt.utcnow()
                end = _now.strftime("%Y-%m-%d %H:%M:%S")
                start = (_now - _td(days=30)).strftime("%Y-%m-%d 00:00:00")
            body = {"apiKey": api_key, "startDate": start, "endDate": end}
            if params.get("limit"):
                body["limit"] = int(params["limit"])
            r = await client.post(f"{base}/get-call-recordings", json=body)
            r.raise_for_status()
            data = r.json() or []
            return {
                "recordings": data,
                "count": len(data),
                "start_date": start,
                "end_date": end,
            }

        if action == "create_appointment":
            start_time = params.get("start")
            if not start_time:
                return {
                    "error": "start time is required (format: YYYY-MM-DD HH:mm:ss, UTC)"
                }
            # Need EITHER leadId OR leadPhoneNumber per Ringy's docs
            lead_id = params.get("lead_id") or params.get("leadId")
            lead_phone = params.get("lead_phone") or params.get("leadPhoneNumber")
            if not lead_id and not lead_phone:
                return {"error": "Either lead_id or lead_phone is required"}
            body = {"apiKey": api_key, "start": start_time}
            if lead_id:
                body["leadId"] = lead_id
            if lead_phone:
                body["leadPhoneNumber"] = lead_phone
            if params.get("lead_first_name"):
                body["leadFirstName"] = params["lead_first_name"]
            if params.get("lead_last_name"):
                body["leadLastName"] = params["lead_last_name"]
            if params.get("lead_email"):
                body["leadEmail"] = params["lead_email"]
            if params.get("comments"):
                body["comments"] = params["comments"]
            if params.get("duration_minutes"):
                body["durationInMinutes"] = int(params["duration_minutes"])
            r = await client.post(f"{base}/create-appointment", json=body)
            r.raise_for_status()
            data = r.json() or {}
            if data.get("status") == 200:
                return {"message": "Appointment created"}
            return {
                "error": data.get("message")
                or f"Ringy returned status {data.get('status')}"
            }

    return {"error": f"Unknown Ringy action: {action}"}


async def _excel_online_adapter(action: str, params: dict, creds: dict) -> dict:
    """Microsoft Excel Online adapter via Microsoft Graph API.

    Auth: OAuth 2.0 bearer token from Azure AD. The token must be scoped for
    Files.ReadWrite (and Sites.ReadWrite.All if the file is on SharePoint).

    Exposed actions:
      - list_workbooks    → find .xlsx files in the user's OneDrive
      - get_worksheets    → list sheets in a workbook
      - read_range        → read a cell range from a worksheet
      - write_range       → overwrite a cell range with a 2D array
      - add_row           → append a single row to a worksheet's used range
      - create_workbook   → create a new empty .xlsx in OneDrive root
    """
    token = creds.get("access_token") or creds.get("api_key")
    if not token:
        return {
            "error": (
                "Excel Online is not connected. Please add a Microsoft Graph "
                "access token in Settings → My Apps → Microsoft Excel Online."
            )
        }

    base = "https://graph.microsoft.com/v1.0"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }

    def _base_name(n: str) -> str:
        """Strip repeated .xlsx extensions and lowercase for fuzzy matching."""
        s = (n or "").lower().strip()
        while s.endswith(".xlsx"):
            s = s[:-5]
        # Normalize whitespace and separators so "test budget" == "test-budget"
        for ch in (" ", "_", "-"):
            s = s.replace(ch, "")
        return s

    async def _collect_all_xlsx(client: httpx.AsyncClient) -> list[dict]:
        """List every .xlsx file we can find in the user's OneDrive."""
        seen: dict[str, dict] = {}
        # Root children
        r = await client.get(
            f"{base}/me/drive/root/children",
            headers=headers,
            params={"$top": 200, "$select": "id,name,folder,webUrl,lastModifiedDateTime"},
        )
        subfolder_ids: list[str] = []
        if r.status_code == 200:
            for it in (r.json() or {}).get("value", []):
                if it.get("folder"):
                    if it.get("id"):
                        subfolder_ids.append(it["id"])
                    continue
                name = it.get("name") or ""
                if name.lower().endswith(".xlsx") and it.get("id"):
                    seen[it["id"]] = it
        # One level of subfolders
        for fid in subfolder_ids[:20]:
            rr = await client.get(
                f"{base}/me/drive/items/{fid}/children",
                headers=headers,
                params={"$top": 200, "$select": "id,name,webUrl,lastModifiedDateTime"},
            )
            if rr.status_code == 200:
                for it in (rr.json() or {}).get("value", []):
                    name = it.get("name") or ""
                    if name.lower().endswith(".xlsx") and it.get("id"):
                        seen.setdefault(it["id"], it)
        # Also try the search index — may find files not in root
        rs = await client.get(
            f"{base}/me/drive/root/search(q='.xlsx')",
            headers=headers,
            params={"$top": 50, "$select": "id,name,webUrl,lastModifiedDateTime"},
        )
        if rs.status_code == 200:
            for it in (rs.json() or {}).get("value", []):
                name = it.get("name") or ""
                if name.lower().endswith(".xlsx") and it.get("id"):
                    seen.setdefault(it["id"], it)
        return list(seen.values())

    async def _resolve_workbook_id(client: httpx.AsyncClient, workbook_ref: str) -> tuple[Optional[str], Optional[str]]:
        """Resolve a filename, partial name, or item id to a Graph item id.

        Behavior:
          - If workbook_ref looks like a Graph item id → use as-is.
          - Otherwise, list all .xlsx files and find the best match by fuzzy
            base-name (ignores case, hyphens, spaces, duplicate .xlsx extensions).
          - If the user gave no ref AND there's only one .xlsx file in the
            drive → pick it automatically.
          - If multiple files match, return the list of candidate names.
        """
        # Graph item ids are typically 40+ chars starting with "01" and contain
        # no dots, slashes, or spaces. Anything else → treat as filename.
        looks_like_item_id = (
            bool(workbook_ref)
            and len(workbook_ref) > 20
            and workbook_ref.startswith("01")
            and "." not in workbook_ref
            and "/" not in workbook_ref
            and " " not in workbook_ref
        )
        if looks_like_item_id:
            return workbook_ref, None

        all_workbooks = await _collect_all_xlsx(client)

        # No files at all
        if not all_workbooks:
            return None, (
                "I don't see any Excel files in your OneDrive. "
                "Create one at onedrive.live.com first, or ask me to create a new workbook."
            )

        # User didn't specify a name → auto-pick when there's only one file
        if not workbook_ref:
            if len(all_workbooks) == 1:
                return all_workbooks[0]["id"], None
            names = ", ".join([w.get("name", "") for w in all_workbooks[:10]])
            return None, f"You have multiple Excel files. Which one? {names}"

        target = _base_name(workbook_ref)

        # Exact base-name match
        exact = [w for w in all_workbooks if _base_name(w.get("name", "")) == target]
        if len(exact) == 1:
            return exact[0]["id"], None
        if len(exact) > 1:
            names = ", ".join([w.get("name", "") for w in exact[:10]])
            return None, f"Multiple files match '{workbook_ref}': {names}. Please be more specific."

        # Substring match (either direction)
        partial = [
            w for w in all_workbooks
            if target in _base_name(w.get("name", ""))
            or _base_name(w.get("name", "")) in target
        ]
        if len(partial) == 1:
            return partial[0]["id"], None
        if len(partial) > 1:
            names = ", ".join([w.get("name", "") for w in partial[:10]])
            return None, f"Multiple files match '{workbook_ref}': {names}. Please pick one."

        # Nothing matched
        all_names = ", ".join([w.get("name", "") for w in all_workbooks[:10]])
        return None, (
            f"No Excel file matches '{workbook_ref}'. "
            f"Your files: {all_names}. Use one of these exact names."
        )

    async with httpx.AsyncClient(timeout=30) as client:
        if action == "list_workbooks":
            workbooks: list[dict] = []
            seen_ids: set[str] = set()

            def _add(it: dict) -> None:
                name = it.get("name") or ""
                if not name.lower().endswith(".xlsx"):
                    return
                wid = it.get("id")
                if not wid or wid in seen_ids:
                    return
                seen_ids.add(wid)
                workbooks.append({
                    "id": wid,
                    "name": name,
                    "url": it.get("webUrl"),
                    "modified": it.get("lastModifiedDateTime"),
                    "size": it.get("size"),
                })

            # 1) Search the drive (fastest when the index is warm)
            r = await client.get(
                f"{base}/me/drive/root/search(q='.xlsx')",
                headers=headers,
                params={"$top": 50, "$select": "id,name,webUrl,lastModifiedDateTime,size"},
            )
            if r.status_code == 401:
                return {"error": "Microsoft Graph token is invalid or expired. Reconnect in Settings → My Apps → Microsoft Excel Online."}
            if r.status_code == 200:
                for it in (r.json() or {}).get("value", []):
                    _add(it)

            # 2) Fallback: list the root folder children directly (handles brand-new
            # files that aren't indexed yet — search can take several minutes).
            if not workbooks:
                r = await client.get(
                    f"{base}/me/drive/root/children",
                    headers=headers,
                    params={"$top": 100, "$select": "id,name,webUrl,lastModifiedDateTime,size,folder"},
                )
                if r.status_code == 200:
                    for it in (r.json() or {}).get("value", []):
                        if it.get("folder"):
                            continue  # skip directories for this first pass
                        _add(it)

            # 3) Last-ditch: scan top-level folders one level deep
            if not workbooks:
                r = await client.get(
                    f"{base}/me/drive/root/children",
                    headers=headers,
                    params={"$top": 50, "$select": "id,name,folder"},
                )
                if r.status_code == 200:
                    for folder in (r.json() or {}).get("value", []):
                        if not folder.get("folder"):
                            continue
                        fid = folder.get("id")
                        rr = await client.get(
                            f"{base}/me/drive/items/{fid}/children",
                            headers=headers,
                            params={"$top": 50, "$select": "id,name,webUrl,lastModifiedDateTime,size"},
                        )
                        if rr.status_code == 200:
                            for it in (rr.json() or {}).get("value", []):
                                _add(it)

            return {"workbooks": workbooks, "count": len(workbooks)}

        if action == "get_worksheets":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("name") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            r = await client.get(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets",
                headers=headers,
            )
            if r.status_code == 401:
                return {"error": "Microsoft Graph token is invalid or expired."}
            r.raise_for_status()
            sheets = (r.json() or {}).get("value", [])
            return {
                "worksheets": [
                    {"id": s.get("id"), "name": s.get("name"), "position": s.get("position"), "visibility": s.get("visibility")}
                    for s in sheets
                ],
                "count": len(sheets),
            }

        if action == "read_range":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("name") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            cell_range = params.get("range") or params.get("address")
            if not cell_range:
                return {"error": "range is required (e.g. A1:C10)"}
            r = await client.get(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/range(address='{cell_range}')",
                headers=headers,
            )
            if r.status_code == 401:
                return {"error": "Microsoft Graph token is invalid or expired."}
            r.raise_for_status()
            data = r.json() or {}
            return {
                "range": cell_range,
                "worksheet": worksheet,
                "values": data.get("values", []),
                "text": data.get("text", []),
                "formulas": data.get("formulas", []),
                "rowCount": data.get("rowCount"),
                "columnCount": data.get("columnCount"),
            }

        if action == "write_range":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("name") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            cell_range = params.get("range") or params.get("address")
            values = params.get("values")
            if isinstance(values, str):
                try:
                    values = json.loads(values)
                except Exception:
                    return {"error": "values must be a JSON 2D array, e.g. [[\"a\",1],[\"b\",2]]"}
            if not cell_range or values is None:
                return {"error": "range and values are required"}
            r = await client.patch(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/range(address='{cell_range}')",
                headers=headers,
                json={"values": values},
            )
            if r.status_code == 401:
                return {"error": "Microsoft Graph token is invalid or expired."}
            r.raise_for_status()
            return {"message": f"Wrote values to {worksheet}!{cell_range}", "range": cell_range}

        if action == "add_row":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("name") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            values = params.get("values")
            if isinstance(values, str):
                try:
                    values = json.loads(values)
                except Exception:
                    # Allow a simple comma-separated list as a fallback so the AI
                    # can say values="coffee,50" instead of a JSON array
                    values = [v.strip() for v in values.split(",")]
            if not isinstance(values, list):
                return {"error": "values must be a JSON array, e.g. [\"name\",42,\"note\"]"}
            # First fetch the used range so we can append below it
            u = await client.get(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/usedRange(valuesOnly=true)",
                headers=headers,
            )
            if u.status_code == 401:
                return {"error": "Microsoft Graph token is invalid or expired."}
            if u.status_code == 404:
                return {"error": f"Worksheet '{worksheet}' not found in the workbook"}
            if u.status_code >= 400:
                # Sheet is probably empty — start at row 1
                row_count = 0
                col_count = len(values)
            else:
                used = u.json() or {}
                row_count = used.get("rowCount") or 0
                col_count = used.get("columnCount") or len(values)
            next_row = row_count + 1
            # Pad or trim values to match column count so the Graph API accepts it
            if len(values) < col_count:
                values = values + [""] * (col_count - len(values))
            elif len(values) > col_count:
                col_count = len(values)
            # Build A1 address for the target row — works up to column ZZ
            def col_letter(n: int) -> str:
                s = ""
                while n > 0:
                    n, rem = divmod(n - 1, 26)
                    s = chr(65 + rem) + s
                return s
            last_col = col_letter(col_count)
            target = f"A{next_row}:{last_col}{next_row}"
            r = await client.patch(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/range(address='{target}')",
                headers=headers,
                json={"values": [values]},
            )
            r.raise_for_status()
            return {"message": f"Appended row to {worksheet}!{target}", "range": target}

        # ── Simple reads ────────────────────────────────────────────────
        if action == "get_cell_value":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("name") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            cell = params.get("cell") or params.get("address") or params.get("range")
            if not cell:
                return {"error": "cell is required (e.g. B2)"}
            r = await client.get(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/range(address='{cell}')",
                headers=headers,
            )
            if r.status_code >= 400:
                return {"error": f"Could not read {cell}: {r.text[:200]}"}
            data = r.json() or {}
            values = data.get("values") or [[]]
            return {
                "cell": cell,
                "value": values[0][0] if values and values[0] else None,
                "text": (data.get("text") or [[None]])[0][0],
                "formula": (data.get("formulas") or [[None]])[0][0],
            }

        # ── Sheet management ────────────────────────────────────────────
        if action == "add_worksheet":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            sheet_name = params.get("name") or params.get("worksheet") or "NewSheet"
            r = await client.post(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets/add",
                headers=headers,
                json={"name": sheet_name},
            )
            if r.status_code >= 400:
                return {"error": f"Could not add worksheet: {r.text[:200]}"}
            return {"message": f"Added worksheet '{sheet_name}'", "worksheet": r.json()}

        if action == "rename_worksheet":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            current = params.get("worksheet") or params.get("sheet") or "Sheet1"
            new_name = params.get("name") or params.get("new_name")
            if not new_name:
                return {"error": "name is required (the new sheet name)"}
            r = await client.patch(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{current}')",
                headers=headers,
                json={"name": new_name},
            )
            if r.status_code >= 400:
                return {"error": f"Could not rename: {r.text[:200]}"}
            return {"message": f"Renamed '{current}' to '{new_name}'"}

        if action == "delete_worksheet":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            sheet = params.get("worksheet") or params.get("sheet") or params.get("name")
            if not sheet:
                return {"error": "worksheet is required"}
            r = await client.delete(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{sheet}')",
                headers=headers,
            )
            if r.status_code >= 400:
                return {"error": f"Could not delete worksheet: {r.text[:200]}"}
            return {"message": f"Deleted worksheet '{sheet}'"}

        if action == "copy_worksheet":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            source = params.get("worksheet") or params.get("sheet") or "Sheet1"
            new_name = params.get("name") or f"{source} copy"
            r = await client.post(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{source}')/copy",
                headers=headers,
                json={"positionType": "End"},
            )
            if r.status_code >= 400:
                return {"error": f"Could not copy: {r.text[:200]}"}
            created = r.json() or {}
            # Rename the copy if user specified a name
            if params.get("name"):
                created_name = created.get("name")
                if created_name:
                    await client.patch(
                        f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{created_name}')",
                        headers=headers,
                        json={"name": new_name},
                    )
            return {"message": f"Copied '{source}' to '{new_name}'"}

        # ── Row / range manipulation ────────────────────────────────────
        if action == "delete_row":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            row = params.get("row") or params.get("row_number")
            if row is None:
                return {"error": "row is required (1-based row number)"}
            try:
                row_num = int(row)
            except Exception:
                return {"error": "row must be an integer"}
            # Delete the entire row by addressing it and shifting up
            r = await client.post(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/range(address='{row_num}:{row_num}')/delete",
                headers=headers,
                json={"shift": "Up"},
            )
            if r.status_code >= 400:
                return {"error": f"Could not delete row: {r.text[:200]}"}
            return {"message": f"Deleted row {row_num} from {worksheet}"}

        if action == "clear_range":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            cell_range = params.get("range") or params.get("address")
            if not cell_range:
                return {"error": "range is required"}
            r = await client.post(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/range(address='{cell_range}')/clear",
                headers=headers,
                json={"applyTo": "Contents"},
            )
            if r.status_code >= 400:
                return {"error": f"Could not clear range: {r.text[:200]}"}
            return {"message": f"Cleared {worksheet}!{cell_range}"}

        if action == "set_formula":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            cell = params.get("cell") or params.get("address") or params.get("range")
            formula = params.get("formula")
            if not cell or not formula:
                return {"error": "cell and formula are required"}
            if not formula.startswith("="):
                formula = "=" + formula
            r = await client.patch(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/range(address='{cell}')",
                headers=headers,
                json={"formulas": [[formula]]},
            )
            if r.status_code >= 400:
                return {"error": f"Could not set formula: {r.text[:200]}"}
            return {"message": f"Set formula {formula} in {worksheet}!{cell}"}

        # ── Tables ──────────────────────────────────────────────────────
        if action == "create_table":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            cell_range = params.get("range") or params.get("address")
            if not cell_range:
                return {"error": "range is required (e.g. A1:D10 with headers on row 1)"}
            table_name = params.get("name") or params.get("table") or "Table1"
            r = await client.post(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/tables/add",
                headers=headers,
                json={"address": f"{worksheet}!{cell_range}", "hasHeaders": True},
            )
            if r.status_code >= 400:
                return {"error": f"Could not create table: {r.text[:200]}"}
            created = r.json() or {}
            created_id = created.get("id")
            # Rename the table to the requested name
            if created_id and params.get("name"):
                await client.patch(
                    f"{base}/me/drive/items/{workbook_id}/workbook/tables/{created_id}",
                    headers=headers,
                    json={"name": table_name},
                )
            return {"message": f"Created table '{table_name}' on {worksheet}!{cell_range}"}

        if action == "add_table_row":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            table = params.get("table") or params.get("name")
            if not table:
                return {"error": "table name is required"}
            values = params.get("values")
            if isinstance(values, str):
                try:
                    values = json.loads(values)
                except Exception:
                    values = [v.strip() for v in values.split(",")]
            if not isinstance(values, list):
                return {"error": "values must be a list"}
            # Tables API expects a 2D array of rows
            if values and not isinstance(values[0], list):
                values = [values]
            r = await client.post(
                f"{base}/me/drive/items/{workbook_id}/workbook/tables/{table}/rows/add",
                headers=headers,
                json={"values": values},
            )
            if r.status_code >= 400:
                return {"error": f"Could not add table row: {r.text[:200]}"}
            return {"message": f"Added row to table '{table}'"}

        # ── Formatting ──────────────────────────────────────────────────
        if action == "format_range":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            cell_range = params.get("range") or params.get("address")
            if not cell_range:
                return {"error": "range is required"}
            range_base = f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/range(address='{cell_range}')"
            # Font (bold, italic, color)
            font_patch: dict = {}
            if "bold" in params:
                font_patch["bold"] = str(params["bold"]).lower() == "true"
            if "italic" in params:
                font_patch["italic"] = str(params["italic"]).lower() == "true"
            if params.get("color"):
                font_patch["color"] = params["color"]
            if font_patch:
                await client.patch(f"{range_base}/format/font", headers=headers, json=font_patch)
            # Fill (background color)
            if params.get("fill"):
                await client.patch(f"{range_base}/format/fill", headers=headers, json={"color": params["fill"]})
            return {"message": f"Formatted {worksheet}!{cell_range}"}

        if action == "set_number_format":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            cell_range = params.get("range") or params.get("address")
            fmt = params.get("format") or params.get("number_format")
            if not cell_range or not fmt:
                return {"error": "range and format are required"}
            r = await client.patch(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/range(address='{cell_range}')",
                headers=headers,
                json={"numberFormat": [[fmt]]},
            )
            if r.status_code >= 400:
                return {"error": f"Could not set number format: {r.text[:200]}"}
            return {"message": f"Set number format {fmt} on {worksheet}!{cell_range}"}

        if action == "autofit_columns":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            cell_range = params.get("range") or params.get("address") or "A:Z"
            r = await client.post(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/range(address='{cell_range}')/format/autofitColumns",
                headers=headers,
            )
            if r.status_code >= 400:
                return {"error": f"Could not autofit columns: {r.text[:200]}"}
            return {"message": f"Autofit columns on {cell_range}"}

        # ── Charts ──────────────────────────────────────────────────────
        if action == "create_chart":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            cell_range = params.get("range") or params.get("address")
            if not cell_range:
                return {"error": "range is required"}
            chart_type = params.get("type") or params.get("chart_type") or "ColumnClustered"
            r = await client.post(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/charts/add",
                headers=headers,
                json={
                    "type": chart_type,
                    "sourceData": f"{worksheet}!{cell_range}",
                    "seriesBy": "Auto",
                },
            )
            if r.status_code >= 400:
                return {"error": f"Could not create chart: {r.text[:200]}"}
            chart = r.json() or {}
            chart_name = chart.get("name")
            # Set the title if given
            if params.get("title") and chart_name:
                await client.patch(
                    f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/charts('{chart_name}')/title",
                    headers=headers,
                    json={"text": params["title"]},
                )
            return {"message": f"Created {chart_type} chart on {worksheet}", "chart": chart_name}

        if action == "delete_chart":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            chart = params.get("chart") or params.get("name")
            if not chart:
                return {"error": "chart name is required"}
            r = await client.delete(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/charts('{chart}')",
                headers=headers,
            )
            if r.status_code >= 400:
                return {"error": f"Could not delete chart: {r.text[:200]}"}
            return {"message": f"Deleted chart '{chart}'"}

        # ── Queries & analysis ──────────────────────────────────────────
        if action == "find_cell":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            search = params.get("search") or params.get("query")
            if not search:
                return {"error": "search is required"}
            # Pull the used range and scan it locally (Graph doesn't have a first-class find endpoint)
            u = await client.get(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/usedRange(valuesOnly=true)",
                headers=headers,
            )
            if u.status_code >= 400:
                return {"error": "Worksheet is empty"}
            used = u.json() or {}
            values = used.get("values") or []
            # Graph returns the used-range top-left in `address` like "Sheet1!A1"
            addr = used.get("address", f"{worksheet}!A1")
            try:
                top_left = addr.split("!")[-1].split(":")[0]
                # Extract row/column from top_left (e.g. A1 → col=A, row=1)
                col_prefix = "".join(c for c in top_left if c.isalpha()) or "A"
                row_prefix = int("".join(c for c in top_left if c.isdigit()) or 1)
            except Exception:
                col_prefix, row_prefix = "A", 1
            matches: list[dict] = []
            needle = str(search).lower()
            for r_idx, row in enumerate(values):
                for c_idx, val in enumerate(row):
                    if val is None:
                        continue
                    if needle in str(val).lower():
                        # Compute A1 address of this cell
                        col_idx = sum((ord(c) - 64) * (26 ** i) for i, c in enumerate(reversed(col_prefix.upper()))) + c_idx
                        cell_col = ""
                        n = col_idx
                        while n > 0:
                            n, rem = divmod(n - 1, 26)
                            cell_col = chr(65 + rem) + cell_col
                        matches.append({
                            "cell": f"{cell_col}{row_prefix + r_idx}",
                            "value": val,
                        })
            return {"matches": matches[:50], "count": len(matches)}

        if action == "sum_column":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            column = (params.get("column") or "").strip().upper()
            if not column:
                return {"error": "column letter is required (e.g. B)"}
            # Read the used range, then sum the requested column locally.
            # This is more reliable than Graph's function API, which is picky
            # about range formats and fails on full-column references.
            u = await client.get(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/usedRange(valuesOnly=true)",
                headers=headers,
            )
            if u.status_code == 401:
                return {"error": "Microsoft Graph token is invalid or expired."}
            if u.status_code >= 400:
                return {"error": f"Worksheet '{worksheet}' is empty or unreadable"}
            used = u.json() or {}
            values = used.get("values") or []
            if not values:
                return {"column": column, "sum": 0, "count": 0}
            # Figure out which index the requested column maps to inside the used range
            addr = used.get("address", f"{worksheet}!A1")
            try:
                top_left = addr.split("!")[-1].split(":")[0]
                start_col_letters = "".join(c for c in top_left if c.isalpha()) or "A"
            except Exception:
                start_col_letters = "A"
            def _letters_to_idx(letters: str) -> int:
                n = 0
                for c in letters.upper():
                    n = n * 26 + (ord(c) - 64)
                return n
            start_idx = _letters_to_idx(start_col_letters)
            target_idx = _letters_to_idx(column)
            col_offset = target_idx - start_idx
            if col_offset < 0:
                return {"column": column, "sum": 0, "count": 0}
            # Sum numeric cells in the chosen column, skipping a header if the
            # first row contains non-numeric text in that slot
            total = 0.0
            count = 0
            rows_iter = values
            # Skip header row if the first cell in this column isn't a number
            if rows_iter:
                first_val = rows_iter[0][col_offset] if col_offset < len(rows_iter[0]) else None
                if first_val is not None and not isinstance(first_val, (int, float)):
                    rows_iter = rows_iter[1:]
            for row in rows_iter:
                if col_offset >= len(row):
                    continue
                val = row[col_offset]
                if isinstance(val, (int, float)):
                    total += float(val)
                    count += 1
                elif isinstance(val, str):
                    # Handle strings that look like numbers ("1500", "$299.99")
                    cleaned = val.replace("$", "").replace(",", "").strip()
                    if cleaned:
                        try:
                            total += float(cleaned)
                            count += 1
                        except ValueError:
                            pass
            # Cast to int when the result is whole so the display is clean
            final: float | int = int(total) if total == int(total) else round(total, 2)
            return {"column": column, "sum": final, "count": count, "worksheet": worksheet}

        if action == "get_last_row":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            u = await client.get(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/usedRange(valuesOnly=true)",
                headers=headers,
            )
            if u.status_code >= 400:
                return {"last_row": 0, "message": "Worksheet is empty"}
            used = u.json() or {}
            row_count = used.get("rowCount") or 0
            addr = used.get("address", "")
            # Extract the ending row from addresses like "Sheet1!A1:C10"
            try:
                end_addr = addr.split("!")[-1].split(":")[-1]
                end_row = int("".join(c for c in end_addr if c.isdigit()))
            except Exception:
                end_row = row_count
            return {"last_row": end_row, "row_count": row_count}

        if action == "filter_rows":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            column = (params.get("column") or "").strip().upper()
            match = params.get("value") or params.get("match")
            if not column or match is None:
                return {"error": "column and value are required"}
            # Read the full used range and filter locally
            u = await client.get(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/usedRange(valuesOnly=true)",
                headers=headers,
            )
            if u.status_code >= 400:
                return {"error": "Worksheet is empty or unreadable"}
            used = u.json() or {}
            values = used.get("values") or []
            if not values:
                return {"rows": [], "count": 0}
            headers_row = values[0]
            # Convert column letter to 0-based index
            col_idx = 0
            for c in column:
                col_idx = col_idx * 26 + (ord(c) - 64)
            col_idx -= 1
            needle = str(match).lower()
            matching = [row for row in values[1:] if col_idx < len(row) and needle in str(row[col_idx] or "").lower()]
            return {
                "headers": headers_row,
                "rows": matching,
                "count": len(matching),
            }

        if action == "sort_range":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            cell_range = params.get("range") or params.get("address")
            if not cell_range:
                return {"error": "range is required"}
            try:
                sort_col = int(params.get("column", 0))
            except Exception:
                sort_col = 0
            ascending = str(params.get("ascending", "true")).lower() != "false"
            r = await client.post(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/range(address='{cell_range}')/sort/apply",
                headers=headers,
                json={
                    "fields": [{
                        "key": sort_col,
                        "sortOn": "Value",
                        "ascending": ascending,
                    }],
                    "hasHeaders": True,
                },
            )
            if r.status_code >= 400:
                return {"error": f"Could not sort: {r.text[:200]}"}
            return {"message": f"Sorted {worksheet}!{cell_range} by column {sort_col}"}

        # ── File-level utilities ────────────────────────────────────────
        if action == "download_as_pdf":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            r = await client.get(
                f"{base}/me/drive/items/{workbook_id}/content?format=pdf",
                headers={"Authorization": f"Bearer {token}"},
                follow_redirects=False,
            )
            # Graph returns a 302 redirect to the actual file content
            if r.status_code in (302, 301):
                return {"pdf_url": r.headers.get("Location"), "download_url": r.headers.get("Location"), "filename": "workbook.pdf", "mime_type": "application/pdf", "message": "PDF ready to download"}
            if r.status_code == 200:
                return {"message": "PDF generated (content inline, not URL)"}
            return {"error": f"Could not generate PDF: {r.status_code} {r.text[:200]}"}

        if action == "download_workbook":
            # Return the workbook for download in whatever format the user
            # asked for. xlsx and pdf are served as short-lived Graph URLs
            # (no conversion needed); everything else runs through the
            # LibreOffice-backed convert_bytes_async pipeline and is
            # returned inline as base64 so the client can write it straight
            # to cache without another round trip.
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}

            target_fmt = (params.get("format") or params.get("as") or params.get("to") or "xlsx").strip().lower().lstrip(".")
            # Normalize a few friendly aliases
            target_fmt = {"excel": "xlsx", "spreadsheet": "xlsx", "word": "docx", "text": "txt", "markdown": "md"}.get(target_fmt, target_fmt)

            # Fetch minimal metadata for name/size. We do NOT use $select
            # here because some Graph tenants drop the @microsoft.graph.downloadUrl
            # annotation when $select is used — and we want it in either
            # case for fallback. We'll get the actual download URL from
            # /content below regardless.
            r = await client.get(
                f"{base}/me/drive/items/{workbook_id}",
                headers=headers,
            )
            if r.status_code >= 400:
                return {"error": f"Could not fetch file metadata: {r.status_code} {r.text[:200]}"}
            meta = r.json() or {}
            raw_name = meta.get("name") or "workbook.xlsx"
            stem = raw_name.rsplit(".", 1)[0]

            async def _graph_content_url(format_hint: str | None = None) -> tuple[str | None, str | None]:
                """Ask Graph for the short-lived download URL via /content.
                Graph responds with a 302 whose Location is a pre-authed
                CDN URL. Works on every tenant, no $select weirdness."""
                url = f"{base}/me/drive/items/{workbook_id}/content"
                if format_hint:
                    url += f"?format={format_hint}"
                rr = await client.get(
                    url,
                    headers={"Authorization": f"Bearer {token}"},
                    follow_redirects=False,
                )
                if rr.status_code in (301, 302):
                    loc = rr.headers.get("Location") or rr.headers.get("location")
                    return loc, None
                return None, f"Graph /content returned {rr.status_code}: {rr.text[:200]}"

            # Fast path: raw .xlsx — use /content (no format param) to get
            # the pre-authed download URL. Fall back to the metadata's
            # downloadUrl annotation if /content doesn't redirect.
            if target_fmt in ("xlsx", ""):
                dl, cerr = await _graph_content_url(None)
                if not dl:
                    dl = meta.get("@microsoft.graph.downloadUrl") or meta.get("@content.downloadUrl")
                if not dl:
                    return {"error": cerr or "Could not obtain a download URL for this file"}
                return {
                    "download_url": dl,
                    "filename": raw_name,
                    "size": meta.get("size"),
                    "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "message": f"{raw_name} ready to download",
                }

            # Fast path: PDF via Graph's native export
            if target_fmt == "pdf":
                dl, cerr = await _graph_content_url("pdf")
                if dl:
                    return {
                        "download_url": dl,
                        "filename": f"{stem}.pdf",
                        "mime_type": "application/pdf",
                        "message": f"{stem}.pdf ready to download",
                    }
                return {"error": cerr or "Graph PDF export failed"}

            # Slow path: fetch the raw xlsx bytes, then convert server-side
            # to whatever target the user asked for (csv, txt, docx, ods,
            # html, png, etc.). The converted bytes come back inline as
            # base64 so the chat client can write them to cacheDirectory.
            src_url, cerr = await _graph_content_url(None)
            if not src_url:
                src_url = meta.get("@microsoft.graph.downloadUrl") or meta.get("@content.downloadUrl")
            if not src_url:
                return {"error": cerr or "Could not obtain the source xlsx download URL"}
            rb = await client.get(src_url, headers={"Authorization": f"Bearer {token}"}, follow_redirects=True)
            if rb.status_code != 200:
                return {"error": f"Could not fetch source xlsx bytes: {rb.status_code}"}
            try:
                from services.file_convert import convert_bytes_async as _convert_bytes_async
                out_bytes, out_mime, out_filename = await _convert_bytes_async(
                    rb.content, "xlsx", target_fmt, out_name=stem,
                )
            except Exception as e:
                return {"error": f"Could not convert to {target_fmt}: {e}"}
            import base64 as _b64
            return {
                "content_base64": _b64.b64encode(out_bytes).decode("ascii"),
                "filename": out_filename,
                "size": len(out_bytes),
                "mime_type": out_mime,
                "message": f"{out_filename} ready to download",
            }

        if action == "protect_sheet":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            r = await client.post(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/protection/protect",
                headers=headers,
                json={},
            )
            if r.status_code >= 400:
                return {"error": f"Could not protect: {r.text[:200]}"}
            return {"message": f"Protected worksheet '{worksheet}'"}

        if action == "share_workbook":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            scope = params.get("scope") or "view"  # view or edit
            link_type = "view" if scope.lower() == "view" else "edit"
            r = await client.post(
                f"{base}/me/drive/items/{workbook_id}/createLink",
                headers=headers,
                json={"type": link_type, "scope": "anonymous"},
            )
            if r.status_code >= 400:
                return {"error": f"Could not create share link: {r.text[:200]}"}
            data = r.json() or {}
            link = (data.get("link") or {}).get("webUrl")
            return {"message": f"Share link ({link_type}) created", "share_url": link}

        # ── Tier 6: advanced features ───────────────────────────────────
        if action == "calculate_workbook":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            calc_type = params.get("type") or params.get("calculation_type") or "Recalculate"
            r = await client.post(
                f"{base}/me/drive/items/{workbook_id}/workbook/application/calculate",
                headers=headers,
                json={"calculationType": calc_type},
            )
            if r.status_code >= 400:
                return {"error": f"Could not calculate: {r.text[:200]}"}
            return {"message": f"Recalculated workbook ({calc_type})"}

        if action == "list_tables":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            r = await client.get(
                f"{base}/me/drive/items/{workbook_id}/workbook/tables",
                headers=headers,
            )
            if r.status_code >= 400:
                return {"error": f"Could not list tables: {r.text[:200]}"}
            tables = (r.json() or {}).get("value", [])
            return {
                "tables": [
                    {"id": t.get("id"), "name": t.get("name"), "showHeaders": t.get("showHeaders"), "style": t.get("style")}
                    for t in tables
                ],
                "count": len(tables),
            }

        if action == "list_pivot_tables":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet")
            if worksheet:
                url = f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/pivotTables"
            else:
                url = f"{base}/me/drive/items/{workbook_id}/workbook/worksheets"
            # If no worksheet provided, iterate and gather pivots from each
            if not worksheet:
                wr = await client.get(url, headers=headers)
                if wr.status_code >= 400:
                    return {"error": f"Could not list worksheets: {wr.text[:200]}"}
                sheets = (wr.json() or {}).get("value", [])
                all_pivots: list[dict] = []
                for s in sheets:
                    sname = s.get("name")
                    pr = await client.get(
                        f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{sname}')/pivotTables",
                        headers=headers,
                    )
                    if pr.status_code == 200:
                        for p in (pr.json() or {}).get("value", []):
                            all_pivots.append({"worksheet": sname, "name": p.get("name"), "id": p.get("id")})
                return {"pivot_tables": all_pivots, "count": len(all_pivots)}
            r = await client.get(url, headers=headers)
            if r.status_code >= 400:
                return {"error": f"Could not list pivots: {r.text[:200]}"}
            pivots = (r.json() or {}).get("value", [])
            return {
                "pivot_tables": [{"name": p.get("name"), "id": p.get("id")} for p in pivots],
                "count": len(pivots),
            }

        if action == "refresh_pivot":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            pivot = params.get("pivot") or params.get("name")
            if pivot:
                r = await client.post(
                    f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/pivotTables('{pivot}')/refresh",
                    headers=headers,
                )
            else:
                r = await client.post(
                    f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/pivotTables/refreshAll",
                    headers=headers,
                )
            if r.status_code >= 400:
                return {"error": f"Could not refresh pivot: {r.text[:200]}"}
            return {"message": f"Refreshed pivot{'s on ' + worksheet if not pivot else ' ' + pivot}"}

        if action == "list_comments":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            r = await client.get(
                f"{base}/me/drive/items/{workbook_id}/workbook/comments",
                headers=headers,
            )
            if r.status_code >= 400:
                return {"error": f"Could not list comments: {r.text[:200]}"}
            comments = (r.json() or {}).get("value", [])
            return {
                "comments": [
                    {
                        "id": c.get("id"),
                        "content": (c.get("content") or {}).get("content"),
                        "cell": c.get("cellAddress"),
                    }
                    for c in comments
                ],
                "count": len(comments),
            }

        if action == "set_cell_comment":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            cell = params.get("cell") or params.get("address")
            text = params.get("text") or params.get("comment")
            if not cell or not text:
                return {"error": "cell and text are required"}
            r = await client.post(
                f"{base}/me/drive/items/{workbook_id}/workbook/comments",
                headers=headers,
                json={
                    "cellAddress": f"{worksheet}!{cell}",
                    "content": {"content": text, "contentType": "plain"},
                },
            )
            if r.status_code >= 400:
                return {"error": f"Could not set comment: {r.text[:200]}"}
            return {"message": f"Added comment on {worksheet}!{cell}"}

        if action == "insert_rows":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            try:
                row = int(params.get("row") or 1)
                count = int(params.get("count") or 1)
            except Exception:
                return {"error": "row and count must be integers"}
            end_row = row + count - 1
            r = await client.post(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/range(address='{row}:{end_row}')/insert",
                headers=headers,
                json={"shift": "Down"},
            )
            if r.status_code >= 400:
                return {"error": f"Could not insert rows: {r.text[:200]}"}
            return {"message": f"Inserted {count} row(s) starting at row {row}"}

        if action == "insert_columns":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            column = (params.get("column") or "A").strip().upper()
            try:
                count = int(params.get("count") or 1)
            except Exception:
                return {"error": "count must be an integer"}
            # Build the column range to insert — e.g. for A with count=3, that's A:C
            start_idx = 0
            for c in column:
                start_idx = start_idx * 26 + (ord(c) - 64)
            end_idx = start_idx + count - 1
            def _col_letter(n: int) -> str:
                s = ""
                while n > 0:
                    n, rem = divmod(n - 1, 26)
                    s = chr(65 + rem) + s
                return s
            end_col = _col_letter(end_idx)
            r = await client.post(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/range(address='{column}:{end_col}')/insert",
                headers=headers,
                json={"shift": "Right"},
            )
            if r.status_code >= 400:
                return {"error": f"Could not insert columns: {r.text[:200]}"}
            return {"message": f"Inserted {count} column(s) starting at {column}"}

        if action == "set_column_width":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            column = (params.get("column") or "A").strip().upper()
            try:
                width = float(params.get("width") or 64)
            except Exception:
                return {"error": "width must be a number"}
            # Accept either a single column "A" or a range "A:C"
            addr = column if ":" in column else f"{column}:{column}"
            r = await client.patch(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/range(address='{addr}')/format",
                headers=headers,
                json={"columnWidth": width},
            )
            if r.status_code >= 400:
                return {"error": f"Could not set column width: {r.text[:200]}"}
            return {"message": f"Set column(s) {addr} width to {width}"}

        if action == "set_row_height":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            row = str(params.get("row") or "1")
            try:
                height = float(params.get("height") or 18)
            except Exception:
                return {"error": "height must be a number"}
            addr = row if ":" in row else f"{row}:{row}"
            r = await client.patch(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/range(address='{addr}')/format",
                headers=headers,
                json={"rowHeight": height},
            )
            if r.status_code >= 400:
                return {"error": f"Could not set row height: {r.text[:200]}"}
            return {"message": f"Set row(s) {addr} height to {height}"}

        if action == "freeze_panes":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            try:
                rows = int(params.get("rows") or 0)
                cols = int(params.get("columns") or params.get("cols") or 0)
            except Exception:
                return {"error": "rows and columns must be integers"}
            if rows <= 0 and cols <= 0:
                return {"error": "At least one of rows or columns must be > 0"}
            if rows > 0 and cols <= 0:
                endpoint = "freezeRows"
                body = {"count": rows}
            elif cols > 0 and rows <= 0:
                endpoint = "freezeColumns"
                body = {"count": cols}
            else:
                # Freeze at a specific cell — col letter for cols, row num for rows
                def _cl(n: int) -> str:
                    s = ""
                    while n > 0:
                        n, rem = divmod(n - 1, 26)
                        s = chr(65 + rem) + s
                    return s
                target = f"{_cl(cols + 1)}{rows + 1}"
                r = await client.post(
                    f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/freezePanes/freezeAt",
                    headers=headers,
                    json={"frozenRange": target},
                )
                if r.status_code >= 400:
                    return {"error": f"Could not freeze panes: {r.text[:200]}"}
                return {"message": f"Froze panes at {target}"}
            r = await client.post(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/freezePanes/{endpoint}",
                headers=headers,
                json=body,
            )
            if r.status_code >= 400:
                return {"error": f"Could not freeze panes: {r.text[:200]}"}
            return {"message": f"Froze {rows or cols} {'row(s)' if rows else 'column(s)'}"}

        if action == "unfreeze_panes":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            r = await client.post(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/freezePanes/unfreeze",
                headers=headers,
            )
            if r.status_code >= 400:
                return {"error": f"Could not unfreeze: {r.text[:200]}"}
            return {"message": f"Unfroze panes on {worksheet}"}

        if action == "merge_cells":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            cell_range = params.get("range") or params.get("address")
            if not cell_range:
                return {"error": "range is required"}
            across = str(params.get("across", "false")).lower() == "true"
            r = await client.post(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/range(address='{cell_range}')/merge",
                headers=headers,
                json={"across": across},
            )
            if r.status_code >= 400:
                return {"error": f"Could not merge: {r.text[:200]}"}
            return {"message": f"Merged {worksheet}!{cell_range}"}

        if action == "unmerge_cells":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            cell_range = params.get("range") or params.get("address")
            if not cell_range:
                return {"error": "range is required"}
            r = await client.post(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/range(address='{cell_range}')/unmerge",
                headers=headers,
            )
            if r.status_code >= 400:
                return {"error": f"Could not unmerge: {r.text[:200]}"}
            return {"message": f"Unmerged {worksheet}!{cell_range}"}

        if action == "create_named_range":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            name = params.get("name")
            reference = params.get("reference") or params.get("range")
            if not name or not reference:
                return {"error": "name and reference are required"}
            r = await client.post(
                f"{base}/me/drive/items/{workbook_id}/workbook/names/add",
                headers=headers,
                json={
                    "name": name,
                    "reference": reference,
                    "comment": params.get("comment", ""),
                },
            )
            if r.status_code >= 400:
                return {"error": f"Could not create named range: {r.text[:200]}"}
            return {"message": f"Created named range '{name}' → {reference}"}

        if action == "list_named_ranges":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            r = await client.get(
                f"{base}/me/drive/items/{workbook_id}/workbook/names",
                headers=headers,
            )
            if r.status_code >= 400:
                return {"error": f"Could not list names: {r.text[:200]}"}
            names = (r.json() or {}).get("value", [])
            return {
                "named_ranges": [
                    {"name": n.get("name"), "value": n.get("value"), "type": n.get("type"), "comment": n.get("comment")}
                    for n in names
                ],
                "count": len(names),
            }

        if action == "add_hyperlink":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            cell = params.get("cell") or params.get("address")
            url = params.get("url") or params.get("link")
            if not cell or not url:
                return {"error": "cell and url are required"}
            display = params.get("display") or url
            # Hyperlinks in Graph are set via the HYPERLINK formula
            formula = f'=HYPERLINK("{url}","{display}")'
            r = await client.patch(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/range(address='{cell}')",
                headers=headers,
                json={"formulas": [[formula]]},
            )
            if r.status_code >= 400:
                return {"error": f"Could not add hyperlink: {r.text[:200]}"}
            return {"message": f"Added hyperlink in {worksheet}!{cell} → {url}"}

        if action == "unprotect_sheet":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            r = await client.post(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/protection/unprotect",
                headers=headers,
            )
            if r.status_code >= 400:
                return {"error": f"Could not unprotect: {r.text[:200]}"}
            return {"message": f"Unprotected worksheet '{worksheet}'"}

        if action == "range_details":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            cell_range = params.get("range") or params.get("address")
            if not cell_range:
                return {"error": "range is required"}
            r = await client.get(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/range(address='{cell_range}')",
                headers=headers,
            )
            if r.status_code >= 400:
                return {"error": f"Could not get range details: {r.text[:200]}"}
            data = r.json() or {}
            return {
                "address": data.get("address"),
                "values": data.get("values"),
                "text": data.get("text"),
                "formulas": data.get("formulas"),
                "numberFormat": data.get("numberFormat"),
                "rowCount": data.get("rowCount"),
                "columnCount": data.get("columnCount"),
                "cellCount": data.get("cellCount"),
                "valueTypes": data.get("valueTypes"),
            }

        if action == "set_conditional_format":
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            worksheet = params.get("worksheet") or params.get("sheet") or "Sheet1"
            cell_range = params.get("range") or params.get("address")
            if not cell_range:
                return {"error": "range is required"}
            rule = (params.get("rule") or "colorScale").strip()
            # Build a minimal body per rule type — Graph accepts richer configs too
            body: dict = {"type": rule[0].upper() + rule[1:]}  # normalize to PascalCase
            if rule.lower() == "colorscale":
                body["colorScale"] = {
                    "criteria": {
                        "minimum": {"type": "LowestValue", "color": params.get("min_color", "#FFFFFF")},
                        "maximum": {"type": "HighestValue", "color": params.get("max_color", "#FF0000")},
                    }
                }
            elif rule.lower() == "databar":
                body["dataBar"] = {"axisFormat": "Automatic", "barDirection": "LeftToRight"}
            elif rule.lower() == "iconset":
                body["iconSet"] = {"style": params.get("style", "ThreeTrafficLights1")}
            elif rule.lower() in ("top", "bottom"):
                body["topBottom"] = {
                    "rank": int(params.get("rank", 10)),
                    "type": "Items" if str(params.get("type", "items")).lower() == "items" else "Percent",
                    "isTopRanked": rule.lower() == "top",
                }
            elif rule.lower() == "aboveaverage":
                body["aboveAverage"] = {"isAboveAverage": True}
            r = await client.post(
                f"{base}/me/drive/items/{workbook_id}/workbook/worksheets('{worksheet}')/range(address='{cell_range}')/conditionalFormats/add",
                headers=headers,
                json=body,
            )
            if r.status_code >= 400:
                return {"error": f"Could not set conditional format: {r.text[:200]}"}
            return {"message": f"Added {rule} conditional format to {worksheet}!{cell_range}"}

        if action == "execute_function":
            # This is the meta-action — lets the AI call ANY Excel worksheet function
            # (400+ functions like VLOOKUP, XLOOKUP, SUMIF, TEXTJOIN, TRIM, etc.)
            ref = params.get("workbook_id") or params.get("workbookId") or params.get("filename")
            workbook_id, err = await _resolve_workbook_id(client, ref or "")
            if err:
                return {"error": err}
            func = params.get("function") or params.get("name")
            if not func:
                return {"error": "function name is required (e.g. VLOOKUP, SUMIF, XLOOKUP)"}
            args = params.get("args")
            if isinstance(args, str):
                try:
                    args = json.loads(args)
                except Exception:
                    # Fall back to comma-split for simple cases
                    args = [a.strip() for a in args.split(",")]
            if not isinstance(args, list):
                args = []
            # Graph expects each arg as a discrete field — map positionally
            body = {f"arg{i+1}": a for i, a in enumerate(args)} if len(args) > 1 else {"values": args[0] if args else ""}
            # Graph's function API uses function names in camelCase (e.g. vLookup). Convert:
            func_name = func[0].lower() + func[1:] if func else ""
            r = await client.post(
                f"{base}/me/drive/items/{workbook_id}/workbook/functions/{func_name}",
                headers=headers,
                json=body,
            )
            if r.status_code >= 400:
                return {
                    "error": f"Could not execute {func}: {r.text[:300]}",
                    "hint": "Make sure args match the function signature. Range refs should be full addresses like 'Sheet1!A1:B10'.",
                }
            data = r.json() or {}
            return {
                "function": func,
                "value": data.get("value"),
                "error_code": data.get("error"),
            }

        if action == "create_workbook":
            name = params.get("name") or "New Workbook"
            if not name.lower().endswith(".xlsx"):
                name = f"{name}.xlsx"
            # Create an empty .xlsx by uploading a minimal file via the Graph upload endpoint
            upload_headers = {
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }
            # Minimal valid .xlsx (empty workbook) — use openpyxl if available, otherwise use a pre-built blank
            try:
                from openpyxl import Workbook  # type: ignore
                from io import BytesIO
                wb = Workbook()
                buf = BytesIO()
                wb.save(buf)
                blob = buf.getvalue()
            except Exception:
                return {"error": "Cannot create empty workbook: openpyxl is not installed on the server."}
            r = await client.put(
                f"{base}/me/drive/root:/{name}:/content",
                headers=upload_headers,
                content=blob,
            )
            if r.status_code == 401:
                return {"error": "Microsoft Graph token is invalid or expired."}
            r.raise_for_status()
            data = r.json() or {}
            return {
                "message": f"Created workbook: {name}",
                "workbook": {
                    "id": data.get("id"),
                    "name": data.get("name"),
                    "url": data.get("webUrl"),
                },
            }

    return {"error": f"Unknown Excel Online action: {action}"}


async def _pipedrive_adapter(action: str, params: dict, creds: dict) -> dict:
    token = creds["api_key"]
    base = "https://api.pipedrive.com/v1"

    async with httpx.AsyncClient(timeout=30) as client:
        if action == "get_deals":
            r = await client.get(f"{base}/deals?api_token={token}&limit={params.get('limit', 20)}")
            r.raise_for_status()
            return {"deals": r.json().get("data", []) or [], "count": len(r.json().get("data", []) or [])}

        if action == "create_deal":
            data = {"title": params.get("name", "New Deal")}
            if params.get("amount"): data["value"] = params["amount"]
            r = await client.post(f"{base}/deals?api_token={token}", json=data)
            r.raise_for_status()
            return {"deal": r.json().get("data"), "message": "Deal created"}

        if action == "get_persons":
            r = await client.get(f"{base}/persons?api_token={token}&limit={params.get('limit', 20)}")
            r.raise_for_status()
            return {"persons": r.json().get("data", []) or [], "count": len(r.json().get("data", []) or [])}

        if action == "create_person":
            data = {"name": params.get("name", "Unknown")}
            if params.get("email"): data["email"] = [{"value": params["email"]}]
            if params.get("phone"): data["phone"] = [{"value": params["phone"]}]
            r = await client.post(f"{base}/persons?api_token={token}", json=data)
            r.raise_for_status()
            return {"person": r.json().get("data"), "message": "Contact created"}

        if action == "search":
            query = params.get("query", "")
            r = await client.get(f"{base}/itemSearch?api_token={token}&term={query}")
            r.raise_for_status()
            return {"results": r.json().get("data", {}).get("items", [])}

    return {"error": f"Unknown action: {action}"}


async def _asana_adapter(action: str, params: dict, creds: dict) -> dict:
    token = creds["api_key"]
    headers = {"Authorization": f"Bearer {token}"}
    base = "https://app.asana.com/api/1.0"

    async with httpx.AsyncClient(timeout=30) as client:
        if action == "get_tasks":
            project = params.get("project")
            if project:
                r = await client.get(f"{base}/projects/{project}/tasks?opt_fields=name,completed,due_on,assignee.name", headers=headers)
            else:
                r = await client.get(f"{base}/tasks?assignee=me&workspace={params.get('workspace', '')}&opt_fields=name,completed,due_on", headers=headers)
            r.raise_for_status()
            return {"tasks": r.json().get("data", []), "count": len(r.json().get("data", []))}

        if action == "create_task":
            data = {"data": {"name": params.get("name", "New Task")}}
            if params.get("project"): data["data"]["projects"] = [params["project"]]
            if params.get("due_on"): data["data"]["due_on"] = params["due_on"]
            if params.get("notes"): data["data"]["notes"] = params["notes"]
            r = await client.post(f"{base}/tasks", headers=headers, json=data)
            r.raise_for_status()
            return {"task": r.json().get("data"), "message": "Task created"}

        if action == "get_projects":
            r = await client.get(f"{base}/projects?opt_fields=name,color,archived", headers=headers)
            r.raise_for_status()
            return {"projects": r.json().get("data", [])}

        if action == "update_task":
            task_id = params.get("task_id", "")
            data = {"data": {}}
            if params.get("completed") is not None: data["data"]["completed"] = params["completed"]
            if params.get("name"): data["data"]["name"] = params["name"]
            r = await client.put(f"{base}/tasks/{task_id}", headers=headers, json=data)
            r.raise_for_status()
            return {"task": r.json().get("data"), "message": "Task updated"}

        if action == "search":
            r = await client.get(f"{base}/workspaces", headers=headers)
            r.raise_for_status()
            workspaces = r.json().get("data", [])
            if workspaces:
                ws = workspaces[0]["gid"]
                sr = await client.get(f"{base}/workspaces/{ws}/tasks/search?text={params.get('query', '')}", headers=headers)
                sr.raise_for_status()
                return {"results": sr.json().get("data", [])}
            return {"results": []}

    return {"error": f"Unknown action: {action}"}


async def _trello_adapter(action: str, params: dict, creds: dict) -> dict:
    key = creds["api_key"]
    token = creds["token"]
    base = "https://api.trello.com/1"
    auth = f"key={key}&token={token}"

    async with httpx.AsyncClient(timeout=30) as client:
        if action == "get_boards":
            r = await client.get(f"{base}/members/me/boards?{auth}&fields=name,url,closed")
            r.raise_for_status()
            return {"boards": r.json()}

        if action == "get_cards":
            board = params.get("board_id", "")
            r = await client.get(f"{base}/boards/{board}/cards?{auth}&fields=name,desc,due,idList,url")
            r.raise_for_status()
            return {"cards": r.json(), "count": len(r.json())}

        if action == "create_card":
            data = {"name": params.get("name", "New Card"), "idList": params.get("list_id", "")}
            if params.get("desc"): data["desc"] = params["desc"]
            if params.get("due"): data["due"] = params["due"]
            r = await client.post(f"{base}/cards?{auth}", json=data)
            r.raise_for_status()
            return {"card": r.json(), "message": "Card created"}

        if action == "move_card":
            card_id = params.get("card_id", "")
            r = await client.put(f"{base}/cards/{card_id}?{auth}", json={"idList": params.get("list_id", "")})
            r.raise_for_status()
            return {"card": r.json(), "message": "Card moved"}

        if action == "search":
            r = await client.get(f"{base}/search?{auth}&query={params.get('query', '')}&modelTypes=cards,boards")
            r.raise_for_status()
            return {"results": r.json()}

    return {"error": f"Unknown action: {action}"}


async def _notion_adapter(action: str, params: dict, creds: dict) -> dict:
    token = creds["api_key"]
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json", "Notion-Version": "2022-06-28"}
    base = "https://api.notion.com/v1"

    async with httpx.AsyncClient(timeout=30) as client:
        if action == "search":
            r = await client.post(f"{base}/search", headers=headers, json={"query": params.get("query", ""), "page_size": 20})
            r.raise_for_status()
            results = [{"id": p["id"], "title": p.get("properties", {}).get("title", {}).get("title", [{}])[0].get("plain_text", "") if p["object"] == "page" else p.get("title", [{}])[0].get("plain_text", ""), "type": p["object"]} for p in r.json().get("results", [])]
            return {"results": results, "count": len(results)}

        if action == "get_pages":
            r = await client.post(f"{base}/search", headers=headers, json={"filter": {"property": "object", "value": "page"}, "page_size": 20})
            r.raise_for_status()
            return {"pages": r.json().get("results", [])}

        if action == "create_page":
            page_data: dict[str, Any] = {"parent": {"page_id": params.get("parent_id", "")}, "properties": {"title": {"title": [{"text": {"content": params.get("title", "New Page")}}]}}}
            if params.get("content"):
                page_data["children"] = [{"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": params["content"]}}]}}]
            r = await client.post(f"{base}/pages", headers=headers, json=page_data)
            r.raise_for_status()
            return {"page": r.json(), "message": "Page created"}

        if action == "update_page":
            page_id = params.get("page_id", "")
            props = {}
            if params.get("title"):
                props["title"] = {"title": [{"text": {"content": params["title"]}}]}
            r = await client.patch(f"{base}/pages/{page_id}", headers=headers, json={"properties": props})
            r.raise_for_status()
            return {"page": r.json(), "message": "Page updated"}

        if action == "query_database":
            db_id = params.get("database_id", "")
            r = await client.post(f"{base}/databases/{db_id}/query", headers=headers, json={"page_size": params.get("limit", 20)})
            r.raise_for_status()
            return {"results": r.json().get("results", []), "count": len(r.json().get("results", []))}

    return {"error": f"Unknown action: {action}"}


async def _slack_adapter(action: str, params: dict, creds: dict) -> dict:
    token = creds["api_key"]
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    base = "https://slack.com/api"

    async with httpx.AsyncClient(timeout=30) as client:
        if action == "get_channels":
            r = await client.get(f"{base}/conversations.list?limit=50", headers=headers)
            r.raise_for_status()
            channels = [{"id": c["id"], "name": c["name"]} for c in r.json().get("channels", [])]
            return {"channels": channels}

        if action == "send_message":
            channel = params.get("channel", "")
            text = params.get("text", "")
            # If channel is a name, find its ID
            if not channel.startswith("C") and not channel.startswith("D"):
                ch_r = await client.get(f"{base}/conversations.list?limit=200", headers=headers)
                ch_r.raise_for_status()
                for c in ch_r.json().get("channels", []):
                    if c["name"] == channel.lstrip("#"):
                        channel = c["id"]
                        break
            r = await client.post(f"{base}/chat.postMessage", headers=headers, json={"channel": channel, "text": text})
            r.raise_for_status()
            return {"message": "Message sent", "channel": channel}

        if action == "get_messages":
            channel = params.get("channel", "")
            r = await client.get(f"{base}/conversations.history?channel={channel}&limit={params.get('limit', 20)}", headers=headers)
            r.raise_for_status()
            return {"messages": r.json().get("messages", [])}

        if action == "search":
            r = await client.get(f"{base}/search.messages?query={params.get('query', '')}&count=20", headers=headers)
            r.raise_for_status()
            return {"results": r.json().get("messages", {}).get("matches", [])}

    return {"error": f"Unknown action: {action}"}


async def _stripe_adapter(action: str, params: dict, creds: dict) -> dict:
    token = creds["api_key"]
    headers = {"Authorization": f"Bearer {token}"}
    base = "https://api.stripe.com/v1"

    async with httpx.AsyncClient(timeout=30) as client:
        if action == "get_payments":
            r = await client.get(f"{base}/charges?limit={params.get('limit', 20)}", headers=headers)
            r.raise_for_status()
            charges = [{"id": c["id"], "amount": c["amount"] / 100, "currency": c["currency"], "status": c["status"], "description": c.get("description")} for c in r.json().get("data", [])]
            return {"payments": charges, "count": len(charges)}

        if action == "get_customers":
            r = await client.get(f"{base}/customers?limit={params.get('limit', 20)}", headers=headers)
            r.raise_for_status()
            customers = [{"id": c["id"], "name": c.get("name"), "email": c.get("email")} for c in r.json().get("data", [])]
            return {"customers": customers, "count": len(customers)}

        if action == "create_payment_link":
            data = {"line_items[0][price_data][currency]": params.get("currency", "usd"), "line_items[0][price_data][product_data][name]": params.get("name", "Payment"), "line_items[0][price_data][unit_amount]": int(float(params.get("amount", 0)) * 100), "line_items[0][quantity]": "1"}
            r = await client.post(f"{base}/payment_links", headers=headers, data=data)
            r.raise_for_status()
            return {"payment_link": r.json().get("url"), "id": r.json().get("id")}

        if action == "get_invoices":
            r = await client.get(f"{base}/invoices?limit={params.get('limit', 20)}", headers=headers)
            r.raise_for_status()
            invoices = [{"id": i["id"], "amount_due": i["amount_due"] / 100, "status": i["status"], "customer": i.get("customer_name")} for i in r.json().get("data", [])]
            return {"invoices": invoices, "count": len(invoices)}

        if action == "get_balance":
            r = await client.get(f"{base}/balance", headers=headers)
            r.raise_for_status()
            available = sum(b["amount"] for b in r.json().get("available", []))
            pending = sum(b["amount"] for b in r.json().get("pending", []))
            return {"available": available / 100, "pending": pending / 100, "currency": r.json().get("available", [{}])[0].get("currency", "usd")}

    return {"error": f"Unknown action: {action}"}


async def _shopify_adapter(action: str, params: dict, creds: dict) -> dict:
    token = creds["api_key"]
    store = creds.get("store_url", "").rstrip("/")
    if not store.startswith("https://"): store = f"https://{store}"
    headers = {"X-Shopify-Access-Token": token, "Content-Type": "application/json"}
    base = f"{store}/admin/api/2024-01"

    async with httpx.AsyncClient(timeout=30) as client:
        if action == "get_orders":
            r = await client.get(f"{base}/orders.json?status=any&limit={params.get('limit', 20)}", headers=headers)
            r.raise_for_status()
            orders = [{"id": o["id"], "name": o["name"], "total": o["total_price"], "status": o["financial_status"], "customer": o.get("customer", {}).get("first_name", "")} for o in r.json().get("orders", [])]
            return {"orders": orders, "count": len(orders)}

        if action == "get_products":
            r = await client.get(f"{base}/products.json?limit={params.get('limit', 20)}", headers=headers)
            r.raise_for_status()
            products = [{"id": p["id"], "title": p["title"], "status": p["status"], "price": p.get("variants", [{}])[0].get("price")} for p in r.json().get("products", [])]
            return {"products": products, "count": len(products)}

        if action == "get_customers":
            r = await client.get(f"{base}/customers.json?limit={params.get('limit', 20)}", headers=headers)
            r.raise_for_status()
            customers = [{"id": c["id"], "name": f"{c.get('first_name', '')} {c.get('last_name', '')}", "email": c.get("email"), "orders_count": c.get("orders_count")} for c in r.json().get("customers", [])]
            return {"customers": customers, "count": len(customers)}

        if action == "create_product":
            data = {"product": {"title": params.get("title", "New Product"), "body_html": params.get("description", ""), "variants": [{"price": params.get("price", "0")}]}}
            r = await client.post(f"{base}/products.json", headers=headers, json=data)
            r.raise_for_status()
            return {"product": r.json().get("product"), "message": "Product created"}

        if action == "get_inventory":
            r = await client.get(f"{base}/products.json?limit=50", headers=headers)
            r.raise_for_status()
            inventory = [{"title": p["title"], "variants": [{"sku": v.get("sku"), "inventory": v.get("inventory_quantity")} for v in p.get("variants", [])]} for p in r.json().get("products", [])]
            return {"inventory": inventory}

    return {"error": f"Unknown action: {action}"}


async def _clickup_adapter(action: str, params: dict, creds: dict) -> dict:
    token = creds["api_key"]
    headers = {"Authorization": token, "Content-Type": "application/json"}
    base = "https://api.clickup.com/api/v2"

    async with httpx.AsyncClient(timeout=30) as client:
        if action == "get_tasks":
            list_id = params.get("list_id", "")
            r = await client.get(f"{base}/list/{list_id}/task", headers=headers)
            r.raise_for_status()
            return {"tasks": r.json().get("tasks", [])}

        if action == "create_task":
            list_id = params.get("list_id", "")
            data = {"name": params.get("name", "New Task")}
            if params.get("description"): data["description"] = params["description"]
            if params.get("due_date"): data["due_date"] = params["due_date"]
            r = await client.post(f"{base}/list/{list_id}/task", headers=headers, json=data)
            r.raise_for_status()
            return {"task": r.json(), "message": "Task created"}

        if action == "get_spaces":
            # First get teams
            tr = await client.get(f"{base}/team", headers=headers)
            tr.raise_for_status()
            teams = tr.json().get("teams", [])
            if teams:
                sr = await client.get(f"{base}/team/{teams[0]['id']}/space", headers=headers)
                sr.raise_for_status()
                return {"spaces": sr.json().get("spaces", [])}
            return {"spaces": []}

        if action == "update_task":
            task_id = params.get("task_id", "")
            data = {}
            if params.get("name"): data["name"] = params["name"]
            if params.get("status"): data["status"] = params["status"]
            r = await client.put(f"{base}/task/{task_id}", headers=headers, json=data)
            r.raise_for_status()
            return {"task": r.json(), "message": "Task updated"}

        if action == "search":
            tr = await client.get(f"{base}/team", headers=headers)
            tr.raise_for_status()
            teams = tr.json().get("teams", [])
            if teams:
                r = await client.get(f"{base}/team/{teams[0]['id']}/task?name={params.get('query', '')}", headers=headers)
                r.raise_for_status()
                return {"results": r.json().get("tasks", [])}
            return {"results": []}

    return {"error": f"Unknown action: {action}"}


async def _todoist_adapter(action: str, params: dict, creds: dict) -> dict:
    token = creds["api_key"]
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    base = "https://api.todoist.com/rest/v2"

    async with httpx.AsyncClient(timeout=30) as client:
        if action == "get_tasks":
            r = await client.get(f"{base}/tasks", headers=headers)
            r.raise_for_status()
            tasks = [{"id": t["id"], "content": t["content"], "due": t.get("due", {}).get("string") if t.get("due") else None, "completed": t.get("is_completed", False)} for t in r.json()]
            return {"tasks": tasks, "count": len(tasks)}

        if action == "create_task":
            data = {"content": params.get("name", "New Task")}
            if params.get("due"): data["due_string"] = params["due"]
            if params.get("project_id"): data["project_id"] = params["project_id"]
            r = await client.post(f"{base}/tasks", headers=headers, json=data)
            r.raise_for_status()
            return {"task": r.json(), "message": "Task created"}

        if action == "update_task":
            task_id = params.get("task_id", "")
            data = {}
            if params.get("content"): data["content"] = params["content"]
            if params.get("due"): data["due_string"] = params["due"]
            r = await client.post(f"{base}/tasks/{task_id}", headers=headers, json=data)
            r.raise_for_status()
            return {"task": r.json(), "message": "Task updated"}

        if action == "get_projects":
            r = await client.get(f"{base}/projects", headers=headers)
            r.raise_for_status()
            return {"projects": r.json()}

    return {"error": f"Unknown action: {action}"}


async def _jira_adapter(action: str, params: dict, creds: dict) -> dict:
    import base64
    email = creds.get("email", "")
    token = creds["api_key"]
    domain = creds.get("domain", "").rstrip("/")
    if not domain.startswith("https://"): domain = f"https://{domain}"
    auth = base64.b64encode(f"{email}:{token}".encode()).decode()
    headers = {"Authorization": f"Basic {auth}", "Content-Type": "application/json"}
    base = f"{domain}/rest/api/3"

    async with httpx.AsyncClient(timeout=30) as client:
        if action == "get_issues":
            jql = params.get("jql", "assignee=currentUser() ORDER BY updated DESC")
            r = await client.get(f"{base}/search?jql={jql}&maxResults={params.get('limit', 20)}", headers=headers)
            r.raise_for_status()
            issues = [{"key": i["key"], "summary": i["fields"]["summary"], "status": i["fields"]["status"]["name"], "type": i["fields"]["issuetype"]["name"]} for i in r.json().get("issues", [])]
            return {"issues": issues, "count": r.json().get("total", 0)}

        if action == "create_issue":
            data = {"fields": {"project": {"key": params.get("project", "")}, "summary": params.get("summary", "New Issue"), "issuetype": {"name": params.get("type", "Task")}}}
            if params.get("description"): data["fields"]["description"] = {"type": "doc", "version": 1, "content": [{"type": "paragraph", "content": [{"type": "text", "text": params["description"]}]}]}
            r = await client.post(f"{base}/issue", headers=headers, json=data)
            r.raise_for_status()
            return {"issue": r.json(), "message": f"Issue {r.json().get('key')} created"}

        if action == "update_issue":
            issue_key = params.get("issue_key", "")
            data = {"fields": {}}
            if params.get("summary"): data["fields"]["summary"] = params["summary"]
            if params.get("status"):
                # Get transitions
                tr = await client.get(f"{base}/issue/{issue_key}/transitions", headers=headers)
                tr.raise_for_status()
                for t in tr.json().get("transitions", []):
                    if t["name"].lower() == params["status"].lower():
                        await client.post(f"{base}/issue/{issue_key}/transitions", headers=headers, json={"transition": {"id": t["id"]}})
                        return {"message": f"Issue {issue_key} moved to {params['status']}"}
            r = await client.put(f"{base}/issue/{issue_key}", headers=headers, json=data)
            r.raise_for_status()
            return {"message": f"Issue {issue_key} updated"}

        if action == "search":
            jql = f'text ~ "{params.get("query", "")}"'
            r = await client.get(f"{base}/search?jql={jql}&maxResults=20", headers=headers)
            r.raise_for_status()
            return {"results": [{"key": i["key"], "summary": i["fields"]["summary"]} for i in r.json().get("issues", [])]}

        if action == "get_projects":
            r = await client.get(f"{base}/project", headers=headers)
            r.raise_for_status()
            return {"projects": [{"key": p["key"], "name": p["name"]} for p in r.json()]}

    return {"error": f"Unknown action: {action}"}


async def _linear_adapter(action: str, params: dict, creds: dict) -> dict:
    token = creds["api_key"]
    headers = {"Authorization": token, "Content-Type": "application/json"}
    base = "https://api.linear.app/graphql"

    async with httpx.AsyncClient(timeout=30) as client:
        if action == "get_issues":
            query = '{ issues(first: 20, orderBy: updatedAt) { nodes { id identifier title state { name } priority assignee { name } } } }'
            r = await client.post(base, headers=headers, json={"query": query})
            r.raise_for_status()
            issues = r.json().get("data", {}).get("issues", {}).get("nodes", [])
            return {"issues": [{"id": i["id"], "key": i["identifier"], "title": i["title"], "status": i.get("state", {}).get("name")} for i in issues]}

        if action == "create_issue":
            query = 'mutation($input: IssueCreateInput!) { issueCreate(input: $input) { issue { id identifier title } } }'
            variables = {"input": {"title": params.get("title", "New Issue"), "teamId": params.get("team_id", "")}}
            if params.get("description"): variables["input"]["description"] = params["description"]
            r = await client.post(base, headers=headers, json={"query": query, "variables": variables})
            r.raise_for_status()
            issue = r.json().get("data", {}).get("issueCreate", {}).get("issue", {})
            return {"issue": issue, "message": f"Issue {issue.get('identifier')} created"}

        if action == "update_issue":
            issue_id = params.get("issue_id", "")
            query = 'mutation($id: String!, $input: IssueUpdateInput!) { issueUpdate(id: $id, input: $input) { issue { id identifier title state { name } } } }'
            inp: dict[str, Any] = {}
            if params.get("title"): inp["title"] = params["title"]
            if params.get("state_id"): inp["stateId"] = params["state_id"]
            r = await client.post(base, headers=headers, json={"query": query, "variables": {"id": issue_id, "input": inp}})
            r.raise_for_status()
            return {"issue": r.json().get("data", {}).get("issueUpdate", {}).get("issue"), "message": "Issue updated"}

        if action == "search":
            query_text = params.get("query", "")
            gql = f'{{ issueSearch(query: "{query_text}", first: 20) {{ nodes {{ id identifier title state {{ name }} }} }} }}'
            r = await client.post(base, headers=headers, json={"query": gql})
            r.raise_for_status()
            return {"results": r.json().get("data", {}).get("issueSearch", {}).get("nodes", [])}

    return {"error": f"Unknown action: {action}"}


async def _webhook_adapter(action: str, params: dict, creds: dict) -> dict:
    """Generic webhook adapter for Zapier, Make, n8n, IFTTT."""
    url = creds.get("webhook_url", "")
    if not url:
        return {"error": "No webhook URL configured"}

    async with httpx.AsyncClient(timeout=30) as client:
        if action == "trigger_webhook":
            payload = params.get("data", params)
            # For IFTTT, append event name
            if creds.get("event_name"):
                url = url.replace("{event}", creds["event_name"])
            r = await client.post(url, json=payload)
            r.raise_for_status()
            return {"message": "Webhook triggered successfully", "status_code": r.status_code}

    return {"error": f"Unknown action: {action}"}


async def _telegram_adapter(action: str, params: dict, creds: dict) -> dict:
    token = creds["api_key"]
    base = f"https://api.telegram.org/bot{token}"

    async with httpx.AsyncClient(timeout=30) as client:
        if action == "send_message":
            chat_id = params.get("chat_id", "")
            text = params.get("text", "")
            r = await client.post(f"{base}/sendMessage", json={"chat_id": chat_id, "text": text})
            r.raise_for_status()
            return {"message": "Message sent", "result": r.json().get("result")}

        if action == "get_updates":
            r = await client.get(f"{base}/getUpdates?limit=20")
            r.raise_for_status()
            return {"updates": r.json().get("result", [])}

    return {"error": f"Unknown action: {action}"}


async def _airtable_adapter(action: str, params: dict, creds: dict) -> dict:
    token = creds["api_key"]
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    base_id = params.get("base_id", "")
    table = params.get("table", "")
    base = f"https://api.airtable.com/v0/{base_id}/{table}"

    async with httpx.AsyncClient(timeout=30) as client:
        if action == "get_records":
            r = await client.get(f"{base}?maxRecords={params.get('limit', 20)}", headers=headers)
            r.raise_for_status()
            return {"records": r.json().get("records", []), "count": len(r.json().get("records", []))}

        if action == "create_record":
            fields = params.get("fields", {})
            r = await client.post(base, headers=headers, json={"fields": fields})
            r.raise_for_status()
            return {"record": r.json(), "message": "Record created"}

        if action == "update_record":
            record_id = params.get("record_id", "")
            fields = params.get("fields", {})
            r = await client.patch(f"{base}/{record_id}", headers=headers, json={"fields": fields})
            r.raise_for_status()
            return {"record": r.json(), "message": "Record updated"}

        if action == "search":
            formula = f'SEARCH("{params.get("query", "")}", ARRAYJOIN({{Name}}))'
            r = await client.get(f"{base}?filterByFormula={formula}&maxRecords=20", headers=headers)
            r.raise_for_status()
            return {"results": r.json().get("records", [])}

    return {"error": f"Unknown action: {action}"}


# Generic adapter for apps that just need basic REST calls
async def _generic_rest_adapter(app_id: str, action: str, params: dict, creds: dict) -> dict:
    """Placeholder adapter — returns a helpful message about the app not being fully implemented."""
    app_name = APP_REGISTRY.get(app_id, {}).get("name", app_id)
    return {
        "message": f"{app_name} is connected but this action ({action}) is coming soon. For now, you can use Zapier webhooks to trigger {app_name} actions.",
        "connected": True,
        "action": action,
    }


# ── Mail adapters ────────────────────────────────────────────────────────
#
# All three mail adapters (Outlook via Graph, Gmail via Gmail API, and the
# generic IMAP catch-all) expose the SAME action names so the LLM can use
# "list_inbox", "search_emails", "read_email", etc. without caring which
# provider is on the other side. The only thing that varies is the action
# hints in the registry (Gmail accepts its native search grammar, Outlook
# uses Graph's $search syntax, IMAP uses simple filters).


def _parse_mail_list(value: Any) -> list[str]:
    """Accept either a comma-separated string or a JSON list of recipients."""
    if not value:
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    return [s.strip() for s in str(value).split(",") if s.strip()]


# ── Autodetect IMAP/SMTP servers from an email address ─────────────────
#
# Users shouldn't have to google "what's Yahoo's IMAP port". Given an email
# address, we try to infer the incoming (IMAP) and outgoing (SMTP) servers
# using the same sources Thunderbird and Apple Mail use:
#
#   1. Mozilla ISPDB  (https://autoconfig.thunderbird.net/v1.1/<domain>)
#      — Thunderbird's public autoconfig database, covers thousands of
#        providers including Gmail, Yahoo, iCloud, Neo, Titan, GoDaddy,
#        FastMail, Zoho, ProtonMail, and most business hosts.
#   2. Provider-hosted autoconfig XML (autoconfig.<domain>/mail/config-v1.1.xml
#      and the .well-known variant) — Mozilla spec, many domains self-host it.
#   3. Common heuristics — imap.<domain>:993, smtp.<domain>:587, then
#      mail.<domain>:993/587 as last resort.
#
# Cached in-process for the lifetime of the worker so repeated connects
# don't re-hit the network.

_MAIL_AUTODETECT_CACHE: dict[str, dict] = {}

# Known preset hosts for providers that don't publish autoconfig. Cheap
# short-circuit before we touch the network.
_MAIL_AUTODETECT_PRESETS: dict[str, dict] = {
    "gmail.com":     {"imap_host": "imap.gmail.com",    "imap_port": 993, "smtp_host": "smtp.gmail.com",    "smtp_port": 587},
    "googlemail.com":{"imap_host": "imap.gmail.com",    "imap_port": 993, "smtp_host": "smtp.gmail.com",    "smtp_port": 587},
    "yahoo.com":     {"imap_host": "imap.mail.yahoo.com","imap_port": 993, "smtp_host": "smtp.mail.yahoo.com","smtp_port": 587},
    "ymail.com":     {"imap_host": "imap.mail.yahoo.com","imap_port": 993, "smtp_host": "smtp.mail.yahoo.com","smtp_port": 587},
    "icloud.com":    {"imap_host": "imap.mail.me.com",  "imap_port": 993, "smtp_host": "smtp.mail.me.com",  "smtp_port": 587},
    "me.com":        {"imap_host": "imap.mail.me.com",  "imap_port": 993, "smtp_host": "smtp.mail.me.com",  "smtp_port": 587},
    "mac.com":       {"imap_host": "imap.mail.me.com",  "imap_port": 993, "smtp_host": "smtp.mail.me.com",  "smtp_port": 587},
    "outlook.com":   {"imap_host": "outlook.office365.com","imap_port": 993, "smtp_host": "smtp.office365.com","smtp_port": 587},
    "hotmail.com":   {"imap_host": "outlook.office365.com","imap_port": 993, "smtp_host": "smtp.office365.com","smtp_port": 587},
    "live.com":      {"imap_host": "outlook.office365.com","imap_port": 993, "smtp_host": "smtp.office365.com","smtp_port": 587},
    "fastmail.com":  {"imap_host": "imap.fastmail.com", "imap_port": 993, "smtp_host": "smtp.fastmail.com", "smtp_port": 587},
    "zoho.com":      {"imap_host": "imap.zoho.com",     "imap_port": 993, "smtp_host": "smtp.zoho.com",     "smtp_port": 587},
    "protonmail.com":{"imap_host": "127.0.0.1",         "imap_port": 1143, "smtp_host": "127.0.0.1",        "smtp_port": 1025},  # Bridge
    "neo.space":     {"imap_host": "imap.neo.space",    "imap_port": 993, "smtp_host": "smtp.neo.space",    "smtp_port": 465},
    "titan.email":   {"imap_host": "imap.titan.email",  "imap_port": 993, "smtp_host": "smtp.titan.email",  "smtp_port": 587},
}


def _parse_autoconfig_xml(xml_text: str) -> dict | None:
    """Parse a Mozilla autoconfig XML response and return the first
    IMAP + SMTP pair found. Returns None if the XML is malformed or has
    no matching servers."""
    try:
        import xml.etree.ElementTree as ET
        root = ET.fromstring(xml_text)
    except Exception:
        return None

    incoming_host = incoming_port = outgoing_host = outgoing_port = None

    # Autoconfig uses either emailProvider/incomingServer[@type="imap"] or
    # clientConfig/emailProvider/incomingServer — just walk and match.
    for server in root.iter("incomingServer"):
        if (server.get("type") or "").lower() == "imap":
            h = server.findtext("hostname")
            p = server.findtext("port")
            if h and not incoming_host:
                incoming_host = h
                try: incoming_port = int(p) if p else 993
                except Exception: incoming_port = 993
            break
    for server in root.iter("outgoingServer"):
        if (server.get("type") or "").lower() == "smtp":
            h = server.findtext("hostname")
            p = server.findtext("port")
            if h and not outgoing_host:
                outgoing_host = h
                try: outgoing_port = int(p) if p else 587
                except Exception: outgoing_port = 587
            break

    if incoming_host and outgoing_host:
        return {
            "imap_host": incoming_host,
            "imap_port": incoming_port or 993,
            "smtp_host": outgoing_host,
            "smtp_port": outgoing_port or 587,
        }
    return None


async def _autodetect_mail_servers(email_address: str) -> dict | None:
    """Look up IMAP + SMTP settings for an email address.

    Returns {imap_host, imap_port, smtp_host, smtp_port} on success or
    None if nothing worked (caller falls back to whatever heuristic it
    prefers). Results are cached per-domain.
    """
    if not email_address or "@" not in email_address:
        return None
    domain = email_address.rsplit("@", 1)[-1].strip().lower()
    if not domain:
        return None

    if domain in _MAIL_AUTODETECT_CACHE:
        cached = _MAIL_AUTODETECT_CACHE[domain]
        return cached  # may be a dict or None (negative cache)
    if domain in _MAIL_AUTODETECT_PRESETS:
        _MAIL_AUTODETECT_CACHE[domain] = _MAIL_AUTODETECT_PRESETS[domain]
        return _MAIL_AUTODETECT_PRESETS[domain]

    async with httpx.AsyncClient(timeout=8, follow_redirects=True) as client:
        # 1. Mozilla ISPDB — the most complete source by far
        urls_to_try = [
            f"https://autoconfig.thunderbird.net/v1.1/{domain}",
            f"https://autoconfig.{domain}/mail/config-v1.1.xml?emailaddress={email_address}",
            f"https://{domain}/.well-known/autoconfig/mail/config-v1.1.xml?emailaddress={email_address}",
        ]
        for url in urls_to_try:
            try:
                r = await client.get(url)
                if r.status_code == 200 and r.text.strip().startswith("<"):
                    parsed = _parse_autoconfig_xml(r.text)
                    if parsed:
                        _MAIL_AUTODETECT_CACHE[domain] = parsed
                        return parsed
            except Exception:
                continue

    # 2. Heuristic fallback — guess from the domain itself. A lot of
    #    custom domains follow the imap.<domain> convention, but plenty
    #    don't, and returning a host that doesn't exist just makes the
    #    connection fail with "Name or service not known". Validate the
    #    guess via DNS before returning it.
    import socket
    async def _resolves(host: str) -> bool:
        try:
            loop = asyncio.get_running_loop()
            await loop.getaddrinfo(host, None)
            return True
        except Exception:
            return False

    guesses_to_try = [
        ("imap", "smtp", 993, 587),
        ("mail", "mail", 993, 587),  # some hosts use mail.<domain> for both
    ]
    for imap_prefix, smtp_prefix, ip, sp in guesses_to_try:
        ih = f"{imap_prefix}.{domain}"
        sh = f"{smtp_prefix}.{domain}"
        if await _resolves(ih) and await _resolves(sh):
            guessed = {
                "imap_host": ih,
                "imap_port": ip,
                "smtp_host": sh,
                "smtp_port": sp,
                "_guessed": True,
            }
            _MAIL_AUTODETECT_CACHE[domain] = guessed
            return guessed

    # Nothing resolved — caller will surface a clean "couldn't autodetect"
    # error asking the user to enter the host manually in Settings.
    _MAIL_AUTODETECT_CACHE[domain] = None  # negative cache
    return None


async def _outlook_mail_adapter(action: str, params: dict, creds: dict) -> dict:
    """Microsoft Outlook Mail adapter via Microsoft Graph.

    Reuses the same OAuth infrastructure as Excel Online. The access token
    must be scoped for Mail.ReadWrite and Mail.Send.
    """
    token = creds.get("access_token") or creds.get("api_key")
    if not token:
        return {"error": "Outlook is not connected. Connect it in Settings → My Apps → Microsoft Outlook."}

    base = "https://graph.microsoft.com/v1.0"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    def _folder_path(folder: str | None) -> str:
        """Resolve a folder name to the Graph URL segment. "Inbox" is a
        well-known folder; anything else we look up by display name."""
        if not folder or folder.lower() == "inbox":
            return "mailFolders/inbox"
        return f"mailFolders('{folder}')"

    async with httpx.AsyncClient(timeout=30) as client:
        if action == "list_inbox":
            folder = params.get("folder") or "Inbox"
            limit = min(int(params.get("limit") or 20), 50)
            # Resolve the folder name to an id when it's not "inbox"
            if folder.lower() == "inbox":
                url = f"{base}/me/mailFolders/inbox/messages"
            else:
                fr = await client.get(f"{base}/me/mailFolders", headers=headers, params={"$top": 100})
                if fr.status_code >= 400:
                    return {"error": f"Could not list folders: {fr.status_code}"}
                match = next(
                    (f for f in (fr.json() or {}).get("value", []) if (f.get("displayName") or "").lower() == folder.lower()),
                    None,
                )
                if not match:
                    return {"error": f"Folder '{folder}' not found"}
                url = f"{base}/me/mailFolders/{match['id']}/messages"
            r = await client.get(
                url,
                headers=headers,
                params={
                    "$top": limit,
                    "$select": "id,subject,from,receivedDateTime,bodyPreview,isRead,hasAttachments",
                    "$orderby": "receivedDateTime desc",
                },
            )
            if r.status_code >= 400:
                return {"error": f"Could not list inbox: {r.status_code} {r.text[:200]}"}
            msgs = [
                {
                    "id": m.get("id"),
                    "subject": m.get("subject") or "(no subject)",
                    "from": (m.get("from") or {}).get("emailAddress", {}).get("address"),
                    "from_name": (m.get("from") or {}).get("emailAddress", {}).get("name"),
                    "received": m.get("receivedDateTime"),
                    "snippet": m.get("bodyPreview"),
                    "unread": not m.get("isRead", True),
                    "has_attachments": m.get("hasAttachments", False),
                }
                for m in (r.json() or {}).get("value", [])
            ]
            return {"messages": msgs, "count": len(msgs), "folder": folder}

        if action == "search_emails":
            # Build a Graph $search filter from the friendly params
            terms: list[str] = []
            if params.get("query"):
                terms.append(f'"{params["query"]}"')
            if params.get("from"):
                terms.append(f'from:{params["from"]}')
            if params.get("subject"):
                terms.append(f'subject:"{params["subject"]}"')
            search_str = " ".join(terms) if terms else '""'
            limit = min(int(params.get("limit") or 20), 50)
            # $search and $orderby can't be combined in Graph mail queries —
            # the API returns results in relevance order already.
            query_params: dict = {
                "$top": limit,
                "$select": "id,subject,from,receivedDateTime,bodyPreview,isRead,hasAttachments",
            }
            if terms:
                query_params["$search"] = search_str
            if params.get("unread") in (True, "true", "True"):
                query_params["$filter"] = "isRead eq false"
            r = await client.get(
                f"{base}/me/messages",
                headers={**headers, "ConsistencyLevel": "eventual"},
                params=query_params,
            )
            if r.status_code >= 400:
                return {"error": f"Search failed: {r.status_code} {r.text[:200]}"}
            msgs = [
                {
                    "id": m.get("id"),
                    "subject": m.get("subject") or "(no subject)",
                    "from": (m.get("from") or {}).get("emailAddress", {}).get("address"),
                    "received": m.get("receivedDateTime"),
                    "snippet": m.get("bodyPreview"),
                    "unread": not m.get("isRead", True),
                    "has_attachments": m.get("hasAttachments", False),
                }
                for m in (r.json() or {}).get("value", [])
            ]
            return {"messages": msgs, "count": len(msgs)}

        if action == "read_email":
            mid = params.get("message_id")
            if not mid:
                return {"error": "message_id is required"}
            r = await client.get(
                f"{base}/me/messages/{mid}",
                headers=headers,
                params={"$select": "id,subject,from,toRecipients,ccRecipients,receivedDateTime,body,bodyPreview,hasAttachments"},
            )
            if r.status_code >= 400:
                return {"error": f"Could not read: {r.status_code}"}
            m = r.json() or {}
            attachments: list[dict] = []
            if m.get("hasAttachments"):
                ar = await client.get(
                    f"{base}/me/messages/{mid}/attachments",
                    headers=headers,
                    params={"$select": "id,name,contentType,size"},
                )
                if ar.status_code == 200:
                    attachments = [
                        {"id": a.get("id"), "name": a.get("name"), "content_type": a.get("contentType"), "size": a.get("size")}
                        for a in (ar.json() or {}).get("value", [])
                    ]
            return {
                "id": m.get("id"),
                "subject": m.get("subject") or "(no subject)",
                "from": (m.get("from") or {}).get("emailAddress", {}).get("address"),
                "to": [r.get("emailAddress", {}).get("address") for r in m.get("toRecipients", [])],
                "cc": [r.get("emailAddress", {}).get("address") for r in m.get("ccRecipients", [])],
                "received": m.get("receivedDateTime"),
                "body": (m.get("body") or {}).get("content"),
                "body_type": (m.get("body") or {}).get("contentType", "html"),
                "attachments": attachments,
            }

        if action == "reply_to_email":
            mid = params.get("message_id")
            body = params.get("body") or ""
            if not mid:
                return {"error": "message_id is required"}
            reply_all = str(params.get("reply_all", "false")).lower() in ("true", "1", "yes")
            endpoint = f"{base}/me/messages/{mid}/{'replyAll' if reply_all else 'reply'}"
            r = await client.post(endpoint, headers=headers, json={"comment": body})
            if r.status_code >= 400:
                return {"error": f"Reply failed: {r.status_code} {r.text[:200]}"}
            return {"message": "Reply sent"}

        if action == "send_email":
            to_list = _parse_mail_list(params.get("to"))
            if not to_list:
                return {"error": "to is required"}
            subject = params.get("subject") or "(no subject)"
            body = params.get("body") or ""
            cc_list = _parse_mail_list(params.get("cc"))
            bcc_list = _parse_mail_list(params.get("bcc"))
            msg = {
                "message": {
                    "subject": subject,
                    "body": {"contentType": "HTML", "content": body},
                    "toRecipients": [{"emailAddress": {"address": a}} for a in to_list],
                },
                "saveToSentItems": True,
            }
            if cc_list:
                msg["message"]["ccRecipients"] = [{"emailAddress": {"address": a}} for a in cc_list]
            if bcc_list:
                msg["message"]["bccRecipients"] = [{"emailAddress": {"address": a}} for a in bcc_list]
            r = await client.post(f"{base}/me/sendMail", headers=headers, json=msg)
            if r.status_code >= 400:
                return {"error": f"Send failed: {r.status_code} {r.text[:200]}"}
            return {"message": f"Email sent to {', '.join(to_list)}", "to": to_list, "subject": subject}

        if action in ("mark_read", "mark_unread"):
            mid = params.get("message_id")
            if not mid:
                return {"error": "message_id is required"}
            r = await client.patch(
                f"{base}/me/messages/{mid}",
                headers=headers,
                json={"isRead": action == "mark_read"},
            )
            if r.status_code >= 400:
                return {"error": f"Update failed: {r.status_code}"}
            return {"message": f"Marked {'read' if action == 'mark_read' else 'unread'}"}

        if action == "archive":
            mid = params.get("message_id")
            if not mid:
                return {"error": "message_id is required"}
            r = await client.post(
                f"{base}/me/messages/{mid}/move",
                headers=headers,
                json={"destinationId": "archive"},
            )
            if r.status_code >= 400:
                return {"error": f"Archive failed: {r.status_code}"}
            return {"message": "Archived"}

        if action == "delete":
            mid = params.get("message_id")
            if not mid:
                return {"error": "message_id is required"}
            # Move to Deleted Items instead of permanent delete
            r = await client.post(
                f"{base}/me/messages/{mid}/move",
                headers=headers,
                json={"destinationId": "deleteditems"},
            )
            if r.status_code >= 400:
                return {"error": f"Delete failed: {r.status_code}"}
            return {"message": "Moved to Deleted Items"}

        if action == "move_to_folder":
            mid = params.get("message_id")
            folder_name = params.get("folder")
            if not mid or not folder_name:
                return {"error": "message_id and folder are required"}
            fr = await client.get(f"{base}/me/mailFolders", headers=headers, params={"$top": 100})
            folders = (fr.json() or {}).get("value", [])
            match = next(
                (f for f in folders if (f.get("displayName") or "").lower() == folder_name.lower()),
                None,
            )
            if not match:
                return {"error": f"Folder '{folder_name}' not found"}
            r = await client.post(
                f"{base}/me/messages/{mid}/move",
                headers=headers,
                json={"destinationId": match["id"]},
            )
            if r.status_code >= 400:
                return {"error": f"Move failed: {r.status_code}"}
            return {"message": f"Moved to {folder_name}"}

        if action == "list_folders":
            r = await client.get(f"{base}/me/mailFolders", headers=headers, params={"$top": 200})
            if r.status_code >= 400:
                return {"error": f"Could not list folders: {r.status_code}"}
            folders = [
                {"id": f.get("id"), "name": f.get("displayName"), "unread": f.get("unreadItemCount", 0), "total": f.get("totalItemCount", 0)}
                for f in (r.json() or {}).get("value", [])
            ]
            return {"folders": folders, "count": len(folders)}

        if action == "download_attachment":
            mid = params.get("message_id")
            aid = params.get("attachment_id")
            if not mid or not aid:
                return {"error": "message_id and attachment_id are required"}
            r = await client.get(f"{base}/me/messages/{mid}/attachments/{aid}", headers=headers)
            if r.status_code >= 400:
                return {"error": f"Could not fetch attachment: {r.status_code}"}
            att = r.json() or {}
            return {
                "filename": att.get("name"),
                "mime_type": att.get("contentType"),
                "size": att.get("size"),
                "content_base64": att.get("contentBytes"),  # Graph already gives base64
            }

    return {"error": f"Unknown Outlook Mail action: {action}"}


async def _gmail_adapter(action: str, params: dict, creds: dict) -> dict:
    """Gmail adapter via the Gmail API.

    Note: Gmail doesn't have native "folders" — it has labels. We map
    move_to_folder → add label, archive → remove INBOX label, etc. so the
    action surface matches Outlook/IMAP.
    """
    token = creds.get("access_token") or creds.get("api_key")
    if not token:
        return {"error": "Gmail is not connected. Connect it in Settings → My Apps → Gmail."}

    base = "https://gmail.googleapis.com/gmail/v1/users/me"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    def _header(msg: dict, name: str) -> str:
        for h in (msg.get("payload") or {}).get("headers", []):
            if h.get("name", "").lower() == name.lower():
                return h.get("value") or ""
        return ""

    def _extract_body(payload: dict) -> tuple[str, str]:
        """Walk a MIME tree and return (html_or_plain_body, content_type)."""
        import base64 as _b64
        if not payload:
            return "", "text/plain"
        mime = payload.get("mimeType", "")
        if mime in ("text/plain", "text/html") and (payload.get("body") or {}).get("data"):
            data = payload["body"]["data"]
            try:
                decoded = _b64.urlsafe_b64decode(data + "==").decode("utf-8", errors="replace")
                return decoded, mime
            except Exception:
                return "", mime
        for part in payload.get("parts") or []:
            body, ctype = _extract_body(part)
            if body:
                return body, ctype
        return "", mime or "text/plain"

    def _build_raw(to: list[str], subject: str, body: str, cc: list[str], bcc: list[str], in_reply_to: str | None = None, references: str | None = None) -> str:
        """Build a base64url-encoded RFC 2822 message for Gmail's API."""
        import base64 as _b64
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        msg = MIMEMultipart("alternative")
        msg["To"] = ", ".join(to)
        if cc:
            msg["Cc"] = ", ".join(cc)
        if bcc:
            msg["Bcc"] = ", ".join(bcc)
        msg["Subject"] = subject
        if in_reply_to:
            msg["In-Reply-To"] = in_reply_to
        if references:
            msg["References"] = references
        msg.attach(MIMEText(body, "html"))
        return _b64.urlsafe_b64encode(msg.as_bytes()).decode("ascii").rstrip("=")

    async with httpx.AsyncClient(timeout=30) as client:
        if action == "list_inbox":
            label = params.get("label") or params.get("folder") or "INBOX"
            limit = min(int(params.get("limit") or 20), 50)
            r = await client.get(
                f"{base}/messages",
                headers=headers,
                params={"labelIds": label.upper() if label.lower() == "inbox" else label, "maxResults": limit},
            )
            if r.status_code >= 400:
                return {"error": f"List failed: {r.status_code} {r.text[:200]}"}
            ids = [m["id"] for m in (r.json() or {}).get("messages", [])]
            # Fetch metadata for each in parallel batches
            msgs: list[dict] = []
            for mid in ids:
                mr = await client.get(
                    f"{base}/messages/{mid}",
                    headers=headers,
                    params={"format": "metadata", "metadataHeaders": ["From", "Subject", "Date"]},
                )
                if mr.status_code != 200:
                    continue
                m = mr.json() or {}
                msgs.append({
                    "id": m.get("id"),
                    "thread_id": m.get("threadId"),
                    "subject": _header(m, "Subject") or "(no subject)",
                    "from": _header(m, "From"),
                    "received": _header(m, "Date"),
                    "snippet": m.get("snippet"),
                    "unread": "UNREAD" in (m.get("labelIds") or []),
                })
            return {"messages": msgs, "count": len(msgs), "label": label}

        if action == "search_emails":
            # Gmail has native search grammar — accept either the raw query
            # or the friendly fields and combine them.
            q_parts: list[str] = []
            if params.get("query"):
                q_parts.append(str(params["query"]))
            if params.get("from"):
                q_parts.append(f"from:{params['from']}")
            if params.get("subject"):
                q_parts.append(f"subject:\"{params['subject']}\"")
            if params.get("unread") in (True, "true", "True"):
                q_parts.append("is:unread")
            q = " ".join(q_parts)
            limit = min(int(params.get("limit") or 20), 50)
            r = await client.get(
                f"{base}/messages",
                headers=headers,
                params={"q": q, "maxResults": limit},
            )
            if r.status_code >= 400:
                return {"error": f"Search failed: {r.status_code} {r.text[:200]}"}
            ids = [m["id"] for m in (r.json() or {}).get("messages", [])]
            msgs: list[dict] = []
            for mid in ids:
                mr = await client.get(
                    f"{base}/messages/{mid}",
                    headers=headers,
                    params={"format": "metadata", "metadataHeaders": ["From", "Subject", "Date"]},
                )
                if mr.status_code != 200:
                    continue
                m = mr.json() or {}
                msgs.append({
                    "id": m.get("id"),
                    "thread_id": m.get("threadId"),
                    "subject": _header(m, "Subject") or "(no subject)",
                    "from": _header(m, "From"),
                    "received": _header(m, "Date"),
                    "snippet": m.get("snippet"),
                    "unread": "UNREAD" in (m.get("labelIds") or []),
                })
            return {"messages": msgs, "count": len(msgs), "query": q}

        if action == "read_email":
            mid = params.get("message_id")
            if not mid:
                return {"error": "message_id is required"}
            r = await client.get(f"{base}/messages/{mid}", headers=headers, params={"format": "full"})
            if r.status_code >= 400:
                return {"error": f"Could not read: {r.status_code}"}
            m = r.json() or {}
            body, ctype = _extract_body(m.get("payload") or {})
            # Collect attachment metadata
            attachments: list[dict] = []
            def _walk_attachments(part: dict) -> None:
                if not part:
                    return
                if part.get("filename") and (part.get("body") or {}).get("attachmentId"):
                    attachments.append({
                        "id": part["body"]["attachmentId"],
                        "name": part.get("filename"),
                        "content_type": part.get("mimeType"),
                        "size": (part.get("body") or {}).get("size"),
                    })
                for sub in part.get("parts") or []:
                    _walk_attachments(sub)
            _walk_attachments(m.get("payload") or {})
            return {
                "id": m.get("id"),
                "thread_id": m.get("threadId"),
                "subject": _header(m, "Subject") or "(no subject)",
                "from": _header(m, "From"),
                "to": _header(m, "To"),
                "cc": _header(m, "Cc"),
                "received": _header(m, "Date"),
                "body": body,
                "body_type": ctype,
                "attachments": attachments,
            }

        if action == "reply_to_email":
            mid = params.get("message_id")
            body_text = params.get("body") or ""
            if not mid:
                return {"error": "message_id is required"}
            # Fetch the original to get thread id + headers
            orig = await client.get(f"{base}/messages/{mid}", headers=headers, params={"format": "metadata", "metadataHeaders": ["From", "Subject", "Message-ID", "References"]})
            if orig.status_code != 200:
                return {"error": f"Could not load original: {orig.status_code}"}
            om = orig.json() or {}
            orig_from = _header(om, "From")
            orig_subj = _header(om, "Subject") or ""
            orig_mid = _header(om, "Message-ID")
            orig_refs = _header(om, "References")
            reply_subject = orig_subj if orig_subj.lower().startswith("re:") else f"Re: {orig_subj}"
            # Pull the email address out of "Name <addr@...>"
            import re as _re
            m_addr = _re.search(r"<([^>]+)>", orig_from)
            to_addr = m_addr.group(1) if m_addr else orig_from
            raw = _build_raw([to_addr], reply_subject, body_text, [], [], in_reply_to=orig_mid, references=(orig_refs + " " + orig_mid).strip() if orig_refs else orig_mid)
            r = await client.post(
                f"{base}/messages/send",
                headers=headers,
                json={"raw": raw, "threadId": om.get("threadId")},
            )
            if r.status_code >= 400:
                return {"error": f"Reply failed: {r.status_code} {r.text[:200]}"}
            return {"message": "Reply sent"}

        if action == "send_email":
            to_list = _parse_mail_list(params.get("to"))
            if not to_list:
                return {"error": "to is required"}
            subject = params.get("subject") or "(no subject)"
            body = params.get("body") or ""
            cc_list = _parse_mail_list(params.get("cc"))
            bcc_list = _parse_mail_list(params.get("bcc"))
            raw = _build_raw(to_list, subject, body, cc_list, bcc_list)
            r = await client.post(f"{base}/messages/send", headers=headers, json={"raw": raw})
            if r.status_code >= 400:
                return {"error": f"Send failed: {r.status_code} {r.text[:200]}"}
            return {"message": f"Email sent to {', '.join(to_list)}", "to": to_list, "subject": subject}

        if action in ("mark_read", "mark_unread"):
            mid = params.get("message_id")
            if not mid:
                return {"error": "message_id is required"}
            mod = {"removeLabelIds": ["UNREAD"]} if action == "mark_read" else {"addLabelIds": ["UNREAD"]}
            r = await client.post(f"{base}/messages/{mid}/modify", headers=headers, json=mod)
            if r.status_code >= 400:
                return {"error": f"Update failed: {r.status_code}"}
            return {"message": f"Marked {'read' if action == 'mark_read' else 'unread'}"}

        if action == "archive":
            mid = params.get("message_id")
            if not mid:
                return {"error": "message_id is required"}
            r = await client.post(f"{base}/messages/{mid}/modify", headers=headers, json={"removeLabelIds": ["INBOX"]})
            if r.status_code >= 400:
                return {"error": f"Archive failed: {r.status_code}"}
            return {"message": "Archived"}

        if action == "delete":
            mid = params.get("message_id")
            if not mid:
                return {"error": "message_id is required"}
            # Gmail's "trash" endpoint moves to Trash rather than permanent delete
            r = await client.post(f"{base}/messages/{mid}/trash", headers=headers)
            if r.status_code >= 400:
                return {"error": f"Delete failed: {r.status_code}"}
            return {"message": "Moved to Trash"}

        if action == "move_to_folder":
            mid = params.get("message_id")
            folder = params.get("folder") or params.get("to")
            if not mid or not folder:
                return {"error": "message_id and folder are required"}
            # Look up the label id
            lr = await client.get(f"{base}/labels", headers=headers)
            if lr.status_code >= 400:
                return {"error": f"Could not list labels: {lr.status_code}"}
            labels = (lr.json() or {}).get("labels", [])
            match = next((l for l in labels if (l.get("name") or "").lower() == folder.lower()), None)
            if not match:
                return {"error": f"Label '{folder}' not found. Use list_folders to see available labels."}
            r = await client.post(
                f"{base}/messages/{mid}/modify",
                headers=headers,
                json={"addLabelIds": [match["id"]], "removeLabelIds": ["INBOX"]},
            )
            if r.status_code >= 400:
                return {"error": f"Move failed: {r.status_code}"}
            return {"message": f"Moved to {folder}"}

        if action == "list_folders":
            r = await client.get(f"{base}/labels", headers=headers)
            if r.status_code >= 400:
                return {"error": f"Could not list labels: {r.status_code}"}
            folders = [{"id": l.get("id"), "name": l.get("name"), "type": l.get("type")} for l in (r.json() or {}).get("labels", [])]
            return {"folders": folders, "count": len(folders)}

        if action == "download_attachment":
            mid = params.get("message_id")
            aid = params.get("attachment_id")
            if not mid or not aid:
                return {"error": "message_id and attachment_id are required"}
            r = await client.get(f"{base}/messages/{mid}/attachments/{aid}", headers=headers)
            if r.status_code >= 400:
                return {"error": f"Could not fetch attachment: {r.status_code}"}
            att = r.json() or {}
            import base64 as _b64
            # Gmail returns base64url — decode and re-encode as standard base64
            try:
                raw = _b64.urlsafe_b64decode(att.get("data", "") + "==")
                std_b64 = _b64.b64encode(raw).decode("ascii")
            except Exception:
                std_b64 = att.get("data", "")
            return {
                "size": att.get("size"),
                "content_base64": std_b64,
            }

    return {"error": f"Unknown Gmail action: {action}"}


async def _imap_mail_adapter(action: str, params: dict, creds: dict) -> dict:
    """Generic IMAP + SMTP adapter. Works with any provider that supports
    IMAP (Yahoo, ProtonMail Bridge, FastMail, iCloud, custom domains, etc.).

    Uses the stdlib imaplib/smtplib in a thread pool because neither has
    async support built-in. That's fine for read/list/search but means
    attachment downloads and bulk actions can be slow — mark as TODO if
    that becomes a pain point.
    """
    import imaplib
    import smtplib
    from email import message_from_bytes
    from email.header import decode_header
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    username = creds.get("username") or creds.get("email")
    password = creds.get("app_password") or creds.get("password")
    access_token = creds.get("access_token")
    # XOAUTH2 flow: if we have an OAuth access token (from Yahoo / Zoho /
    # Fastmail / Mail.ru / Yandex), use it instead of asking for a password.
    # The user never touches an app password in that case.
    if not (username and (password or access_token)):
        return {"error": "IMAP mailbox needs at least a username (email) and either an app password or an OAuth access token."}

    imap_host = creds.get("imap_host")
    smtp_host = creds.get("smtp_host")
    imap_port = int(creds.get("imap_port") or 0) or None
    smtp_port = int(creds.get("smtp_port") or 0) or None

    # If the caller didn't pin hosts, auto-detect them from the email
    # domain using Mozilla ISPDB / autoconfig / heuristics.
    if not (imap_host and smtp_host):
        detected = await _autodetect_mail_servers(username)
        if detected:
            imap_host = imap_host or detected.get("imap_host")
            smtp_host = smtp_host or detected.get("smtp_host")
            imap_port = imap_port or detected.get("imap_port")
            smtp_port = smtp_port or detected.get("smtp_port")

    imap_port = imap_port or 993
    smtp_port = smtp_port or 587

    if not (imap_host and smtp_host):
        domain = username.rsplit("@", 1)[-1] if "@" in username else username
        return {"error": f"We couldn't autodetect the email servers for {domain}. Go to Settings → Connect Apps → Email (IMAP) and enter the IMAP and SMTP hosts manually (your email provider's docs usually call them 'imap.yourdomain.com' and 'smtp.yourdomain.com' or similar)."}

    def _decode(val: str | bytes | None) -> str:
        if val is None:
            return ""
        if isinstance(val, bytes):
            try:
                return val.decode("utf-8", errors="replace")
            except Exception:
                return str(val)
        # Handle RFC 2047 encoded-word headers ("=?utf-8?b?...?=")
        try:
            parts = decode_header(val)
            out = ""
            for txt, enc in parts:
                if isinstance(txt, bytes):
                    out += txt.decode(enc or "utf-8", errors="replace")
                else:
                    out += txt
            return out
        except Exception:
            return str(val)

    def _xoauth2_string(user: str, token: str) -> bytes:
        """Build an XOAUTH2 authentication string per RFC 7628.
        Format: user=<email>\\x01auth=Bearer <token>\\x01\\x01"""
        return f"user={user}\x01auth=Bearer {token}\x01\x01".encode("ascii")

    def _connect_imap() -> "imaplib.IMAP4_SSL":
        c = imaplib.IMAP4_SSL(imap_host, imap_port)
        if access_token:
            # XOAUTH2 login — no password involved
            auth_blob = _xoauth2_string(username, access_token)
            c.authenticate("XOAUTH2", lambda _: auth_blob)
        else:
            c.login(username, password)
        return c

    def _list_inbox_sync(folder: str, limit: int) -> list[dict]:
        c = _connect_imap()
        try:
            c.select(folder, readonly=True)
            typ, data = c.search(None, "ALL")
            if typ != "OK":
                return []
            ids = (data[0] or b"").split()
            ids = ids[-limit:][::-1]  # newest first
            out: list[dict] = []
            for uid in ids:
                typ, msg_data = c.fetch(uid, "(RFC822.HEADER FLAGS)")
                if typ != "OK" or not msg_data:
                    continue
                header_bytes = next((d[1] for d in msg_data if isinstance(d, tuple)), None)
                if not header_bytes:
                    continue
                m = message_from_bytes(header_bytes)
                flags = ""
                for d in msg_data:
                    if isinstance(d, bytes) and b"FLAGS" in d:
                        flags = d.decode("utf-8", errors="replace")
                out.append({
                    "id": uid.decode("ascii"),
                    "subject": _decode(m.get("Subject")) or "(no subject)",
                    "from": _decode(m.get("From")),
                    "received": _decode(m.get("Date")),
                    "snippet": "",
                    "unread": "\\Seen" not in flags,
                })
            return out
        finally:
            try: c.logout()
            except Exception: pass

    def _read_email_sync(folder: str, uid: str) -> dict:
        c = _connect_imap()
        try:
            c.select(folder, readonly=True)
            typ, data = c.fetch(uid.encode("ascii"), "(RFC822)")
            if typ != "OK" or not data:
                return {"error": "Message not found"}
            raw = next((d[1] for d in data if isinstance(d, tuple)), None)
            if not raw:
                return {"error": "No message body"}
            m = message_from_bytes(raw)
            body_text = ""
            body_html = ""
            attachments: list[dict] = []
            for i, part in enumerate(m.walk() if m.is_multipart() else [m]):
                ctype = part.get_content_type()
                disp = str(part.get("Content-Disposition") or "")
                if "attachment" in disp.lower() or part.get_filename():
                    attachments.append({
                        "index": i,
                        "name": _decode(part.get_filename()) or f"attachment-{i}",
                        "content_type": ctype,
                        "size": len(part.get_payload(decode=True) or b""),
                    })
                    continue
                if ctype == "text/plain" and not body_text:
                    try:
                        body_text = (part.get_payload(decode=True) or b"").decode("utf-8", errors="replace")
                    except Exception:
                        pass
                elif ctype == "text/html" and not body_html:
                    try:
                        body_html = (part.get_payload(decode=True) or b"").decode("utf-8", errors="replace")
                    except Exception:
                        pass
            return {
                "id": uid,
                "subject": _decode(m.get("Subject")) or "(no subject)",
                "from": _decode(m.get("From")),
                "to": _decode(m.get("To")),
                "cc": _decode(m.get("Cc")),
                "received": _decode(m.get("Date")),
                "body": body_html or body_text,
                "body_type": "text/html" if body_html else "text/plain",
                "attachments": attachments,
            }
        finally:
            try: c.logout()
            except Exception: pass

    def _set_flag_sync(folder: str, uid: str, flag: str, add: bool) -> None:
        c = _connect_imap()
        try:
            c.select(folder)
            c.uid("STORE", uid, "+FLAGS" if add else "-FLAGS", f"({flag})")
        finally:
            try: c.logout()
            except Exception: pass

    def _move_sync(folder: str, uid: str, dest: str) -> None:
        c = _connect_imap()
        try:
            c.select(folder)
            # Try MOVE first (RFC 6851), fall back to COPY + delete
            typ, _ = c.uid("MOVE", uid, dest)
            if typ != "OK":
                c.uid("COPY", uid, dest)
                c.uid("STORE", uid, "+FLAGS", "(\\Deleted)")
                c.expunge()
        finally:
            try: c.logout()
            except Exception: pass

    def _list_folders_sync() -> list[dict]:
        c = _connect_imap()
        try:
            typ, data = c.list()
            if typ != "OK":
                return []
            out: list[dict] = []
            for line in data or []:
                if not line:
                    continue
                parts = line.decode("utf-8", errors="replace").split(' "/" ')
                if len(parts) >= 2:
                    name = parts[-1].strip().strip('"')
                    out.append({"name": name})
            return out
        finally:
            try: c.logout()
            except Exception: pass

    def _send_sync(to_list: list[str], subject: str, body: str, cc_list: list[str], bcc_list: list[str]) -> None:
        msg = MIMEMultipart("alternative")
        msg["From"] = username
        msg["To"] = ", ".join(to_list)
        if cc_list: msg["Cc"] = ", ".join(cc_list)
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "html"))
        recipients = to_list + cc_list + bcc_list
        with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as s:
            s.ehlo()
            try:
                s.starttls()
                s.ehlo()
            except Exception:
                pass
            if access_token:
                # SMTP XOAUTH2: base64-encode the same auth blob we use
                # for IMAP and send it via AUTH XOAUTH2.
                import base64 as _b64
                auth_blob = _xoauth2_string(username, access_token)
                encoded = _b64.b64encode(auth_blob).decode("ascii")
                code, resp = s.docmd("AUTH", f"XOAUTH2 {encoded}")
                if code != 235:
                    # Some servers return 334 with a challenge that we just
                    # acknowledge with an empty line.
                    if code == 334:
                        code, resp = s.docmd("")
                    if code != 235:
                        raise smtplib.SMTPAuthenticationError(code, resp)
            else:
                s.login(username, password)
            s.sendmail(username, recipients, msg.as_string())

    import asyncio as _aio
    try:
        if action == "list_inbox":
            folder = params.get("folder") or "INBOX"
            limit = min(int(params.get("limit") or 20), 50)
            msgs = await _aio.to_thread(_list_inbox_sync, folder, limit)
            return {"messages": msgs, "count": len(msgs), "folder": folder}

        if action == "search_emails":
            # IMAP search is its own mini-language. We build a basic query
            # from the friendly params; the user can also pass a raw query.
            folder = params.get("folder") or "INBOX"
            limit = min(int(params.get("limit") or 20), 50)
            criteria: list[str] = []
            if params.get("from"):
                criteria.append(f'FROM "{params["from"]}"')
            if params.get("subject"):
                criteria.append(f'SUBJECT "{params["subject"]}"')
            if params.get("query"):
                criteria.append(f'BODY "{params["query"]}"')
            if params.get("unread") in (True, "true", "True"):
                criteria.append("UNSEEN")
            query = " ".join(criteria) or "ALL"

            def _search_sync() -> list[dict]:
                c = _connect_imap()
                try:
                    c.select(folder, readonly=True)
                    typ, data = c.search(None, query)
                    if typ != "OK":
                        return []
                    ids = (data[0] or b"").split()[-limit:][::-1]
                    out: list[dict] = []
                    for uid in ids:
                        typ, msg_data = c.fetch(uid, "(RFC822.HEADER)")
                        if typ != "OK" or not msg_data:
                            continue
                        header_bytes = next((d[1] for d in msg_data if isinstance(d, tuple)), None)
                        if not header_bytes:
                            continue
                        m = message_from_bytes(header_bytes)
                        out.append({
                            "id": uid.decode("ascii"),
                            "subject": _decode(m.get("Subject")) or "(no subject)",
                            "from": _decode(m.get("From")),
                            "received": _decode(m.get("Date")),
                        })
                    return out
                finally:
                    try: c.logout()
                    except Exception: pass

            msgs = await _aio.to_thread(_search_sync)
            return {"messages": msgs, "count": len(msgs), "query": query}

        if action == "read_email":
            folder = params.get("folder") or "INBOX"
            uid = params.get("message_id") or params.get("uid")
            if not uid:
                return {"error": "message_id is required"}
            return await _aio.to_thread(_read_email_sync, folder, str(uid))

        if action == "reply_to_email":
            folder = params.get("folder") or "INBOX"
            uid = params.get("message_id")
            body = params.get("body") or ""
            if not uid:
                return {"error": "message_id is required"}
            # Load the original to get From + Subject for the reply
            orig = await _aio.to_thread(_read_email_sync, folder, str(uid))
            if "error" in orig:
                return orig
            to_addr = orig.get("from") or ""
            import re as _re
            m_addr = _re.search(r"<([^>]+)>", to_addr)
            if m_addr:
                to_addr = m_addr.group(1)
            reply_subject = orig.get("subject") or ""
            if not reply_subject.lower().startswith("re:"):
                reply_subject = f"Re: {reply_subject}"
            await _aio.to_thread(_send_sync, [to_addr], reply_subject, body, [], [])
            return {"message": "Reply sent"}

        if action == "send_email":
            to_list = _parse_mail_list(params.get("to"))
            if not to_list:
                return {"error": "to is required"}
            await _aio.to_thread(
                _send_sync,
                to_list,
                params.get("subject") or "(no subject)",
                params.get("body") or "",
                _parse_mail_list(params.get("cc")),
                _parse_mail_list(params.get("bcc")),
            )
            return {"message": f"Email sent to {', '.join(to_list)}"}

        if action in ("mark_read", "mark_unread"):
            folder = params.get("folder") or "INBOX"
            uid = params.get("message_id")
            if not uid:
                return {"error": "message_id is required"}
            await _aio.to_thread(_set_flag_sync, folder, str(uid), "\\Seen", action == "mark_read")
            return {"message": f"Marked {'read' if action == 'mark_read' else 'unread'}"}

        if action == "archive":
            folder = params.get("folder") or "INBOX"
            uid = params.get("message_id")
            if not uid:
                return {"error": "message_id is required"}
            # Common archive folder names across providers
            await _aio.to_thread(_move_sync, folder, str(uid), params.get("archive_folder") or "Archive")
            return {"message": "Archived"}

        if action == "delete":
            folder = params.get("folder") or "INBOX"
            uid = params.get("message_id")
            if not uid:
                return {"error": "message_id is required"}
            await _aio.to_thread(_set_flag_sync, folder, str(uid), "\\Deleted", True)
            return {"message": "Marked for deletion (will be removed on next expunge)"}

        if action == "move_to_folder":
            src = params.get("folder") or "INBOX"
            dest = params.get("to") or params.get("destination")
            uid = params.get("message_id")
            if not uid or not dest:
                return {"error": "message_id and to (destination folder) are required"}
            await _aio.to_thread(_move_sync, src, str(uid), dest)
            return {"message": f"Moved to {dest}"}

        if action == "list_folders":
            folders = await _aio.to_thread(_list_folders_sync)
            return {"folders": folders, "count": len(folders)}

        if action == "download_attachment":
            folder = params.get("folder") or "INBOX"
            uid = params.get("message_id")
            idx = int(params.get("attachment_index") or 0)
            if not uid:
                return {"error": "message_id is required"}

            def _fetch_sync() -> dict:
                c = _connect_imap()
                try:
                    c.select(folder, readonly=True)
                    typ, data = c.fetch(uid.encode("ascii"), "(RFC822)")
                    if typ != "OK" or not data:
                        return {"error": "Message not found"}
                    raw = next((d[1] for d in data if isinstance(d, tuple)), None)
                    if not raw:
                        return {"error": "No message body"}
                    m = message_from_bytes(raw)
                    import base64 as _b64
                    current = -1
                    for part in (m.walk() if m.is_multipart() else [m]):
                        disp = str(part.get("Content-Disposition") or "")
                        if "attachment" in disp.lower() or part.get_filename():
                            current += 1
                            if current == idx:
                                payload = part.get_payload(decode=True) or b""
                                return {
                                    "filename": _decode(part.get_filename()) or f"attachment-{idx}",
                                    "mime_type": part.get_content_type(),
                                    "size": len(payload),
                                    "content_base64": _b64.b64encode(payload).decode("ascii"),
                                }
                    return {"error": f"No attachment at index {idx}"}
                finally:
                    try: c.logout()
                    except Exception: pass

            return await _aio.to_thread(_fetch_sync)

    except Exception as e:
        logger.exception("IMAP mail action failed")
        msg = str(e)
        domain = username.rsplit("@", 1)[-1].lower() if "@" in username else ""
        # DNS resolution failure — the host we were given doesn't exist.
        if "[Errno -2]" in msg or "Name or service not known" in msg or "nodename nor servname" in msg or "getaddrinfo" in msg.lower():
            return {"error": f"Could not connect to {imap_host}:{imap_port} — hostname not found. Go to Settings → Connect Apps → Email (IMAP) and double-check the IMAP Server field, or ask your email provider for the correct IMAP hostname."}
        # Auth failure — clearly tell the user the password is wrong,
        # with provider-specific pointers where we can give them.
        if "authentication" in msg.lower() or "login failed" in msg.lower() or "AUTHENTICATIONFAILED" in msg or "[AUTHENTICATIONFAILED]" in msg:
            # Titan / Neo specific guidance — these share the same admin
            # panel and both require an app password when 2FA is enabled.
            if "titan" in (imap_host or "").lower() or "neo.space" in (imap_host or "").lower() or imap_host in ("imap.titan.email", "imap.neo.space"):
                return {"error": f"Titan/Neo rejected the login for {username}. Server said: {msg}\n\nFixes to try:\n1. Double-check the password (log in to app.titan.email first to confirm).\n2. If you have 2FA on, generate an app password in Titan's admin panel → Settings → Security → App Passwords.\n3. Disconnect and reconnect in Settings → Connect Apps → Titan Email."}
            return {"error": f"IMAP login failed for {username}. Server said: {msg}\n\nMost providers require an 'app password' instead of your regular password — check your email provider's security settings and generate one."}
        # Connection refused / timed out
        if "refused" in msg.lower() or "timed out" in msg.lower() or "timeout" in msg.lower():
            return {"error": f"Could not reach {imap_host}:{imap_port} ({msg}). Check that your provider allows IMAP access and the port is correct (usually 993)."}
        return {"error": f"IMAP error ({imap_host}:{imap_port}): {msg}"}

    return {"error": f"Unknown IMAP action: {action}"}


# ── Preset IMAP wrappers (Neo, Titan, etc.) ─────────────────────────────
#
# These are just the generic IMAP adapter with the host/port pre-filled so
# the user only has to provide their email + password. New presets should
# follow the same one-liner pattern.

async def _neo_mail_adapter(action: str, params: dict, creds: dict) -> dict:
    """Neo Business Email (neo.space) — IMAP/SMTP preset."""
    enriched = dict(creds)
    enriched.setdefault("imap_host", "imap.neo.space")
    enriched.setdefault("imap_port", 993)
    enriched.setdefault("smtp_host", "smtp.neo.space")
    enriched.setdefault("smtp_port", 465)
    return await _imap_mail_adapter(action, params, enriched)


async def _titan_mail_adapter(action: str, params: dict, creds: dict) -> dict:
    """Titan Email (titan.email) — IMAP/SMTP preset."""
    enriched = dict(creds)
    enriched.setdefault("imap_host", "imap.titan.email")
    enriched.setdefault("imap_port", 993)
    enriched.setdefault("smtp_host", "smtp.titan.email")
    enriched.setdefault("smtp_port", 587)
    return await _imap_mail_adapter(action, params, enriched)


def _make_imap_preset(imap_host: str, imap_port: int, smtp_host: str, smtp_port: int):
    """Build a thin adapter wrapper that pins IMAP/SMTP hosts for a preset."""
    async def _adapter(action: str, params: dict, creds: dict) -> dict:
        enriched = dict(creds)
        enriched.setdefault("imap_host", imap_host)
        enriched.setdefault("imap_port", imap_port)
        enriched.setdefault("smtp_host", smtp_host)
        enriched.setdefault("smtp_port", smtp_port)
        return await _imap_mail_adapter(action, params, enriched)
    return _adapter


_yahoo_mail_adapter      = _make_imap_preset("imap.mail.yahoo.com",  993, "smtp.mail.yahoo.com",  587)
_icloud_mail_adapter     = _make_imap_preset("imap.mail.me.com",     993, "smtp.mail.me.com",     587)
_zoho_mail_adapter       = _make_imap_preset("imap.zoho.com",        993, "smtp.zoho.com",        587)
_fastmail_mail_adapter   = _make_imap_preset("imap.fastmail.com",    993, "smtp.fastmail.com",    587)
_aol_mail_adapter        = _make_imap_preset("imap.aol.com",         993, "smtp.aol.com",         587)
_gmx_mail_adapter        = _make_imap_preset("imap.gmx.com",         993, "mail.gmx.com",         587)
_mailru_mail_adapter     = _make_imap_preset("imap.mail.ru",         993, "smtp.mail.ru",         587)
_yandex_mail_adapter     = _make_imap_preset("imap.yandex.com",      993, "smtp.yandex.com",      587)
_protonmail_mail_adapter = _make_imap_preset("127.0.0.1",            1143, "127.0.0.1",           1025)  # ProtonMail Bridge
_hostinger_mail_adapter  = _make_imap_preset("imap.hostinger.com",   993, "smtp.hostinger.com",   587)
_godaddy_mail_adapter    = _make_imap_preset("imap.secureserver.net", 993, "smtpout.secureserver.net", 587)
_namecheap_mail_adapter  = _make_imap_preset("mail.privateemail.com", 993, "mail.privateemail.com", 587)
_ionos_mail_adapter      = _make_imap_preset("imap.ionos.com",       993, "smtp.ionos.com",       587)
_mailboxorg_mail_adapter = _make_imap_preset("imap.mailbox.org",     993, "smtp.mailbox.org",     587)
_posteo_mail_adapter     = _make_imap_preset("posteo.de",            993, "posteo.de",            587)
_mailfence_mail_adapter  = _make_imap_preset("imap.mailfence.com",   993, "smtp.mailfence.com",   587)


# ── Adapter map ──────────────────────────────────────────────────────────

ADAPTERS: dict[str, Any] = {
    # CRM
    "hubspot": _hubspot_adapter,
    "salesforce": _salesforce_adapter,
    "ringy": _ringy_adapter,
    "pipedrive": _pipedrive_adapter,
    # PM
    "asana": _asana_adapter,
    "trello": _trello_adapter,
    "notion": _notion_adapter,
    "clickup": _clickup_adapter,
    "todoist": _todoist_adapter,
    "jira": _jira_adapter,
    "linear": _linear_adapter,
    # Communication
    "slack": _slack_adapter,
    "telegram": _telegram_adapter,
    # E-commerce
    "stripe": _stripe_adapter,
    "shopify": _shopify_adapter,
    # Storage
    "airtable": _airtable_adapter,
    "excel_online": _excel_online_adapter,
    # Email (mailbox access)
    "outlook_mail": _outlook_mail_adapter,
    "gmail": _gmail_adapter,
    "neo_mail": _neo_mail_adapter,
    "titan_mail": _titan_mail_adapter,
    # IMAP presets (share _imap_mail_adapter via a preset wrapper that
    # pins the host/port so the user only enters email + password)
    "yahoo_mail": _yahoo_mail_adapter,
    "icloud_mail": _icloud_mail_adapter,
    "zoho_mail": _zoho_mail_adapter,
    "fastmail_mail": _fastmail_mail_adapter,
    "aol_mail": _aol_mail_adapter,
    "gmx_mail": _gmx_mail_adapter,
    "mailru_mail": _mailru_mail_adapter,
    "yandex_mail": _yandex_mail_adapter,
    "protonmail_mail": _protonmail_mail_adapter,
    "hostinger_mail": _hostinger_mail_adapter,
    "godaddy_mail": _godaddy_mail_adapter,
    "namecheap_mail": _namecheap_mail_adapter,
    "ionos_mail": _ionos_mail_adapter,
    "mailboxorg_mail": _mailboxorg_mail_adapter,
    "posteo_mail": _posteo_mail_adapter,
    "mailfence_mail": _mailfence_mail_adapter,
    "imap_mail": _imap_mail_adapter,
    # Automation
    "zapier": _webhook_adapter,
    "make": _webhook_adapter,
    "n8n": _webhook_adapter,
    "ifttt": _webhook_adapter,
}

# Fill in generic adapter for all apps without a specific implementation
for _app_id in APP_REGISTRY:
    if _app_id not in ADAPTERS:
        ADAPTERS[_app_id] = lambda action, params, creds, aid=_app_id: _generic_rest_adapter(aid, action, params, creds)
