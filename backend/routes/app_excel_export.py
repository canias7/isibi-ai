from __future__ import annotations

"""
Excel Export — export app data tables as .xlsx files.
Uses openpyxl if available, falls back to CSV with .xlsx extension.
"""

import io
import re
import uuid
from typing import Any

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from auth import get_current_org_id
from db import DATABASE_URL, get_db
from generator.app_db import get_schema_name, _get_raw_connection, list_schema_tables

router = APIRouter(prefix="/apps", tags=["App Excel Export"])

_IDENT_RE = re.compile(r"^[a-zA-Z_][a-zA-Z0-9_]*$")


def _safe_ident(name: str) -> str:
    clean = name.strip().lower()
    if not _IDENT_RE.match(clean):
        raise HTTPException(status_code=400, detail=f"Invalid identifier: {name}")
    return clean


def _cell_value(v: Any) -> Any:
    """Convert a value to something safe for Excel."""
    if v is None:
        return ""
    if hasattr(v, "hex"):
        return str(v)
    if hasattr(v, "isoformat"):
        return v.isoformat()
    if isinstance(v, (dict, list)):
        import json
        return json.dumps(v)
    return v


@router.get("/{project_id}/data/{table_name}/export-xlsx")
async def export_xlsx(
    project_id: str,
    table_name: str,
    org_id: uuid.UUID = Depends(get_current_org_id),
    db: AsyncSession = Depends(get_db),
):
    """Export a table as a formatted .xlsx file."""
    table = _safe_ident(table_name)
    schema = get_schema_name(project_id)

    tables = await list_schema_tables(project_id, DATABASE_URL)
    if table not in tables:
        raise HTTPException(status_code=404, detail=f"Table '{table}' not found in app schema")

    conn = await _get_raw_connection(DATABASE_URL)
    try:
        await conn.execute(f'SET search_path TO "{schema}"')
        rows = await conn.fetch(
            f'SELECT * FROM "{table}" WHERE "deleted_at" IS NULL ORDER BY "created_at" DESC LIMIT 10000'
        )

        if not rows:
            raise HTTPException(status_code=404, detail="No data to export")

        columns = list(rows[0].keys())

        # Try openpyxl first
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = table

            # Header row with styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Data rows
            for row_idx, row in enumerate(rows, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    ws.cell(row=row_idx, column=col_idx, value=_cell_value(row[col_name]))

            # Auto-width columns
            for col_idx, col_name in enumerate(columns, 1):
                max_len = len(col_name)
                for row in rows[:100]:
                    val = str(_cell_value(row[col_name]))
                    if len(val) > max_len:
                        max_len = len(val)
                ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else f"A{chr(64 + col_idx - 26)}"].width = min(max_len + 2, 50)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{table}.xlsx"'},
            )

        except ImportError:
            # Fallback to CSV
            import csv

            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(columns)
            for row in rows:
                writer.writerow([_cell_value(row[col]) for col in columns])

            csv_bytes = output.getvalue().encode("utf-8")
            return StreamingResponse(
                io.BytesIO(csv_bytes),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{table}.xlsx"'},
            )
    finally:
        await conn.close()
